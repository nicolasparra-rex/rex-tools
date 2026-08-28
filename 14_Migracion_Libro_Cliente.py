# -*- coding: utf-8 -*-
"""Rex-tools · Migración detalle desde el Libro de Remuneraciones de cualquier cliente.
Lee un libro con estructura arbitraria, autodetecta período/estructura, propone el mapeo de
conceptos contra el catálogo del cliente y genera la planilla de migración detalle, cuadrada
al peso. Empresa/mutual/contrato/caja se resuelven por RUT desde la dotación."""
import streamlit as st
import pandas as pd
import io, os, json, csv, zipfile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from libro_engine import (norm, _num, load_grid, detect_header_row, match_struct, classify_and_map,
                          generar_detalle, cargar_homologacion, cargar_dotacion, detectar_periodo,
                          leer_catalogo_rex, cargar_base_estandar, OUT_COLS)

try:
    from lib.branding import aplicar_branding, aplicar_footer, hero
except Exception:
    def aplicar_branding(**k): pass
    def aplicar_footer(): pass
    def hero(t, d="", i=""): st.title(t); st.caption(d)

DATA_DIR = "data"
AMBAR = "#FFF3CD"; ROJO = "#F8D7DA"; VERDE = "#D4EDDA"
_XLMIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_ETIQ_CLASIF = {"af": "AFP", "is": "Salud", "mu": "Mutual", "ca": "Caja/CCAF",
                "ap": "APV", "ie": "Educacional", "ot": "Otro"}

st.set_page_config(page_title="Rex+ | Conversor a Migración Detalle", page_icon="📘", layout="wide")
aplicar_branding(titulo_pagina="Migración Detalle", badge="BETA")

# ---- CSS propio de esta página: uploaders en español, contraste de captions, íconos, stepper y puntos de color ----
st.markdown("""
<style>
/* Captions con más contraste (AA) */
[data-testid="stCaptionContainer"], [data-testid="stCaptionContainer"] p { color:#5c6773 !important; }

/* Componentes de carga en español */
[data-testid="stFileUploaderDropzoneInstructions"] span{ font-size:0; }
[data-testid="stFileUploaderDropzoneInstructions"] span::after{
    content:"Arrastra el archivo aquí"; font-size:.875rem; color:#31333F; font-weight:500; }
[data-testid="stFileUploaderDropzoneInstructions"] small{ font-size:0; }
[data-testid="stFileUploaderDropzoneInstructions"] small::after{
    content:"Límite 1 GB por archivo"; font-size:.8rem; color:#5c6773; }
[data-testid="stFileUploaderDropzone"] button{ color:transparent !important; position:relative; min-width:8rem; }
[data-testid="stFileUploaderDropzone"] button *{ color:transparent !important; fill:transparent !important; }
[data-testid="stFileUploaderDropzone"] button::after{
    content:"Subir archivo"; color:#fff !important; font-weight:600; font-size:.875rem;
    position:absolute; inset:0; display:flex; align-items:center; justify-content:center; }

/* Fallback de tipografía para los íconos del expander (evita ver 'keyboard_arrow_right' si no carga Material Symbols) */
[data-testid="stExpanderToggleIcon"]{
    font-family:'Material Symbols Rounded','Material Symbols Outlined','Material Symbols Sharp',sans-serif; }

/* Puntos de color (lista 'Antes de empezar') */
.rex-dot{display:inline-block;width:.7rem;height:.7rem;border-radius:50%;margin-right:.4rem;vertical-align:middle;}

/* Stepper de pasos */
.rex-steps{display:flex;align-items:center;gap:0;margin:1.4rem 0 1.3rem;flex-wrap:wrap;}
.rex-step{display:flex;align-items:center;gap:.5rem;}
.rex-step .dot{width:1.7rem;height:1.7rem;border-radius:50%;display:flex;align-items:center;justify-content:center;
    font-weight:800;font-size:.85rem;flex:0 0 auto;background:#e3e8ef;color:#8a93a2;border:2px solid #e3e8ef;}
.rex-step .lbl{font-size:.9rem;font-weight:600;color:#8a93a2;line-height:1.15;white-space:nowrap;}
.rex-bar{flex:1;height:2px;background:#e3e8ef;margin:0 .6rem;min-width:16px;}
.rex-step.done .dot{background:#1e5591;color:#fff;border-color:#1e5591;}
.rex-step.done .lbl{color:#1e5591;}
.rex-step.active .dot{background:#fff;color:#1e5591;border-color:#1e5591;box-shadow:0 0 0 3px rgba(30,85,145,.18);}
.rex-step.active .lbl{color:#1e5591;font-weight:800;}
</style>
""", unsafe_allow_html=True)

hero("📘 Conversor de Libro a Migración Detalle Rex+",
     "Convierte el libro de remuneraciones (o detalle del proceso) del cliente y <b>construye</b> el archivo de "
     "<b>migración detalle de Rex+</b> — con mapeo de conceptos, homologación de instituciones y cuadratura al peso.")
# Aviso legal — estilo informativo (no de error)
st.markdown(
    '<div style="background:#EAF2FB;border:1px solid #cfe0f3;border-left:4px solid #1e5591;color:#22364a;'
    'padding:.7rem 1rem;border-radius:.5rem;line-height:1.4;margin:.55rem 0 1rem;">'
    'ℹ️ Los datos personales de este proceso se tratan <b>exclusivamente para la migración a Rex+</b>, '
    'conforme a la Ley N° 21.719 de protección de datos personales; no se destinan a otros fines '
    'ni a infringir la normativa vigente.</div>',
    unsafe_allow_html=True)

# ---- Stepper: mapa del proceso (se pinta con el paso real a lo largo del script) ----
_PASOS = ["Subir archivos", "Mapear conceptos", "Cuadratura", "Descargar"]
def _stepper_html(activo):
    out = ['<div class="rex-steps">']
    for i, nom in enumerate(_PASOS, start=1):
        cls = "done" if i < activo else ("active" if i == activo else "")
        out.append(f'<div class="rex-step {cls}"><div class="dot">{i}</div><div class="lbl">{nom}</div></div>')
        if i < len(_PASOS):
            out.append('<div class="rex-bar"></div>')
    out.append('</div>')
    return "".join(out)
_stepper_ph = st.empty()
def pintar_stepper(activo):
    _stepper_ph.markdown(_stepper_html(activo), unsafe_allow_html=True)

# ---------- helpers de datos de referencia ----------
# Nota: la caché se invalida por la fecha de modificación del archivo (mtime). Así, si el
# archivo en data/ se agrega o se actualiza, la tabla se recarga sola (no queda un "vacío" pegado).
def _mtime(path):
    return os.path.getmtime(path) if os.path.exists(path) else 0

@st.cache_data(show_spinner=False)
def _parametros_de(path, _mtime):
    wp = load_workbook(path, data_only=True).worksheets[0]
    ph = [wp.cell(row=1, column=c).value for c in range(1, wp.max_column + 1)]
    return {wp.cell(row=r, column=1).value: {ph[c-1]: wp.cell(row=r, column=c).value
            for c in range(1, wp.max_column + 1)} for r in range(2, wp.max_row + 1) if wp.cell(row=r, column=1).value}

def cargar_parametros(file=None):
    if file is not None:
        return _parametros_de(file, getattr(file, "name", "subido"))
    p = os.path.join(DATA_DIR, "parametrosMesuales.xlsx")
    return _parametros_de(p, _mtime(p)) if os.path.exists(p) else {}

@st.cache_data(show_spinner=False)
def cargar_cot_hist():
    ch = load_workbook(os.path.join(DATA_DIR, "cot_afp_hist.xlsx"), data_only=True).worksheets[0]
    return {f"{ch.cell(row=r,column=2).value}{ch.cell(row=r,column=3).value}": (ch.cell(row=r,column=5).value or 0)
            for r in range(2, ch.max_row + 1)}

@st.cache_data(show_spinner=False)
def _homolog_de(path, _mtime):
    return cargar_homologacion(path) if os.path.exists(path) else []

def cargar_homolog_default():
    p = os.path.join(DATA_DIR, "listado_instituciones.xlsx")
    return _homolog_de(p, _mtime(p))

@st.cache_data(show_spinner=False)
def cargar_ids_estandar():
    """Lista maestra de IDs de concepto estándar de Rex (desde data/equiv_conceptos.xlsx).
    Sirve como universo base para el desplegable, además del catálogo del cliente."""
    p = os.path.join(DATA_DIR, "equiv_conceptos.xlsx")
    if not os.path.exists(p):
        return []
    try:
        x = pd.read_excel(p)
        col = "concepto_detalle" if "concepto_detalle" in x.columns else x.columns[1]
        return sorted({str(v).strip() for v in x[col].dropna() if str(v).strip()})
    except Exception:
        return []

def leer_catalogo(file):
    df = pd.read_excel(file, header=None)
    hr = 0
    for r in range(min(6, len(df))):
        vals = [norm(x) for x in df.iloc[r].values]
        if "concepto" in vals and any("nombre" in v for v in vals): hr = r; break
    hdr = [norm(x) for x in df.iloc[hr].values]
    ci = hdr.index("concepto") if "concepto" in hdr else 0
    ni = next((i for i, v in enumerate(hdr) if "nombre" in v), 2)
    names, valid = {}, {}
    for _, row in df.iloc[hr+1:].iterrows():
        cid = row[ci]; nom = row[ni]
        if pd.notna(cid) and str(cid).strip():
            valid[str(cid).strip()] = (str(nom).strip() if pd.notna(nom) else "")
            if pd.notna(nom): names[norm(nom)] = str(cid).strip()
    return names, valid

def to_csv(filas):
    """Archivo de salida en CSV UTF-8 (con BOM, separador coma) — formato de carga a Rex."""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(OUT_COLS)
    for rr in filas:
        w.writerow(rr)
    return buf.getvalue().encode("utf-8-sig")

# ---------- Barra lateral: se deja libre para el menú de las apps ----------
cliente = ""  # los archivos de salida se nombran por período
# valores por defecto (se completan desde la dotación / homologación)
apv_inst = "afp"
caja_inst = "losandes"
jornada = "C"
# fallbacks internos (se completan desde la dotación)
empresa_id, mutual_id, num_contrato = "", "", 1

_DOTACION_SQL = '''SELECT
    contr.empleado                                                          AS "Id empleado",
    contr.contrato                                                          AS "Numero de contrato",
    contr."fechaInic"                                                       AS "fechaInic",
    contr."tipoCont"                                                        AS "tipoCont",
    (SELECT emp.jubilado FROM T$empleados emp WHERE emp.empleado = contr.empleado) AS "jubilado",
    e.identificador_nacional                                                AS "Rut empresa",
    e.empresa                                                               AS "idempresa",
    e.mutual                                                                AS "Mutual",
    e."cotizacionMutu"                                                      AS "% mutual"
FROM T$empleadoscontr contr
INNER JOIN T$empresas e ON e.empresa = contr.empresa'''

# ---------- 0. Antes de empezar (orientación) ----------
st.markdown("### Antes de empezar")
st.markdown("Para migrar la historia de un cliente necesitas preparar **3 archivos** y luego subirlos abajo:")
st.markdown(
    '<div style="line-height:1.85;margin:.1rem 0 .5rem;">'
    '<div><span class="rex-dot" style="background:#1e5591"></span> <b>Libro de remuneraciones</b> del cliente '
    '— te lo entrega el cliente. <b>(Obligatorio)</b></div>'
    '<div><span class="rex-dot" style="background:#F5A623"></span> <b>Catálogo de conceptos</b> del cliente '
    '— se <b>exporta de su Rex</b> (hoja <i>Lista de conceptos</i>). <b>(Obligatorio)</b></div>'
    '<div><span class="rex-dot" style="background:#7CB342"></span> <b>Dotación</b> '
    '— la <b>generas tú</b> corriendo la consulta de abajo y exportando el resultado a Excel. <b>(Recomendada)</b></div>'
    '</div>', unsafe_allow_html=True)
with st.expander("📋 Paso previo — genera la dotación con esta consulta (para el consultor)"):
    st.caption("**Dónde ejecutarla:** corre esta consulta SQL directamente sobre la base de datos del cliente "
               "(desde el gestor/consola SQL que uses para ese cliente). Luego **exporta el resultado a Excel**: "
               "ese archivo es el que subes como ③ Dotación. Con él la app resuelve contrato / empresa / mutual por RUT.")
    st.code(_DOTACION_SQL, language="sql")
st.divider()

# ---------- Tablas de referencia (homologación + parámetros) — visibles antes de subir archivos ----------
def _ultimo_mes_param(P):
    con = sorted(m for m, r in P.items()
                 if _num(r.get("topeImp_pesos_afp", 0)) > 0 and _num(r.get("sis", 0)) > 0)
    return con[-1] if con else None

def render_config_avanzada():
    st.markdown('<div id="config-avanzada"></div>', unsafe_allow_html=True)
    st.markdown("### 📋 Tablas de referencia")
    st.caption("Estas **dos tablas** son las que usa la migración. Ábrelas para revisarlas y, si hace falta, "
               "actualizarlas: la **homologación de instituciones** (traduce AFP/salud/mutual/caja al ID de Rex) "
               "y los **parámetros mensuales** (tope de salud y % de SIS por mes).")

    # --- Homologación de instituciones ---
    _hp = os.path.join(DATA_DIR, "listado_instituciones.xlsx")
    _homolog_disp = cargar_homolog_default()
    with st.expander(f"🏛️ Homologación de instituciones — {len(_homolog_disp)} registros  ·  ver / descargar / actualizar / subir"):
        st.caption("Traduce el texto del libro (‘MODELO’, ‘Colmena…’) al **ID de Rex**, por tipo de institución "
                   "(AFP, salud, mutual, caja, APV…). Se sube desde Rex cuando corresponde.")
        if _homolog_disp:
            _hrows = [{"Tipo": _ETIQ_CLASIF.get(r.get("clasif"), r.get("clasif")),
                       "ID Rex": r.get("id"), "Nombre": str(r.get("nombre_n", "")).title(),
                       "Código": r.get("cod")} for r in _homolog_disp]
            _hdf = pd.DataFrame(_hrows).sort_values(["Tipo", "Nombre"]).reset_index(drop=True)
            st.dataframe(_hdf, hide_index=True, use_container_width=True, height=280)
        else:
            st.warning("⚠️ No se pudo cargar la homologación (`data/listado_instituciones.xlsx`).")
        if os.path.exists(_hp):
            with open(_hp, "rb") as f:
                st.download_button("⬇️ Descargar listado_instituciones.xlsx actual", f.read(),
                                   file_name="listado_instituciones.xlsx", mime=_XLMIME, key="dl_homolog")
        st.file_uploader("Reemplazar listado_instituciones.xlsx (opcional)", type=["xlsx"], key="up_homolog")

    # --- Parámetros mensuales ---
    _pp = os.path.join(DATA_DIR, "parametrosMesuales.xlsx")
    _P = _parametros_de(_pp, _mtime(_pp)) if os.path.exists(_pp) else {}
    _ult_mes = _ultimo_mes_param(_P)
    with st.expander(f"📅 Parámetros mensuales — con datos hasta {_ult_mes or '—'}  ·  ver / descargar / actualizar / subir"):
        if _ult_mes:
            st.info(f"Tope de salud y % SIS cargados hasta **{_ult_mes}**. "
                    "Si migras un mes posterior, descarga la tabla, completa la fila del mes y vuelve a subirla.")
        else:
            st.warning("⚠️ La tabla no tiene meses con datos válidos.")
        if os.path.exists(_pp):
            _cols = {"mes_Proc": "Mes", "topeImp_pesos_afp": "Tope imponible ($)", "topeImp_Uf_afp": "Tope (UF)",
                     "sis": "SIS %", "topeCes_pesos": "Tope cesantía ($)", "uf_Mes": "UF", "monto_Utm": "UTM", "imm": "IMM"}
            _rows = [{_cols.get(k, k): v for k, v in {**{"mes_Proc": m}, **r}.items() if k in _cols}
                     for m, r in sorted(_P.items(), reverse=True) if _num(r.get("topeImp_pesos_afp", 0)) > 0]
            if _rows:
                st.caption("Los que usa el motor son **Tope imponible ($)** y **SIS %**; el resto es de referencia.")
                st.dataframe(pd.DataFrame(_rows), hide_index=True, use_container_width=True, height=260)
            with open(_pp, "rb") as f:
                st.download_button("⬇️ Descargar parametrosMesuales.xlsx actual", f.read(),
                                   file_name="parametrosMesuales.xlsx", mime=_XLMIME, key="dl_param")
        st.file_uploader("Reemplazar parametrosMesuales.xlsx (opcional)", type=["xlsx"], key="up_param")

# Se muestran ANTES de subir archivos: el usuario debe conocer y revisar estas dos tablas.
render_config_avanzada()
st.divider()

# ---------- 1. Archivos ----------
st.markdown("### 1 · Sube los archivos")
# Estilos SOLO de nuestro apartado (clases propias rex-upc-*, no tocan el menú lateral ni nada global).
st.markdown("""
<style>
.rex-upc{margin:-1rem -1rem .7rem -1rem;padding:.6rem 1rem;font-weight:700;color:#fff;
         border-radius:.5rem .5rem 0 0;font-size:.95rem;line-height:1.2;
         display:flex;align-items:center;min-height:2.7rem;box-sizing:border-box;}
.rex-upc .n{opacity:.85;font-weight:800;margin-right:.35rem;}
.rex-upc .rex-badge{margin-left:auto;font-size:.68rem;font-weight:700;background:rgba(255,255,255,.28);
         padding:.12rem .5rem;border-radius:1rem;white-space:nowrap;}
.rex-upc-ambar .rex-badge{background:rgba(0,0,0,.14);}
.rex-upc-azul{background:#1e5591;}
.rex-upc-ambar{background:#F5A623;color:#3d2c00;}
.rex-upc-verde{background:#7CB342;}
</style>
""", unsafe_allow_html=True)
c1, c2, c3 = st.columns(3)
with c1:
    with st.container(border=True):
        st.markdown('<div class="rex-upc rex-upc-azul"><span class="n">①</span>Libro de remuneraciones<span class="rex-badge">Obligatorio</span></div>', unsafe_allow_html=True)
        libro_files = st.file_uploader("Libro", type=["xlsx", "xls"], key="up_libro", label_visibility="collapsed",
                                       accept_multiple_files=True)
        st.caption("Del cliente · .xlsx o .xls · puedes subir **varios meses** (o arrastrar la carpeta)")
with c2:
    with st.container(border=True):
        st.markdown('<div class="rex-upc rex-upc-ambar"><span class="n">②</span>Catálogo de conceptos<span class="rex-badge">Obligatorio</span></div>', unsafe_allow_html=True)
        cat_file = st.file_uploader("Catálogo", type=["xlsx"], key="up_cat", label_visibility="collapsed")
        st.caption("Exportado del Rex del cliente · .xlsx")
with c3:
    with st.container(border=True):
        st.markdown('<div class="rex-upc rex-upc-verde"><span class="n">③</span>Dotación<span class="rex-badge">Recomendada</span></div>', unsafe_allow_html=True)
        dot_file = st.file_uploader("Dotación", type=["xlsx"], key="up_dot", label_visibility="collapsed")
        st.caption("RUT → contrato / empresa / mutual / caja · .xlsx")
        st.caption("¿No tienes la dotación? Genérala con la **consulta del paso previo** (arriba).")

if not libro_files:
    pintar_stepper(1)
    st.info("Sube el **libro** para comenzar (puedes subir **varios meses**). El **catálogo** ayuda a proponer "
            "los IDs y la **dotación** completa empresa/mutual/contrato/caja.")
    aplicar_footer(); st.stop()

by_id, name_to_id = ({}, {})
if cat_file: by_id, name_to_id = leer_catalogo_rex(cat_file)
base_estandar = cargar_base_estandar(os.path.join(DATA_DIR, "conceptos_base.xlsx"))
if not by_id:
    pintar_stepper(1)
    st.error("🔴 Falta el **catálogo de conceptos del cliente** (② en el paso 1). Es **obligatorio**: "
             "de él salen los IDs válidos de Rex. Expórtalo de su Rex (hoja *Lista de conceptos*) y súbelo.")
    aplicar_footer(); st.stop()
saved = {}
# Los reemplazos opcionales de tablas de referencia se suben en ⚙️ Configuración avanzada (al final);
# se leen desde session_state para no depender del orden de render.
_up_homolog = st.session_state.get("up_homolog")
_up_param = st.session_state.get("up_param")
homolog = cargar_homologacion(_up_homolog) if _up_homolog else cargar_homolog_default()
if not homolog:
    st.error("🔴 **No se cargó la tabla de homologación de instituciones.** Las AFP, salud, mutual y caja "
             "saldrán **sin homologar** (texto crudo del libro). Revisa que exista `data/listado_instituciones.xlsx` "
             "o súbela en **📋 Tablas de referencia** (arriba).")
dotacion = cargar_dotacion(dot_file) if dot_file else {}
params_all = cargar_parametros(_up_param) if _up_param else cargar_parametros()

# Carga TODOS los libros una vez (uno por mes).
libros = []
for f in libro_files:
    dfi, sheeti = load_grid(f)
    hri = detect_header_row(dfi)
    hdri = [x if str(x) != "nan" else "" for x in dfi.iloc[hri].values]
    libros.append({"name": f.name, "df": dfi, "sheet": sheeti, "hr": hri,
                   "hdr": hdri, "struct": match_struct(hdri), "periodo": detectar_periodo(dfi, f.name)})
ref = libros[0]
df, sheet, hr, hdr, struct = ref["df"], ref["sheet"], ref["hr"], ref["hdr"], ref["struct"]
# Mapeo por UNIÓN de columnas de TODOS los libros: cada mes puede traer columnas distintas.
# Se clasifica CADA libro con su propia estructura (así el grupo/posición de cada columna es correcto)
# y se fusionan por nombre (primera aparición). Una sola revisión cubre todas; cada mes toma las suyas.
_seen, propuesta = set(), []
for lb in libros:
    for r in classify_and_map(lb["hdr"], lb["struct"], catalog_names=name_to_id, saved=saved, valid_ids=set(by_id)):
        n = norm(r["header"])
        if n not in _seen:
            _seen.add(n); propuesta.append(r)

# ---------- 2. Período(s) + estructura ----------
st.markdown("### 2 · Períodos y estructura detectados")
# si es un solo libro y no se detectó el período, permitir escribirlo
if len(libros) == 1 and not ref["periodo"]:
    ref["periodo"] = st.text_input("📅 No pude detectar el período — escríbelo (AAAA-MM)", value="", placeholder="2026-06")
st.dataframe(pd.DataFrame([{"Archivo": lb["name"], "Período": lb["periodo"] or "— (no detectado)",
                            "Hoja": lb["sheet"], "Encabezado (fila)": lb["hr"] + 1} for lb in libros]),
             hide_index=True, use_container_width=True)
meses = sorted({lb["periodo"] for lb in libros if lb["periodo"]})
periodo_ok = len(meses) > 0
if len(libros) > 1:
    st.caption(f"Se procesarán **{len(meses)} mes(es)** en una pasada. El mapeo de abajo cubre las **columnas de "
               "todos los meses** (unión); cada mes usa solo las suyas, así que pueden tener estructuras distintas.")
_sinper = [lb["name"] for lb in libros if not lb["periodo"]]
if _sinper:
    st.warning("⚠️ No pude detectar el período de: " + ", ".join(f"`{n}`" for n in _sinper)
               + ". Esos archivos **no se generarán**. Renombra el archivo incluyendo el período (AAAA-MM).")
_todos = [lb["periodo"] for lb in libros if lb["periodo"]]
_dups = sorted({p for p in _todos if _todos.count(p) > 1})
if _dups:
    st.warning("⚠️ Hay **períodos repetidos** entre los libros: " + ", ".join(_dups)
               + ". Revisa que no hayas subido el mismo mes dos veces.")

# Estructurales imprescindibles (sin estas no se puede generar/cuadrar).
faltan = [k for k in ["rut", "total_haberes", "total_descuentos", "liquido"] if k not in struct]
if faltan:
    st.error(f"❌ No detecté columnas estructurales en el primer libro: **{', '.join(faltan)}**. Revísalo antes de continuar.")
# La renta imponible es deseable pero no bloquea: si no viene, se deriva (AFP ÷ tasa o SIS ÷ tasa).
if "base_afp" not in struct:
    st.warning("⚠️ El libro **no trae la renta imponible** (base AFP). Se **derivará** del monto de AFP (o SIS) ÷ su tasa "
               "para llenar la columna *Afecto*. La cuadratura no se ve afectada (va por montos), pero **revisa** el "
               "*Afecto* resultante con el consultor.")
if not dotacion:
    st.warning("⚠️ No subiste la **dotación**: empresa, mutual, contrato y caja no se podrán resolver por RUT.")

# Validación proactiva: ¿cada mes tiene parámetros mensuales cargados (tope salud / SIS)?
_sin_param = [m for m in meses
              if not (_num(params_all.get(m, {}).get("topeImp_pesos_afp", 0)) > 0
                      and _num(params_all.get(m, {}).get("sis", 0)) > 0)]
if _sin_param:
    _ult = _ultimo_mes_param(params_all)
    _hasta = f" Hoy hay datos **hasta {_ult}**." if _ult else ""
    st.error("🔴 **Faltan los parámetros mensuales de: " + ", ".join(_sin_param) + "** (tope de salud y/o % de SIS en 0)."
             + _hasta +
             " Si generas así, el **afecto de isapre saldrá sin tope** y la **cotización de SIS en 0**. "
             "Agrega esos meses en [📋 Tablas de referencia](#config-avanzada) (arriba): descarga "
             "`parametrosMesuales.xlsx`, completa las filas y vuelve a subirla.")

# ---------- 3. Mapeo ----------
st.markdown("### 3 · Confirma el mapeo de conceptos")
st.caption("Cada columna del libro se asocia a un concepto **del catálogo del cliente**. "
           "El **nombre y el tipo** (haber/descuento/aporte) salen del catálogo. "
           "En **ámbar** los que faltan por asociar; en **rojo** los que no existen en el catálogo (hay que corregirlos). "
           "Si una columna **no es un concepto** (ej. *Centro de Trabajo*, *Pluriempleo*), marca **Omitir** y no bloqueará.")

valid_ids = sorted(by_id.keys())
# Opciones del desplegable como "ID - Nombre" (para que al asignar se vea qué es cada ID).
_disp_id = {cid: f"{cid} - {by_id[cid]['nombre']}" for cid in valid_ids}
_id_de_disp = {v: k for k, v in _disp_id.items()}   # inverso: "ID - Nombre" -> ID
_opciones_disp = [_disp_id[cid] for cid in valid_ids]
def _id_real(v):
    """Del valor mostrado en el desplegable devuelve el ID (acepta 'ID - Nombre' o el ID pelado)."""
    v = str(v).strip()
    return _id_de_disp.get(v, v)

def _info(cid):
    """Datos derivados del catálogo para un ID: nombre, legal/propio, bloque y estado."""
    if not cid:
        return {"nombre": "", "tipo": "LEGAL/PROPIO", "bloque": "", "estado": "🟠 falta asociar"}
    if cid in by_id:
        return {"nombre": by_id[cid]["nombre"], "tipo": ("LEGAL" if cid in base_estandar else "PROPIO"),
                "bloque": by_id[cid]["bloque"] or "—", "estado": "✅ en catálogo"}
    return {"nombre": "—", "tipo": "—", "bloque": "—", "estado": "❌ no está en catálogo"}

# Bloque por posición del libro (respaldo cuando el catálogo no define bloque o el concepto no está asociado).
_GRUPO_POS = {"haber": "haber", "descuento": "desc", "aporte": "aporte", "?": ""}
filas_map = []
for r in propuesta:
    inf = _info(r["id_rex"])
    # Bloque efectivo: del catálogo si lo define; si no (pendiente o Tipo Dato/Valor Guardado), el de la posición.
    if r["id_rex"] in by_id and by_id[r["id_rex"]]["bloque"]:
        inf["bloque"] = by_id[r["id_rex"]]["bloque"]
    else:
        inf["bloque"] = _GRUPO_POS.get(r["grupo"], "")
    _cid = r["id_rex"]
    filas_map.append({"Columna del libro": r["header"], "Concepto Rex (nombre)": inf["nombre"],
                      "Tipo": inf["tipo"], "Bloque": inf["bloque"],
                      "ID Rex": (_disp_id.get(_cid, _cid) if _cid else ""),
                      "Omitir": False, "Estado": inf["estado"]})
map_df = pd.DataFrame(filas_map)

tot = len(map_df)
auto0 = int((map_df["ID Rex"].astype(str).str.len() > 0).sum())
m1, m2, m3 = st.columns(3)
m1.metric("Conceptos", tot); m2.metric("Asociados", auto0); m3.metric("Por asociar", tot - auto0)

# posición del libro por columna (para detectar conflictos y como respaldo del bloque)
pos_by_header = {norm(r["header"]): _GRUPO_POS.get(r["grupo"], "") for r in propuesta}

editor = st.data_editor(
    map_df, use_container_width=True, hide_index=True, key="mapeo", height=430,
    column_config={
        "Columna del libro": st.column_config.TextColumn(disabled=True),
        "Concepto Rex (nombre)": st.column_config.TextColumn(disabled=True,
                    help="Nombre del concepto en el catálogo del cliente."),
        "Tipo": st.column_config.TextColumn("Legal/Propio", disabled=True, width="small"),
        "Bloque": st.column_config.SelectboxColumn("Bloque", options=["haber", "desc", "aporte"], required=False,
                    width="small",
                    help="Cómo se trata en la migración. Por defecto sale del catálogo (o de la posición del "
                         "libro). Cámbialo solo si en ESTE libro corresponde otro (es un ajuste opcional)."),
        "ID Rex": st.column_config.SelectboxColumn("ID Rex (elige del catálogo)", options=_opciones_disp, required=False,
                    width="large",
                    help="Elige el concepto del catálogo del cliente (ID - Nombre). Escribe para filtrar."),
        "Omitir": st.column_config.CheckboxColumn("Omitir", width="small",
                    help="Marca las columnas que NO son conceptos (ej. Centro de Trabajo, Pluriempleo). "
                         "No se mapean, no bloquean y no entran al archivo."),
        "Estado": st.column_config.TextColumn(disabled=True, width="medium"),
    })

def _omitida(r): return bool(r.get("Omitir", False))
mapping = {norm(r["Columna del libro"]): _id_real(r["ID Rex"])
           for _, r in editor.iterrows() if str(r["ID Rex"]).strip() and not _omitida(r)}
# tipo_map = el bloque que quedó en la tabla (catálogo, posición o el override del implementador)
tipo_map = {}
for _, r in editor.iterrows():
    if _omitida(r): continue
    cid = _id_real(r["ID Rex"]); bl = str(r["Bloque"]).strip()
    if cid and bl in ("haber", "desc", "aporte"): tipo_map[cid] = bl
invalidos = sorted(v for v in set(mapping.values()) if v not in by_id)
# Pendientes: sin ID y NO omitidas (las omitidas no bloquean).
pend_df = editor[(editor["ID Rex"].astype(str).str.strip() == "") & (~editor["Omitir"].astype(bool))][["Columna del libro", "Bloque"]]

# Aviso GENERAL (no atado a ningún concepto puntual): dónde el catálogo y el libro difieren en el bloque.
conflictos = []
for _, r in editor.iterrows():
    cid = _id_real(r["ID Rex"])
    if cid in by_id and by_id[cid]["bloque"]:
        pos = pos_by_header.get(norm(r["Columna del libro"]), "")
        bl = str(r["Bloque"]).strip()
        if pos in ("haber", "desc", "aporte") and by_id[cid]["bloque"] != pos and bl == by_id[cid]["bloque"]:
            conflictos.append((r["Columna del libro"], pos, by_id[cid]["bloque"]))

if len(pend_df):
    st.warning(f"🟠 Faltan **{len(pend_df)}** columna(s) por asociar — elige su ID del catálogo arriba "
               "(o marca **Omitir** si no es un concepto):")
    st.dataframe(pend_df.style.set_properties(**{"background-color": AMBAR}),
                 hide_index=True, use_container_width=True)
elif not invalidos:
    st.success("✅ Todos los conceptos están asociados a un ID del catálogo.")
if invalidos:
    st.error("❌ Estos IDs **no existen en el catálogo del cliente**: " + ", ".join(f"`{i}`" for i in invalidos)
             + ". Rex los rechazaría al importar. Corrige el mapeo (elige uno del catálogo) "
               "o agrega el concepto al catálogo del cliente y vuelve a subirlo.")
if conflictos:
    with st.expander(f"ℹ️ {len(conflictos)} concepto(s) donde el libro y el catálogo difieren en el bloque — conviene revisar"):
        st.caption("No siempre es un error (puede ser un concepto bien clasificado). Se aplica el bloque del "
                   "**catálogo**. Si esto te descuadra al generar, revisa el mapeo (quizá va a otro concepto) "
                   "o ajusta la columna **Bloque** arriba.")
        st.dataframe(pd.DataFrame(conflictos, columns=["Columna del libro", "Bloque en el libro", "Bloque en el catálogo"]),
                     hide_index=True, use_container_width=True)

# ---------- 4. Generar ----------
st.markdown("### 4 · Generar y validar")
listo = periodo_ok and not faltan and not len(pend_df) and not invalidos
if not listo:
    faltantes = []
    if not periodo_ok: faltantes.append("detectar al menos un período")
    if faltan: faltantes.append("completar la estructura")
    if len(pend_df): faltantes.append(f"asociar {len(pend_df)} concepto(s)")
    if invalidos: faltantes.append(f"corregir {len(invalidos)} ID(s) que no están en el catálogo")
    st.info("Para habilitar **Generar**, falta: " + " · ".join(faltantes) + ".")

_btn = f"🚀 Generar migración detalle ({len(meses)} mes(es))" if periodo_ok else "🚀 Generar migración detalle"
if st.button(_btn, type="primary", disabled=not listo, use_container_width=True):
    st.session_state["_generado"] = True
    cot = cargar_cot_hist()
    _gen = [lb for lb in libros if lb["periodo"]]
    resultados = []
    _prog = st.progress(0.0, text="Generando…")
    for i, lb in enumerate(_gen, start=1):
        per = lb["periodo"]
        cfg = dict(empresa_id=empresa_id, mutual_id=mutual_id, apv_inst=apv_inst, caja_inst=caja_inst,
                   num_contrato=int(num_contrato), jornada=jornada, periodo=per)
        filas, res = generar_detalle(lb["df"], lb["hr"], lb["struct"], mapping, params_all.get(per, {}),
                                     cot, cfg, homolog=homolog, dotacion=dotacion, tipo_map=tipo_map)
        resultados.append({"periodo": per, "filas": filas, "res": res})
        _prog.progress(i / len(_gen), text=f"Generando… {per}")
    _prog.empty()

    # --- Resumen por mes ---
    def _desc(r): return r["descuadre_haberes"] + r["descuadre_descuentos"] + r["descuadre_liquido"]
    st.markdown("#### Resumen por mes")
    st.dataframe(pd.DataFrame([{"Período": r["periodo"], "Empleados": r["res"]["empleados"],
                                "Omitidos": r["res"].get("omitidos", 0), "Descuadres": _desc(r["res"]),
                                "Estado": "✅ Cuadra" if _desc(r["res"]) == 0 else "⚠️ Revisar"}
                               for r in resultados]), hide_index=True, use_container_width=True)
    _tot_desc = sum(_desc(r["res"]) for r in resultados)
    if _tot_desc:
        st.error(f"⚠️ Hay **{_tot_desc}** descuadre(s) en total — revisa el resumen por mes antes de cargar a Rex.")

    # --- RUT omitidos (todos los meses) ---
    _omit = [{"Período": r["periodo"], "RUT": lc.get("rut"), "Motivo": lc.get("motivo")}
             for r in resultados for lc in r["res"].get("log_contratos", [])]
    if _omit:
        st.warning(f"🔴 **{len(_omit)} RUT omitidos** (no están en la dotación) en total — se excluyeron del archivo.")
        st.dataframe(pd.DataFrame(_omit).style.set_properties(**{"background-color": ROJO}),
                     hide_index=True, use_container_width=True)

    # --- Instituciones sin homologar (todos los meses) ---
    _sin = [{"Período": r["periodo"], "Tipo": li["tipo"], "Valor en el libro": li["valor_libro"]}
            for r in resultados for li in r["res"].get("log_inst", []) if li["estado"] == "SIN HOMOLOGAR"]
    if _sin:
        with st.expander(f"🏛️ {len(_sin)} institución(es) SIN HOMOLOGAR — revisar", expanded=True):
            st.dataframe(pd.DataFrame(_sin).drop_duplicates(), hide_index=True, use_container_width=True)
            st.warning("🟠 Salieron con el texto crudo del libro. Agrégalas en **📋 Tablas de referencia** "
                       "(arriba) y vuelve a generar.")

    # --- Avisos (unión de flags de todos los meses) ---
    for f in sorted({fl for r in resultados for fl in r["res"]["flags"]}):
        st.warning("🟠 " + f)

    # --- Descargas ---
    if len(resultados) == 1:
        r = resultados[0]
        st.download_button("⬇️ Descargar migración detalle (.csv)", to_csv(r["filas"]),
                           file_name=f"migracion_detalle_{r['periodo']}.csv", mime="text/csv", type="primary")
    else:
        _zbuf = io.BytesIO()
        with zipfile.ZipFile(_zbuf, "w", zipfile.ZIP_DEFLATED) as z:
            for r in resultados:
                z.writestr(f"migracion_detalle_{r['periodo']}.csv", to_csv(r["filas"]))
        st.download_button(f"⬇️ Descargar {len(resultados)} archivos (.zip)", _zbuf.getvalue(),
                           file_name=f"migracion_detalle_{len(resultados)}meses.zip",
                           mime="application/zip", type="primary")
        with st.expander("Descargar por mes"):
            for r in resultados:
                st.download_button(f"⬇️ {r['periodo']}.csv", to_csv(r["filas"]),
                                   file_name=f"migracion_detalle_{r['periodo']}.csv",
                                   mime="text/csv", key=f"dl_{r['periodo']}")

    st.divider()
    st.caption("¿Terminaste? Limpia todo para empezar otra migración desde cero.")
    if st.button("🔄 Refrescar datos", help="Borra los archivos subidos (libro, catálogo, dotación), "
                 "el mapeo y la caché, y reinicia la sesión."):
        for _k in ["up_libro", "up_cat", "up_dot", "up_homolog", "up_param", "mapeo", "nuevo_id_txt", "_generado"]:
            st.session_state.pop(_k, None)
        st.cache_data.clear()
        st.rerun()

# ---- Stepper: pinta el paso real alcanzado en esta corrida ----
_paso = 2
if periodo_ok and not faltan and not len(pend_df) and not invalidos:
    _paso = 3
if st.session_state.get("_generado"):
    _paso = 4
pintar_stepper(_paso)

aplicar_footer()
