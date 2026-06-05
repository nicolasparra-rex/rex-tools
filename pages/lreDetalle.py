"""
Rex+ Tools — IreDetalle
Gestión de Parámetros Previsionales, Instituciones y Liquidaciones.
"""

import io
import os
import re
import json
import sqlite3
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from lib.branding import aplicar_branding, aplicar_footer, hero
    BRANDING = True
except ImportError:
    BRANDING = False

st.set_page_config(page_title="IreDetalle | Rex+ Tools", page_icon="📊", layout="wide")

if BRANDING:
    aplicar_branding(titulo_pagina="IreDetalle", badge="PRODUCCIÓN")
    hero("📊 IreDetalle", "Gestión de parámetros previsionales, instituciones AFP/Salud y liquidaciones.")
else:
    st.title("📊 IreDetalle")
    st.caption("Gestión de parámetros previsionales, instituciones AFP/Salud y liquidaciones.")

# ─────────────────────────────────────────────
# BASE DE DATOS
# ─────────────────────────────────────────────

DB_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "iredetalle.db")

PARAM_COLUMNS = [
    ("mes_proc", "TEXT PRIMARY KEY"), ("uf_mes", "REAL"), ("tope_imp_uf_afp", "REAL"),
    ("tope_imp_pesos_afp", "REAL"), ("tope_ces_uf", "REAL"), ("tope_ces_pesos", "REAL"),
    ("sis", "REAL"), ("factor_sis", "REAL"), ("tope_salud_uf", "REAL"),
    ("tope_salud_pesos", "REAL"), ("imm", "REAL"), ("tope_gratif", "REAL"),
    ("monto_utm", "REAL"), ("ult_dia_mes", "INTEGER"), ("aporte_ccaf", "REAL"),
    ("aporte_fonasa", "REAL"), ("formato_fecha", "TEXT"), ("aporte_afp", "REAL"),
    ("seg_social_exp_vida", "REAL"),
]

AFP_COLUMNS = [
    ("id_afp", "TEXT PRIMARY KEY"), ("clas_afp", "TEXT"), ("nombre_afp", "TEXT"),
    ("rut_afp", "TEXT"), ("codPrev_afp", "TEXT"), ("cot_afp", "REAL"), ("observ_afp", "TEXT"),
]

PARAM_HEADERS = [
    "Mes Proceso", "UF Mes", "Tope Imp. UF AFP", "Tope Imp. $ AFP",
    "Tope Ces. UF", "Tope Ces. $", "SIS (%)", "Factor SIS",
    "Tope Salud UF", "Tope Salud $", "IMM", "Tope Gratif.",
    "UTM", "Últ. Día", "Aporte CCAF", "Aporte FONASA",
    "Fecha", "Aporte AFP", "Seg. Social"
]

AFP_HEADERS = ["ID AFP", "Clasificación", "Nombre AFP", "RUT", "Cód. Previred", "Cotización (%)", "Observaciones"]
SALUD_HEADERS = ["ID Institución", "Clasificación", "Nombre Institución", "RUT", "Equiv. Previred"]
CAJAS_HEADERS = ["ID Inst.", "Clasif.", "Nombre Institución", "Doc. Identidad", "Cod. Equiv.", "Valor", "Valor 2", "Valor 3", "Dato Adicional"]
MUTUALES_HEADERS = CAJAS_HEADERS
APV_HEADERS = ["ID APV", "Clasificación", "Nombre Institución APV", "RUT", "Cod. Previred"]

PREVIRED_SALUD = {
    "00": "Sin Isapre", "01": "Banmedica", "02": "Consalud", "03": "VidaTres",
    "04": "Colmena", "05": "Cruz Blanca", "07": "Fonasa", "10": "Nueva Masvida",
    "11": "Isalud", "12": "Fundacion", "25": "Cruz del Norte", "28": "Esencial"
}

COLS_SALIDA = [
    "Fecha de proceso", "Id empleado", "Número de contrato", "Id del concepto",
    "Monto del concepto", "Afecto", "Id de institución", "Cotización de jubilación",
    "Días de licencias", "Días trabajados", "Fecha de aplicación", "Empresa",
    "Total de rebajas por LLSS", "Rentas no gravadas", "Rebaja por zona extrema", "Jornada",
]


@st.cache_resource
def get_db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    cur = conn.cursor()
    col_defs = ", ".join(f"{c} {t}" for c, t in PARAM_COLUMNS)
    cur.execute(f"CREATE TABLE IF NOT EXISTS parametros ({col_defs})")
    afp_defs = ", ".join(f"{c} {t}" for c, t in AFP_COLUMNS)
    cur.execute(f"CREATE TABLE IF NOT EXISTS inst_afp ({afp_defs})")
    cur.execute("""CREATE TABLE IF NOT EXISTS inst_salud (
        id_inst TEXT PRIMARY KEY, clasif TEXT, nombre_inst TEXT, rut_inst TEXT, equiv_previred TEXT)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS inst_cajas (
        id_inst TEXT PRIMARY KEY, clasif TEXT, nombre_inst TEXT,
        doc_ident TEXT, cod_equiv TEXT, valor TEXT, valor2 TEXT, valor3 TEXT, dato_adic TEXT)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS inst_mutuales (
        id_inst TEXT PRIMARY KEY, clasif TEXT, nombre_inst TEXT,
        doc_ident TEXT, cod_equiv TEXT, valor TEXT, valor2 TEXT, valor3 TEXT, dato_adic TEXT)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS inst_apv (
        id_apv TEXT PRIMARY KEY, clasif_apv TEXT, nombre_apv TEXT, rut_apv TEXT, codprevi_apv TEXT)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS afp_previred (
        codPrev_afp TEXT PRIMARY KEY, nombre_afp TEXT)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS liquidaciones (
        id INTEGER PRIMARY KEY AUTOINCREMENT, fecha_carga TEXT, origen TEXT, archivo TEXT,
        mes_periodo TEXT, rut_trabajador TEXT, nombre TEXT,
        sueldo_base REAL, total_haberes REAL, total_descuentos REAL, liquido_pagar REAL,
        afp_monto REAL, salud_monto REAL, datos_raw TEXT)""")
    conn.commit()
    return conn


def db():
    return get_db()


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def calcular_campos(uf, tope_imp_uf, tope_ces_uf, sis, ccaf, mes_proc):
    tope_imp_pesos = round(uf * tope_imp_uf) if uf and tope_imp_uf else 0
    tope_ces_pesos = round(uf * tope_ces_uf) if uf and tope_ces_uf else 0
    factor_sis = round(sis / 100, 4) if sis else 0
    tope_salud_uf = round(tope_imp_uf * 0.07, 3) if tope_imp_uf else 0
    tope_salud_pesos = round(tope_salud_uf * uf) if tope_salud_uf and uf else 0
    try:
        mm = mes_proc[-2:]
        ult_dia = 31 if mm in ("01", "03", "05", "07", "08", "10", "12") else 30
    except Exception:
        ult_dia = 30
    aporte_fonasa = round(7 - ccaf, 2) if ccaf is not None else 0
    return {
        "tope_imp_pesos_afp": tope_imp_pesos, "tope_ces_pesos": tope_ces_pesos,
        "factor_sis": factor_sis, "tope_salud_uf": tope_salud_uf,
        "tope_salud_pesos": tope_salud_pesos, "ult_dia_mes": ult_dia,
        "aporte_fonasa": aporte_fonasa,
    }


def validar_rut(rut):
    rut = rut.upper().replace(".", "").replace("-", "").strip()
    if len(rut) < 2:
        return False
    cuerpo, dv = rut[:-1], rut[-1]
    if not cuerpo.isdigit():
        return False
    suma, mult = 0, 2
    for c in reversed(cuerpo):
        suma += int(c) * mult
        mult = mult + 1 if mult < 7 else 2
    resto = suma % 11
    dv_calc = str(11 - resto) if resto not in (0, 1) else ("0" if resto == 0 else "K")
    return dv == dv_calc


def fmt_num(val):
    if val is None or val == "":
        return ""
    try:
        f = float(val)
        return f"{int(f):,}" if f == int(f) else f"{f:,.4f}".rstrip("0").rstrip(".")
    except Exception:
        return str(val)


# ─────────────────────────────────────────────
# CSS EXTRA
# ─────────────────────────────────────────────

st.markdown("""
<style>
.metric-card {
    background: white;
    border: 1px solid #e2e8f0;
    border-radius: 10px;
    padding: 16px 20px;
    margin-bottom: 8px;
}
.metric-card .label { color: #64748b; font-size: 12px; font-weight: 500; text-transform: uppercase; letter-spacing: 0.5px; }
.metric-card .value { color: #1e293b; font-size: 22px; font-weight: 700; }
.section-header {
    background: linear-gradient(135deg, #1A3A5F, #2563eb);
    color: white;
    padding: 8px 16px;
    border-radius: 8px;
    font-weight: 600;
    margin: 12px 0 8px 0;
}
.tag-ok { background: #dcfce7; color: #16a34a; padding: 2px 8px; border-radius: 12px; font-size: 12px; font-weight: 600; }
.tag-warn { background: #fef9c3; color: #ca8a04; padding: 2px 8px; border-radius: 12px; font-size: 12px; font-weight: 600; }
.tag-err { background: #fee2e2; color: #dc2626; padding: 2px 8px; border-radius: 12px; font-size: 12px; font-weight: 600; }
div[data-testid="stTabs"] button { font-size: 14px !important; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# INICIALIZAR DB
# ─────────────────────────────────────────────
conn = db()

# ─────────────────────────────────────────────
# TABS PRINCIPALES
# ─────────────────────────────────────────────
tab_param, tab_afp, tab_salud, tab_cajas, tab_mutuales, tab_apv, tab_liq = st.tabs([
    "📅 Parámetros", "🏦 AFP", "🏥 Salud", "🏢 Cajas CCAF",
    "🦺 Mutuales", "💰 APV", "📄 Liquidaciones"
])


# ══════════════════════════════════════════════
# TAB 1: PARÁMETROS MENSUALES
# ══════════════════════════════════════════════

with tab_param:
    st.markdown('<div class="section-header">Parámetros Previsionales Mensuales</div>', unsafe_allow_html=True)

    rows_param = conn.execute("SELECT * FROM parametros ORDER BY mes_proc DESC").fetchall()
    df_param = pd.DataFrame(rows_param, columns=[c[0] for c in PARAM_COLUMNS]) if rows_param else pd.DataFrame(columns=[c[0] for c in PARAM_COLUMNS])

    col_search, col_add = st.columns([3, 1])
    with col_search:
        busq_param = st.text_input("🔍 Buscar mes (ej: 2025-03)", key="busq_param", label_visibility="collapsed", placeholder="Buscar mes (ej: 2025-03)")
    with col_add:
        agregar_param = st.button("➕ Nuevo parámetro", use_container_width=True, key="btn_nuevo_param")

    if busq_param:
        df_param = df_param[df_param["mes_proc"].str.contains(busq_param, na=False)]

    if not df_param.empty:
        df_show = df_param.copy()
        df_show.columns = PARAM_HEADERS
        st.dataframe(df_show, use_container_width=True, hide_index=True, height=280)
        st.caption(f"{len(df_param)} registro(s)")
    else:
        st.info("Sin registros. Agrega el primer parámetro mensual.")

    # ── Formulario nuevo / editar ──
    if agregar_param:
        st.session_state["param_form_open"] = True
        st.session_state["param_edit_mes"] = None

    # Editar desde selectbox
    if not df_param.empty:
        with st.expander("✏️ Editar / eliminar registro existente"):
            mes_sel = st.selectbox("Seleccionar mes", df_param["mes_proc"].tolist(), key="param_sel_edit")
            c1, c2 = st.columns(2)
            if c1.button("✏️ Editar", key="btn_edit_param"):
                st.session_state["param_form_open"] = True
                st.session_state["param_edit_mes"] = mes_sel
            if c2.button("🗑️ Eliminar", key="btn_del_param"):
                conn.execute("DELETE FROM parametros WHERE mes_proc=?", (mes_sel,))
                conn.commit()
                st.success(f"Parámetro '{mes_sel}' eliminado.")
                st.rerun()

    if st.session_state.get("param_form_open"):
        edit_mes = st.session_state.get("param_edit_mes")
        rec = None
        if edit_mes:
            rec = conn.execute("SELECT * FROM parametros WHERE mes_proc=?", (edit_mes,)).fetchone()

        with st.form("form_param"):
            st.subheader("✏️ Editar parámetro" if rec else "➕ Nuevo parámetro")
            c1, c2, c3 = st.columns(3)
            mes_proc    = c1.text_input("Mes Proceso (YYYY-MM)*", value=rec[0] if rec else "", disabled=bool(rec))
            uf_mes      = c2.number_input("UF del Mes*", value=float(rec[1] or 0) if rec else 0.0, format="%.2f")
            tope_imp_uf = c3.number_input("Tope Imponible UF AFP*", value=float(rec[2] or 0) if rec else 0.0, format="%.2f")
            c4, c5, c6 = st.columns(3)
            tope_ces_uf = c4.number_input("Tope Cesantía UF*", value=float(rec[4] or 0) if rec else 0.0, format="%.2f")
            sis         = c5.number_input("SIS (%)*", value=float(rec[6] or 0) if rec else 0.0, format="%.2f")
            imm         = c6.number_input("IMM*", value=int(rec[10] or 0) if rec else 0)
            c7, c8, c9 = st.columns(3)
            tope_gratif = c7.number_input("Tope Gratificación*", value=int(rec[11] or 0) if rec else 0)
            utm         = c8.number_input("UTM*", value=int(rec[12] or 0) if rec else 0)
            ccaf        = c9.number_input("Aporte CCAF (%)*", value=float(rec[14] or 0) if rec else 0.0, max_value=7.0, format="%.2f")
            c10, c11    = st.columns(2)
            fecha_fmt   = c10.text_input("Formato Fecha (dd-mm-aaaa)*", value=str(rec[16] or "") if rec else "")
            aporte_afp  = c11.number_input("Aporte AFP (%)*", value=float(rec[17] or 0) if rec else 0.0, format="%.2f")
            seg_social  = st.number_input("Seg. Social Exp. Vida*", value=float(rec[18] or 0) if rec else 0.0, format="%.2f")

            # Preview calculados
            if uf_mes and tope_imp_uf:
                calc = calcular_campos(uf_mes, tope_imp_uf, tope_ces_uf, sis, ccaf, mes_proc)
                st.markdown("**Campos calculados automáticamente:**")
                pc1, pc2, pc3, pc4 = st.columns(4)
                pc1.metric("Tope Imp. $ AFP", f"{calc['tope_imp_pesos_afp']:,}")
                pc2.metric("Tope Ces. $", f"{calc['tope_ces_pesos']:,}")
                pc3.metric("Tope Salud UF", f"{calc['tope_salud_uf']:.3f}")
                pc4.metric("Aporte FONASA", f"{calc['aporte_fonasa']:.2f}%")

            col_g, col_c = st.columns(2)
            guardar = col_g.form_submit_button("💾 Guardar", use_container_width=True, type="primary")
            cancelar = col_c.form_submit_button("✖ Cancelar", use_container_width=True)

            if cancelar:
                st.session_state["param_form_open"] = False
                st.rerun()

            if guardar:
                errores = []
                if not mes_proc or len(mes_proc) != 7 or mes_proc[4] != "-":
                    errores.append("Mes Proceso inválido (use YYYY-MM)")
                if uf_mes == 0: errores.append("UF del Mes no puede ser 0")
                if errores:
                    for e in errores: st.error(e)
                else:
                    calc = calcular_campos(uf_mes, tope_imp_uf, tope_ces_uf, sis, ccaf, mes_proc)
                    values = [
                        mes_proc, uf_mes, tope_imp_uf, calc["tope_imp_pesos_afp"],
                        tope_ces_uf, calc["tope_ces_pesos"], sis, calc["factor_sis"],
                        calc["tope_salud_uf"], calc["tope_salud_pesos"],
                        imm, tope_gratif, utm, calc["ult_dia_mes"],
                        ccaf, calc["aporte_fonasa"], fecha_fmt, aporte_afp, seg_social
                    ]
                    if rec:
                        cols = [c[0] for c in PARAM_COLUMNS if c[0] != "mes_proc"]
                        set_cl = ", ".join(f"{c}=?" for c in cols)
                        conn.execute(f"UPDATE parametros SET {set_cl} WHERE mes_proc=?", values[1:] + [mes_proc])
                    else:
                        ph = ",".join("?" * len(PARAM_COLUMNS))
                        col_names = ",".join(c[0] for c in PARAM_COLUMNS)
                        conn.execute(f"INSERT OR REPLACE INTO parametros ({col_names}) VALUES ({ph})", values)
                    conn.commit()
                    st.session_state["param_form_open"] = False
                    st.success(f"✅ Parámetro '{mes_proc}' guardado.")
                    st.rerun()

    # ── Exportar Excel ──
    if not df_param.empty:
        if st.button("📥 Exportar parámetros a Excel", key="exp_param"):
            wb = Workbook()
            ws = wb.active
            ws.title = "Parámetros Mensuales"
            ws.append(PARAM_HEADERS)
            for row in df_param.values.tolist():
                ws.append(row)
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            st.download_button("⬇️ Descargar Excel", buf, "parametros_mensuales.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ══════════════════════════════════════════════
# FUNCIÓN GENÉRICA PARA TABLAS DE INSTITUCIONES
# ══════════════════════════════════════════════

def render_institucion_tab(tab_key, tabla, col_names, headers, title, icon,
                           ejemplo_id="", ejemplo_nombre=""):
    with globals()[f"tab_{tab_key}"]:
        st.markdown(f'<div class="section-header">{icon} {title}</div>', unsafe_allow_html=True)

        rows = conn.execute(f"SELECT * FROM {tabla} ORDER BY 3").fetchall()
        df = pd.DataFrame(rows, columns=col_names) if rows else pd.DataFrame(columns=col_names)

        col_s, col_a = st.columns([3, 1])
        with col_s:
            busq = st.text_input("🔍 Buscar", key=f"busq_{tab_key}", label_visibility="collapsed", placeholder="Buscar...")
        with col_a:
            nuevo = st.button(f"➕ Nuevo", use_container_width=True, key=f"btn_nuevo_{tab_key}")

        if busq:
            mask = df.apply(lambda col: col.astype(str).str.contains(busq, case=False, na=False)).any(axis=1)
            df = df[mask]

        if not df.empty:
            df_show = df.copy()
            df_show.columns = headers
            st.dataframe(df_show, use_container_width=True, hide_index=True, height=260)
            st.caption(f"{len(df)} registro(s)")
        else:
            st.info(f"Sin registros en {title}.")

        if nuevo:
            st.session_state[f"{tab_key}_form_open"] = True
            st.session_state[f"{tab_key}_edit_id"] = None

        if not df.empty:
            with st.expander("✏️ Editar / eliminar"):
                id_sel = st.selectbox("Seleccionar", df[col_names[0]].tolist(), key=f"sel_{tab_key}")
                c1, c2 = st.columns(2)
                if c1.button("✏️ Editar", key=f"edit_{tab_key}"):
                    st.session_state[f"{tab_key}_form_open"] = True
                    st.session_state[f"{tab_key}_edit_id"] = id_sel
                if c2.button("🗑️ Eliminar", key=f"del_{tab_key}"):
                    conn.execute(f"DELETE FROM {tabla} WHERE {col_names[0]}=?", (id_sel,))
                    conn.commit()
                    st.success(f"'{id_sel}' eliminado.")
                    st.rerun()

        return df


# ══════════════════════════════════════════════
# TAB 2: AFP
# ══════════════════════════════════════════════

with tab_afp:
    st.markdown('<div class="section-header">🏦 Instituciones AFP</div>', unsafe_allow_html=True)

    rows_afp = conn.execute("SELECT * FROM inst_afp ORDER BY nombre_afp").fetchall()
    df_afp = pd.DataFrame(rows_afp, columns=[c[0] for c in AFP_COLUMNS]) if rows_afp else pd.DataFrame(columns=[c[0] for c in AFP_COLUMNS])

    col_s, col_a = st.columns([3, 1])
    with col_s:
        busq_afp = st.text_input("🔍", key="busq_afp", label_visibility="collapsed", placeholder="Buscar AFP...")
    with col_a:
        nuevo_afp = st.button("➕ Nueva AFP", use_container_width=True, key="btn_nuevo_afp")

    if busq_afp:
        mask = df_afp.apply(lambda c: c.astype(str).str.contains(busq_afp, case=False, na=False)).any(axis=1)
        df_afp = df_afp[mask]

    if not df_afp.empty:
        df_show_afp = df_afp.copy()
        df_show_afp.columns = AFP_HEADERS
        st.dataframe(df_show_afp, use_container_width=True, hide_index=True, height=260)
        st.caption(f"{len(df_afp)} AFP(s)")
    else:
        st.info("Sin AFPs registradas.")

    if nuevo_afp:
        st.session_state["afp_form_open"] = True
        st.session_state["afp_edit_id"] = None

    if not df_afp.empty:
        with st.expander("✏️ Editar / eliminar AFP"):
            id_sel_afp = st.selectbox("Seleccionar AFP", df_afp["id_afp"].tolist(), key="sel_afp")
            c1, c2 = st.columns(2)
            if c1.button("✏️ Editar AFP", key="edit_afp"):
                st.session_state["afp_form_open"] = True
                st.session_state["afp_edit_id"] = id_sel_afp
            if c2.button("🗑️ Eliminar AFP", key="del_afp"):
                conn.execute("DELETE FROM inst_afp WHERE id_afp=?", (id_sel_afp,))
                conn.commit()
                st.success(f"AFP '{id_sel_afp}' eliminada.")
                st.rerun()

    if st.session_state.get("afp_form_open"):
        edit_id = st.session_state.get("afp_edit_id")
        rec = None
        if edit_id:
            rec = conn.execute("SELECT * FROM inst_afp WHERE id_afp=?", (edit_id,)).fetchone()
        with st.form("form_afp"):
            st.subheader("✏️ Editar AFP" if rec else "➕ Nueva AFP")
            c1, c2 = st.columns(2)
            f_id     = c1.text_input("ID AFP*", value=rec[0] if rec else "", disabled=bool(rec))
            f_nombre = c2.text_input("Nombre AFP*", value=rec[2] if rec else "")
            c3, c4 = st.columns(2)
            f_rut    = c3.text_input("RUT AFP", value=rec[3] if rec else "")
            f_cod    = c4.text_input("Cód. Previred*", value=rec[4] if rec else "")
            c5, c6 = st.columns(2)
            f_cot    = c5.number_input("Cotización (%)*", value=float(rec[5] or 0) if rec else 0.0, format="%.2f")
            f_obs    = c6.text_input("Observaciones", value=rec[6] if rec else "")

            # Validar cod previred en tiempo real
            if f_cod:
                previred_rows = conn.execute("SELECT codPrev_afp, nombre_afp FROM afp_previred").fetchall()
                previred_dict = {r[0]: r[1] for r in previred_rows}
                cod_norm = f_cod.strip().zfill(2)
                if cod_norm in previred_dict:
                    st.success(f"✅ Código válido: {cod_norm} → {previred_dict[cod_norm]}")
                elif previred_dict:
                    st.warning(f"⚠️ Código '{cod_norm}' no encontrado en tabla Previred")

            cg, cc = st.columns(2)
            guardar = cg.form_submit_button("💾 Guardar", use_container_width=True, type="primary")
            cancelar = cc.form_submit_button("✖ Cancelar", use_container_width=True)

            if cancelar:
                st.session_state["afp_form_open"] = False
                st.rerun()
            if guardar:
                if not f_id or not f_nombre:
                    st.error("ID y Nombre son obligatorios.")
                else:
                    values = (f_id, "af", f_nombre, f_rut, f_cod.strip().zfill(2) if f_cod else "", f_cot, f_obs)
                    if rec:
                        conn.execute("UPDATE inst_afp SET clas_afp=?,nombre_afp=?,rut_afp=?,codPrev_afp=?,cot_afp=?,observ_afp=? WHERE id_afp=?",
                                     values[1:] + (f_id,))
                    else:
                        conn.execute("INSERT OR REPLACE INTO inst_afp VALUES (?,?,?,?,?,?,?)", values)
                    conn.commit()
                    st.session_state["afp_form_open"] = False
                    st.success(f"✅ AFP '{f_nombre}' guardada.")
                    st.rerun()


# ══════════════════════════════════════════════
# TAB 3: SALUD
# ══════════════════════════════════════════════

with tab_salud:
    st.markdown('<div class="section-header">🏥 Instituciones de Salud</div>', unsafe_allow_html=True)

    rows_sal = conn.execute("SELECT * FROM inst_salud ORDER BY nombre_inst").fetchall()
    df_sal = pd.DataFrame(rows_sal, columns=["id_inst", "clasif", "nombre_inst", "rut_inst", "equiv_previred"]) if rows_sal else pd.DataFrame(columns=["id_inst", "clasif", "nombre_inst", "rut_inst", "equiv_previred"])

    col_s, col_a = st.columns([3, 1])
    with col_s:
        busq_sal = st.text_input("🔍", key="busq_sal", label_visibility="collapsed", placeholder="Buscar institución salud...")
    with col_a:
        nuevo_sal = st.button("➕ Nueva", use_container_width=True, key="btn_nuevo_sal")

    if busq_sal:
        mask = df_sal.apply(lambda c: c.astype(str).str.contains(busq_sal, case=False, na=False)).any(axis=1)
        df_sal = df_sal[mask]

    if not df_sal.empty:
        df_show_sal = df_sal.copy(); df_show_sal.columns = SALUD_HEADERS
        st.dataframe(df_show_sal, use_container_width=True, hide_index=True, height=260)
        st.caption(f"{len(df_sal)} institución(es)")
    else:
        st.info("Sin instituciones de salud.")

    if nuevo_sal:
        st.session_state["sal_form_open"] = True
        st.session_state["sal_edit_id"] = None

    if not df_sal.empty:
        with st.expander("✏️ Editar / eliminar"):
            id_sel_sal = st.selectbox("Seleccionar", df_sal["id_inst"].tolist(), key="sel_sal")
            c1, c2 = st.columns(2)
            if c1.button("✏️ Editar", key="edit_sal"):
                st.session_state["sal_form_open"] = True
                st.session_state["sal_edit_id"] = id_sel_sal
            if c2.button("🗑️ Eliminar", key="del_sal"):
                conn.execute("DELETE FROM inst_salud WHERE id_inst=?", (id_sel_sal,))
                conn.commit(); st.success("Eliminado."); st.rerun()

    if st.session_state.get("sal_form_open"):
        edit_id = st.session_state.get("sal_edit_id")
        rec = conn.execute("SELECT * FROM inst_salud WHERE id_inst=?", (edit_id,)).fetchone() if edit_id else None
        with st.form("form_sal"):
            st.subheader("✏️ Editar" if rec else "➕ Nueva institución salud")
            c1, c2 = st.columns(2)
            f_id     = c1.text_input("ID*", value=rec[0] if rec else "", disabled=bool(rec))
            f_clasif = c2.text_input("Clasificación*", value=rec[1] if rec else "is")
            c3, c4 = st.columns(2)
            f_nombre = c3.text_input("Nombre*", value=rec[2] if rec else "")
            f_rut    = c4.text_input("RUT", value=rec[3] if rec else "")
            f_equiv  = st.text_input("Equiv. Previred*", value=rec[4] if rec else "")

            ref = "  |  ".join(f"{k}={v}" for k, v in sorted(PREVIRED_SALUD.items()))
            st.caption(f"Códigos válidos: {ref}")
            if f_equiv:
                cod_n = f_equiv.strip().zfill(2)
                if cod_n in PREVIRED_SALUD:
                    st.success(f"✅ {cod_n} → {PREVIRED_SALUD[cod_n]}")
                else:
                    st.warning(f"⚠️ Código '{cod_n}' no válido")

            cg, cc = st.columns(2)
            guardar = cg.form_submit_button("💾 Guardar", use_container_width=True, type="primary")
            cancelar = cc.form_submit_button("✖ Cancelar", use_container_width=True)
            if cancelar:
                st.session_state["sal_form_open"] = False; st.rerun()
            if guardar:
                if not f_id or not f_nombre:
                    st.error("ID y Nombre obligatorios.")
                else:
                    vals = (f_id, f_clasif, f_nombre, f_rut, f_equiv.strip().zfill(2) if f_equiv else "")
                    if rec:
                        conn.execute("UPDATE inst_salud SET clasif=?,nombre_inst=?,rut_inst=?,equiv_previred=? WHERE id_inst=?", vals[1:] + (f_id,))
                    else:
                        conn.execute("INSERT OR REPLACE INTO inst_salud VALUES (?,?,?,?,?)", vals)
                    conn.commit()
                    st.session_state["sal_form_open"] = False
                    st.success(f"✅ '{f_nombre}' guardada."); st.rerun()


# ══════════════════════════════════════════════
# FUNCIÓN GENÉRICA CAJAS / MUTUALES / APV
# ══════════════════════════════════════════════

def render_gen_tab(tab_obj, key, tabla, col_names, headers, title):
    with tab_obj:
        st.markdown(f'<div class="section-header">{title}</div>', unsafe_allow_html=True)
        rows = conn.execute(f"SELECT * FROM {tabla} ORDER BY 3").fetchall()
        df = pd.DataFrame(rows, columns=col_names) if rows else pd.DataFrame(columns=col_names)

        col_s, col_a = st.columns([3, 1])
        with col_s:
            busq = st.text_input("🔍", key=f"busq_{key}", label_visibility="collapsed", placeholder="Buscar...")
        with col_a:
            nuevo = st.button("➕ Nueva", use_container_width=True, key=f"btn_nuevo_{key}")

        if busq:
            mask = df.apply(lambda c: c.astype(str).str.contains(busq, case=False, na=False)).any(axis=1)
            df = df[mask]

        if not df.empty:
            df_s = df.copy(); df_s.columns = headers
            st.dataframe(df_s, use_container_width=True, hide_index=True, height=240)
            st.caption(f"{len(df)} registro(s)")
        else:
            st.info(f"Sin registros.")

        if nuevo:
            st.session_state[f"{key}_form_open"] = True
            st.session_state[f"{key}_edit_id"] = None

        if not df.empty:
            with st.expander("✏️ Editar / eliminar"):
                id_s = st.selectbox("Seleccionar", df[col_names[0]].tolist(), key=f"sel_{key}")
                c1, c2 = st.columns(2)
                if c1.button("✏️ Editar", key=f"edit_{key}"):
                    st.session_state[f"{key}_form_open"] = True
                    st.session_state[f"{key}_edit_id"] = id_s
                if c2.button("🗑️ Eliminar", key=f"del_{key}"):
                    conn.execute(f"DELETE FROM {tabla} WHERE {col_names[0]}=?", (id_s,))
                    conn.commit(); st.success("Eliminado."); st.rerun()

        if st.session_state.get(f"{key}_form_open"):
            edit_id = st.session_state.get(f"{key}_edit_id")
            rec = conn.execute(f"SELECT * FROM {tabla} WHERE {col_names[0]}=?", (edit_id,)).fetchone() if edit_id else None
            is_apv = len(col_names) == 5

            with st.form(f"form_{key}"):
                st.subheader("✏️ Editar" if rec else "➕ Nuevo")
                c1, c2 = st.columns(2)
                f_id   = c1.text_input(f"{headers[0]}*", value=rec[0] if rec else "", disabled=bool(rec))
                f_clas = c2.text_input(f"{headers[1]}*", value=rec[1] if rec else "")
                c3, c4 = st.columns(2)
                f_nom  = c3.text_input(f"{headers[2]}*", value=rec[2] if rec else "")
                f_rut  = c4.text_input(f"{headers[3]}", value=rec[3] if rec else "")
                f_cod  = st.text_input(f"{headers[4]}*", value=rec[4] if rec else "")

                if not is_apv:
                    c5, c6, c7 = st.columns(3)
                    f_val  = c5.text_input("Valor", value=rec[5] if rec else "")
                    f_val2 = c6.text_input("Valor 2", value=rec[6] if rec else "")
                    f_val3 = c7.text_input("Valor 3", value=rec[7] if rec else "")
                    f_dato = st.text_input("Dato Adicional", value=rec[8] if rec else "")

                cg, cc = st.columns(2)
                guardar = cg.form_submit_button("💾 Guardar", use_container_width=True, type="primary")
                cancelar = cc.form_submit_button("✖ Cancelar", use_container_width=True)
                if cancelar:
                    st.session_state[f"{key}_form_open"] = False; st.rerun()
                if guardar:
                    if not f_id or not f_nom:
                        st.error("ID y Nombre obligatorios.")
                    else:
                        if is_apv:
                            vals = (f_id, f_clas, f_nom, f_rut, f_cod)
                        else:
                            vals = (f_id, f_clas, f_nom, f_rut, f_cod, f_val, f_val2, f_val3, f_dato)
                        ph = ",".join("?" * len(col_names))
                        if rec:
                            set_cl = ", ".join(f"{c}=?" for c in col_names[1:])
                            conn.execute(f"UPDATE {tabla} SET {set_cl} WHERE {col_names[0]}=?", vals[1:] + (f_id,))
                        else:
                            conn.execute(f"INSERT OR REPLACE INTO {tabla} ({','.join(col_names)}) VALUES ({ph})", vals)
                        conn.commit()
                        st.session_state[f"{key}_form_open"] = False
                        st.success(f"✅ '{f_nom}' guardado."); st.rerun()


render_gen_tab(tab_cajas, "cajas", "inst_cajas",
               ["id_inst","clasif","nombre_inst","doc_ident","cod_equiv","valor","valor2","valor3","dato_adic"],
               CAJAS_HEADERS, "🏢 Cajas de Compensación (CCAF)")

render_gen_tab(tab_mutuales, "mutuales", "inst_mutuales",
               ["id_inst","clasif","nombre_inst","doc_ident","cod_equiv","valor","valor2","valor3","dato_adic"],
               MUTUALES_HEADERS, "🦺 Instituciones Mutuales")

render_gen_tab(tab_apv, "apv", "inst_apv",
               ["id_apv","clasif_apv","nombre_apv","rut_apv","codprevi_apv"],
               APV_HEADERS, "💰 Instituciones APV")


# ══════════════════════════════════════════════
# TAB 7: LIQUIDACIONES
# ══════════════════════════════════════════════

with tab_liq:
    st.markdown('<div class="section-header">📄 Carga y Gestión de Liquidaciones</div>', unsafe_allow_html=True)

    sub1, sub2 = st.tabs(["📂 Cargar archivo", "🗃️ Historial"])

    # ── Sub-tab Cargar ──
    with sub1:
        st.markdown("Carga un archivo Excel con liquidaciones para procesar y exportar al formato estándar de 16 columnas.")

        uploaded = st.file_uploader("Subir archivo Excel de liquidaciones", type=["xlsx", "xls", "xlsm"], key="liq_upload")

        if uploaded:
            try:
                xl = pd.ExcelFile(uploaded)
                hojas = xl.sheet_names
                hoja_sel = st.selectbox("Seleccionar hoja", ["(auto-detectar)"] + hojas, key="liq_hoja")
                hoja = None if hoja_sel == "(auto-detectar)" else hoja_sel

                if st.button("▶ Procesar", type="primary", key="btn_procesar_liq"):
                    with st.spinner("Procesando..."):
                        df_raw = pd.read_excel(uploaded, sheet_name=hoja or 0, header=None)

                        # Detectar fila de encabezados
                        idx_hdr = 0
                        ALIAS_RUT = {"rut", "r.u.t", "rut trabajador", "run", "rut empleado"}
                        ALIAS_PROC = {"proceso", "mes proceso", "periodo", "período", "mes"}
                        for i, row in df_raw.iterrows():
                            vals = [str(v).strip().lower() for v in row if v is not None]
                            if any(v in ALIAS_RUT for v in vals) and any(v in ALIAS_PROC for v in vals):
                                idx_hdr = i
                                break

                        df = pd.read_excel(uploaded, sheet_name=hoja or 0, header=idx_hdr)
                        df.columns = [str(c).strip() for c in df.columns]
                        df = df.dropna(how="all")

                        # Normalizar nombres de columna (sin tildes, minúsculas)
                        def norm_h(s):
                            s = str(s).lower().strip()
                            for a, b in [('á','a'),('é','e'),('í','i'),('ó','o'),('ú','u'),('ñ','n')]:
                                s = s.replace(a, b)
                            return s

                        ALIAS_MAP = {
                            "rut": ["rut","r.u.t","rut trabajador","run","rut empleado"],
                            "proceso": ["proceso","mes proceso","mes","periodo","período"],
                            "empresa": ["empresa","rut empresa","id empresa"],
                            "sueldo_base": ["sueldo base","sueldo","remuneracion base","haber base"],
                            "total_haberes": ["total haberes","total imponible","total remuneraciones"],
                            "total_descuentos": ["total descuentos","total deducciones"],
                            "liquido": ["liquido a pagar","liquido","liquido total","total liquido","neto a pagar","sueldo liquido"],
                            "afp": ["afp","cotizacion afp","descuento afp"],
                            "salud": ["salud","isapre","fonasa","cotizacion salud"],
                            "nombre": ["nombre","nombre trabajador","trabajador","empleado"],
                        }
                        alias_norm = {}
                        for campo, lista in ALIAS_MAP.items():
                            for a in lista:
                                alias_norm[norm_h(a)] = campo

                        col_map = {}
                        for col in df.columns:
                            n = norm_h(col)
                            if n in alias_norm:
                                campo = alias_norm[n]
                                if campo not in col_map:
                                    col_map[campo] = col

                        filas_salida = []
                        for _, row in df.iterrows():
                            rut = str(row.get(col_map.get("rut",""), "")).strip()
                            if not rut or rut in ("nan", ""):
                                continue
                            proc = str(row.get(col_map.get("proceso",""), "")).strip()
                            nom  = str(row.get(col_map.get("nombre",""), "")).strip()

                            # Generar filas para conceptos encontrados
                            CONCEPTOS_MAP = {
                                "sueldo_base": "sueldoBase",
                                "total_haberes": "totalHaberes",
                                "total_descuentos": "totalDescuentos",
                                "liquido": "totalesEmpl",
                                "afp": "afp",
                                "salud": "isapre",
                            }
                            for campo, id_conc in CONCEPTOS_MAP.items():
                                col_name = col_map.get(campo)
                                if not col_name or col_name not in row:
                                    continue
                                monto = row[col_name]
                                try:
                                    monto = int(float(str(monto).replace(".","").replace(",",".")))
                                except Exception:
                                    continue
                                if monto == 0:
                                    continue
                                filas_salida.append({
                                    "Fecha de proceso": proc,
                                    "Id empleado": rut,
                                    "Número de contrato": 1,
                                    "Id del concepto": id_conc,
                                    "Monto del concepto": monto,
                                    "Afecto": 0,
                                    "Id de institución": "",
                                    "Cotización de jubilación": 0,
                                    "Días de licencias": 0,
                                    "Días trabajados": 30,
                                    "Fecha de aplicación": "x",
                                    "Empresa": str(row.get(col_map.get("empresa",""), "")).strip(),
                                    "Total de rebajas por LLSS": 0,
                                    "Rentas no gravadas": 0,
                                    "Rebaja por zona extrema": 0,
                                    "Jornada": "C",
                                    "_nombre": nom,
                                })

                        st.session_state["liq_filas"] = filas_salida
                        st.session_state["liq_archivo"] = uploaded.name

                        if filas_salida:
                            st.success(f"✅ {len(filas_salida)} fila(s) generadas desde {uploaded.name}")
                        else:
                            st.warning("No se encontraron datos mapeables. Verifica los encabezados del Excel.")

            except Exception as e:
                st.error(f"Error procesando el archivo: {e}")

        # ── Previsualización y exportación ──
        if st.session_state.get("liq_filas"):
            filas = st.session_state["liq_filas"]
            df_prev = pd.DataFrame([{k: v for k, v in f.items() if not k.startswith("_")} for f in filas])
            st.dataframe(df_prev, use_container_width=True, hide_index=True, height=280)
            st.caption(f"{len(filas)} fila(s) procesadas")

            col_e, col_g = st.columns(2)

            with col_e:
                if st.button("📥 Exportar a Excel estándar", use_container_width=True, key="exp_liq"):
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Liquidaciones detalladas"
                    thin = Side(style="thin", color="C8D4E0")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    ws.merge_cells(f"A1:{get_column_letter(len(COLS_SALIDA))}1")
                    ws["A1"].value = "Liquidaciones detalladas"
                    ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
                    ws["A1"].fill = PatternFill("solid", start_color="1e3a5f")
                    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
                    ws.row_dimensions[1].height = 26
                    for c, h in enumerate(COLS_SALIDA, 1):
                        cell = ws.cell(row=2, column=c, value=h)
                        cell.font = Font(bold=True, color="FFFFFF", size=10)
                        cell.fill = PatternFill("solid", start_color="2563eb")
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        cell.border = border
                    ws.row_dimensions[2].height = 30
                    ALT = "EFF6FF"
                    for r_idx, fila in enumerate(filas):
                        alt = r_idx % 2 == 0
                        for c_idx, col in enumerate(COLS_SALIDA):
                            val = fila.get(col, "")
                            cell = ws.cell(row=r_idx + 3, column=c_idx + 1, value=val)
                            cell.font = Font(name="Arial", size=10)
                            cell.border = border
                            cell.alignment = Alignment(
                                horizontal="right" if isinstance(val, (int, float)) else "center",
                                vertical="center"
                            )
                            if alt:
                                cell.fill = PatternFill("solid", start_color=ALT)
                    anchos = [12,16,12,18,16,16,16,16,12,12,12,14,18,14,16,10]
                    for i, w in enumerate(anchos, 1):
                        ws.column_dimensions[get_column_letter(i)].width = w
                    ws.freeze_panes = "A3"
                    buf = io.BytesIO()
                    wb.save(buf); buf.seek(0)
                    st.download_button(
                        "⬇️ Descargar Excel estándar", buf,
                        f"liquidaciones_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

            with col_g:
                if st.button("💾 Guardar en historial", use_container_width=True, key="guardar_liq"):
                    archivo = st.session_state.get("liq_archivo", "desconocido")
                    fecha_c = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    # Agrupar por empleado
                    por_emp = {}
                    for f in filas:
                        rut = f["Id empleado"]
                        if rut not in por_emp:
                            por_emp[rut] = {
                                "nombre": f.get("_nombre",""), "proceso": f["Fecha de proceso"],
                                "empresa": f["Empresa"], "sb": 0, "th": 0, "td": 0, "liq": 0, "afp": 0, "sal": 0
                            }
                        id_c = f["Id del concepto"]
                        m = f["Monto del concepto"]
                        if id_c == "sueldoBase": por_emp[rut]["sb"] = m
                        elif id_c == "totalHaberes": por_emp[rut]["th"] = m
                        elif id_c == "totalDescuentos": por_emp[rut]["td"] = m
                        elif id_c == "totalesEmpl": por_emp[rut]["liq"] = m
                        elif id_c == "afp": por_emp[rut]["afp"] = m
                        elif id_c == "isapre": por_emp[rut]["sal"] = m
                    n = 0
                    for rut, d in por_emp.items():
                        conn.execute("""INSERT INTO liquidaciones
                            (fecha_carga,origen,archivo,mes_periodo,rut_trabajador,nombre,
                             sueldo_base,total_haberes,total_descuentos,liquido_pagar,afp_monto,salud_monto,datos_raw)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                            (fecha_c,"EXCEL",archivo,d["proceso"],rut,d["nombre"],
                             d["sb"],d["th"],d["td"],d["liq"],d["afp"],d["sal"],"{}"))
                        n += 1
                    conn.commit()
                    st.success(f"✅ {n} registro(s) guardados en el historial.")
                    st.session_state["liq_filas"] = None

    # ── Sub-tab Historial ──
    with sub2:
        st.markdown("Registros guardados de liquidaciones procesadas.")

        c1, c2, c3 = st.columns(3)
        f_periodo = c1.text_input("Filtrar período", placeholder="ej: 2025-03", key="filt_periodo")
        f_rut     = c2.text_input("Filtrar RUT", placeholder="ej: 12345678-9", key="filt_rut")
        f_nombre  = c3.text_input("Filtrar nombre", placeholder="ej: García", key="filt_nombre")

        sql = "SELECT * FROM liquidaciones WHERE 1=1"
        params = []
        if f_periodo: sql += " AND mes_periodo LIKE ?"; params.append(f"%{f_periodo}%")
        if f_rut:     sql += " AND rut_trabajador LIKE ?"; params.append(f"%{f_rut}%")
        if f_nombre:  sql += " AND nombre LIKE ?"; params.append(f"%{f_nombre}%")
        sql += " ORDER BY id DESC LIMIT 500"

        rows_liq = conn.execute(sql, params).fetchall()
        cols_liq = ["id","fecha_carga","origen","archivo","mes_periodo","rut_trabajador",
                    "nombre","sueldo_base","total_haberes","total_descuentos",
                    "liquido_pagar","afp_monto","salud_monto"]
        df_liq = pd.DataFrame(rows_liq, columns=["id","fecha_carga","origen","archivo",
                               "mes_periodo","rut_trabajador","nombre","sueldo_base",
                               "total_haberes","total_descuentos","liquido_pagar",
                               "afp_monto","salud_monto","datos_raw"])[cols_liq] if rows_liq else pd.DataFrame(columns=cols_liq)

        if not df_liq.empty:
            st.dataframe(df_liq, use_container_width=True, hide_index=True, height=340)
            st.caption(f"{len(df_liq)} registro(s)")

            col_exp, col_del = st.columns(2)

            with col_exp:
                if st.button("📥 Exportar historial a Excel", key="exp_hist"):
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Historial Liquidaciones"
                    ws.append(cols_liq)
                    for row in df_liq.values.tolist():
                        ws.append(row)
                    buf = io.BytesIO()
                    wb.save(buf); buf.seek(0)
                    st.download_button("⬇️ Descargar", buf,
                        f"historial_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            with col_del:
                id_del = st.number_input("ID a eliminar", min_value=1, step=1, key="id_del_liq")
                if st.button("🗑️ Eliminar registro", key="btn_del_liq"):
                    conn.execute("DELETE FROM liquidaciones WHERE id=?", (int(id_del),))
                    conn.commit()
                    st.success(f"Registro {id_del} eliminado.")
                    st.rerun()
        else:
            st.info("Sin registros en el historial.")

if BRANDING:
    aplicar_footer()
