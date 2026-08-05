"""
Rex+ Tools — Minutas de Implementación
Formulario para generar minutas de Remuneraciones, Asistencia y DO en Excel.
Autocompletado desde Zoho Projects al ingresar la OT.
"""

import io
import json
import requests
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    from lib.branding import aplicar_branding, aplicar_footer, hero
    BRANDING = True
except ImportError:
    BRANDING = False

st.set_page_config(page_title="Minutas | Rex+ Tools", page_icon="📋", layout="wide")

if BRANDING:
    aplicar_branding(titulo_pagina="Minutas", badge="PRODUCCIÓN")
    hero("📋 Minutas de Implementación", "Completa los datos y descarga la minuta en Excel lista para entregar.")
else:
    st.title("📋 Minutas de Implementación")
    st.caption("Completa los datos y descarga la minuta en Excel lista para entregar.")

# ── ZOHO HELPERS ──────────────────────────────────────────────────────────────

@st.cache_data(ttl=3000, show_spinner=False)
def get_access_token(refresh_token, client_id, client_secret):
    r = requests.post("https://accounts.zoho.com/oauth/v2/token", params={
        "refresh_token": refresh_token,
        "client_id":     client_id,
        "client_secret": client_secret,
        "grant_type":    "refresh_token",
    })
    return r.json().get("access_token")

@st.cache_data(ttl=600, show_spinner=False)
def buscar_proyecto_por_ot(access_token, portal_id, ot):
    url = f"https://projectsapi.zoho.com/restapi/portal/{portal_id}/projects/"
    headers = {"Authorization": f"Zoho-oauthtoken {access_token}"}
    index = 1
    while True:
        r = requests.get(url, headers=headers, params={"range": 100, "index": index})
        batch = r.json().get("projects", [])
        if not batch:
            break
        for p in batch:
            ot_upper = ot.strip().upper()
            if ot_upper == p.get("key", "").upper() or ot_upper in p.get("name", "").upper():
                return p
        if len(batch) < 100:
            break
        index += 100
    return None

@st.cache_data(ttl=600, show_spinner=False)
def listar_ots_activas(access_token, portal_id):
    ESTADOS = ["inicio sin agenda", "reunion ko", "reunión ko", "agenda por confirmar"]
    url = f"https://projectsapi.zoho.com/restapi/portal/{portal_id}/projects/"
    headers = {"Authorization": f"Zoho-oauthtoken {access_token}"}
    rows = []
    index = 1
    while True:
        r = requests.get(url, headers=headers, params={"range": 100, "index": index})
        batch = r.json().get("projects", [])
        if not batch:
            break
        for p in batch:
            status = p.get("custom_status_name", "")
            if status.lower() in ESTADOS:
                cfields = parse_custom_fields(p.get("custom_fields", []))
                rows.append({
                    "OT":        p.get("key", ""),
                    "Proyecto":  p.get("name", ""),
                    "Consultor": cf(cfields, "Consultor 1"),
                    "Estado":    status,
                })
        if len(batch) < 100:
            break
        index += 100
    return rows

def parse_custom_fields(custom_fields):
    result = {}
    if not isinstance(custom_fields, list):
        return result
    for item in custom_fields:
        if isinstance(item, dict):
            for k, v in item.items():
                result[k] = v
    return result

def cf(fields, *keys):
    for k in keys:
        if k in fields and fields[k] not in (None, "", "false", False):
            val = fields[k]
            if isinstance(val, str) and val.startswith("["):
                try:
                    parsed = json.loads(val)
                    return ", ".join(parsed) if isinstance(parsed, list) else val
                except Exception:
                    pass
            return str(val)
    return ""

def extraer_datos_zoho(proyecto):
    if not proyecto:
        return {}
    cfields = parse_custom_fields(proyecto.get("custom_fields", []))
    return {
        "empresa":       cf(cfields, "Razón social"),
        "rut":           cf(cfields, "RUT Empresa"),
        "vendedor":      cf(cfields, "Vendedor"),
        "jefe_proyecto": proyecto.get("owner_name", ""),
        "direccion":     cf(cfields, "Dirección"),
        "correo":        cf(cfields, "Correo del contacto"),
        "telefono":      cf(cfields, "Telefono de contacto"),
        "plan":          cf(cfields, "Plan Contratado"),
        "colaboradores": cf(cfields, "Cantidad de empleados"),
        "razones":       cf(cfields, "Cantidad de empresas"),
        "empresa_venta": cf(cfields, "Empresa Venta"),
        "contacto":      cf(cfields, "Jefe de Proyecto Cliente (Contacto)"),
    }

# ── EXCEL HELPERS ─────────────────────────────────────────────────────────────

COLOR_TITLE    = "1B3A6B"
COLOR_SECTION  = "2E6DB4"
COLOR_ROW_ODD  = "D9E8F5"
COLOR_ROW_EVEN = "EEF5FB"
COLOR_CHECK1   = "F0F7E6"
COLOR_CHECK2   = "E8F5E9"
WHITE          = "FFFFFF"
FONT_WHITE     = "FFFFFF"
FONT_DARK      = "1A3A5F"
SI_NO          = ["si", "no", "No aplica"]
SI_NO2         = ["si", "no"]

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=FONT_DARK, size=11):
    return Font(name="Arial", bold=bold, color=color, size=size)

def _border():
    thin = Side(style="thin", color="BBCDE0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _set_title(ws, row, text):
    cell = ws.cell(row=row, column=3, value=text)
    cell.fill = _fill(COLOR_TITLE)
    cell.font = Font(name="Arial", bold=True, color=FONT_WHITE, size=13)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
    ws.row_dimensions[row].height = 28

def _set_section(ws, row, text):
    cell = ws.cell(row=row, column=3, value=text)
    cell.fill = _fill(COLOR_SECTION)
    cell.font = Font(name="Arial", bold=True, color=FONT_WHITE, size=11)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
    ws.row_dimensions[row].height = 22

def _set_row(ws, row, label, value, even=False):
    bg_label = COLOR_ROW_EVEN if even else COLOR_ROW_ODD
    bg_value = COLOR_ROW_EVEN if even else WHITE
    lc = ws.cell(row=row, column=3, value=label)
    lc.fill = _fill(bg_label); lc.font = _font()
    lc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    lc.border = _border()
    vc = ws.cell(row=row, column=4, value=value)
    vc.fill = _fill(bg_value); vc.font = _font(color="1A3A5F")
    vc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    vc.border = _border()
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
    ws.row_dimensions[row].height = 20

def _set_check(ws, row, text, alt=False):
    bg = COLOR_CHECK2 if alt else COLOR_CHECK1
    c1 = ws.cell(row=row, column=3, value="☐ Pendiente")
    c1.fill = _fill(bg); c1.font = Font(name="Arial", bold=True, color="4CAF50", size=11)
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c1.border = _border()
    c2 = ws.cell(row=row, column=4, value=text)
    c2.fill = _fill(bg); c2.font = _font()
    c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c2.border = _border()
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
    ws.row_dimensions[row].height = 20

def _set_note(ws, row):
    c = ws.cell(row=row, column=3,
                value="💡  Los campos en azul son editables. Use los desplegables para seleccionar valores estándar.")
    c.font = Font(name="Arial", italic=True, color="555555", size=9)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
    ws.row_dimensions[row].height = 18

def _col_widths(ws):
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 3
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 5

def build_remuneraciones(ws, d):
    _col_widths(ws)
    _set_title(ws, 2, "📋  MINUTA DE IMPLEMENTACIÓN — REX REMUNERACIONES")
    _set_section(ws, 4, "  DATOS GENERALES DEL CLIENTE")
    fields = [
        ("OT (Orden de Trabajo)", d["ot"]),
        ("Empresa / Razón Social", d["empresa"]),
        ("RUT Empresa", d["rut"]),
        ("Vendedor", d["vendedor"]),
        ("Jefe de Proyecto", d["jefe_proyecto"]),
        ("Dirección", d["direccion"]),
        ("Correo de Contacto", d["correo"]),
        ("Número de Contacto", d["telefono"]),
        ("Plan Contratado", d["plan"]),
        ("Cantidad de Colaboradores", d["colaboradores"]),
        ("Cantidad de Razones Sociales", d["razones_sociales"]),
    ]
    for i, (l, v) in enumerate(fields):
        _set_row(ws, 5+i, l, v, even=(i%2==1))
    _set_section(ws, 16, "  CONFIGURACIÓN DE REMUNERACIONES")
    cfg = [
        ("Estructura de Remuneraciones", d["estructura"]),
        ("Comisión / Semana Corrida", d["comision"]),
        ("Reliquidación / Renta Accesoria", d["reliquidacion"]),
        ("3 Primeros Días (Art. 195)", d["tres_dias"]),
        ("Zona Extrema", d["zona_extrema"]),
        ("Provisión Vacaciones", d["provision"]),
        ("Centralización Contable", d["centralizacion"]),
        ("Transferencia Bancaria", d["transferencia"]),
        ("¿Utiliza API?", d["usa_api"]),
    ]
    for i, (l, v) in enumerate(cfg):
        _set_row(ws, 17+i, l, v, even=(i%2==1))
    _set_section(ws, 26, "  COMENTARIOS GENERALES")
    _set_row(ws, 27, "Minuta / Observaciones", d["observaciones"])
    ws.row_dimensions[27].height = 40
    _set_section(ws, 30, "  ✅  CHECKLIST — INFORMACIÓN NECESARIA PARA COMENZAR")
    checklist = [
        "Liquidaciones de todo 2026",
        "Libro de remuneraciones (en Excel) o el que sube a la DT",
        "Contratos y finiquitos (en Word)",
        "Saldo y vacaciones (en Excel)",
        "Licencias médicas (Registros)",
        "Ausentismos (en Excel)",
    ]
    for i, item in enumerate(checklist):
        _set_check(ws, 31+i, item, alt=(i%2==1))
    _set_note(ws, 38)

def build_asistencia(ws, d):
    _col_widths(ws)
    _set_title(ws, 2, "📋  MINUTA DE IMPLEMENTACIÓN — REX ASISTENCIA")
    _set_section(ws, 4, "  DATOS DE LA EMPRESA")
    emp = [
        ("Empresa", d["empresa"]),
        ("RUT", d["rut"]),
        ("Jefe de Proyecto", d["jefe_proyecto"]),
        ("Dirección", d["direccion"]),
        ("Vendedor", d["vendedor"]),
        ("Empresa Venta", d["empresa_venta"]),
        ("Contacto (Nombre Completo)", d["contacto_nombre"]),
        ("Contacto (Número)", d["contacto_numero"]),
        ("Contacto (Email)", d["contacto_email"]),
    ]
    for i, (l, v) in enumerate(emp):
        _set_row(ws, 5+i, l, v, even=(i%2==1))
    _set_section(ws, 14, "  PLAN DE IMPLEMENTACIÓN")
    _set_row(ws, 15, "Plan Asistencia", d["plan_asistencia"])
    _set_row(ws, 16, "Plan Casino", d["plan_casino"], even=True)
    _set_row(ws, 17, "Adicionales", d["adicionales_plan"])
    _set_section(ws, 20, "  CONSULTAS TÉCNICAS")
    tec = [
        ("Sistema de Asistencia Actual", d["sistema_actual"]),
        ("Dispositivo de Marcaje", d["dispositivo"]),
        ("¿Tiene Rex+?", d["tiene_rex"]),
        ("Cantidad de RUT", d["cantidad_rut"]),
        ("Cantidad de Empleados", d["cantidad_empleados"]),
        ("Empleados Art. 22", d["art22"]),
        ("Tipos de Horario", d["tipos_horario"]),
        ("Cantidad de Ubicaciones", d["ubicaciones"]),
    ]
    for i, (l, v) in enumerate(tec):
        _set_row(ws, 21+i, l, v, even=(i%2==1))
    _set_section(ws, 30, "  CONFIGURACIÓN ADICIONAL")
    _set_row(ws, 31, "Concepto de Asistencia (Remuneración)", d["concepto"])
    _set_row(ws, 32, "Cortes Mensuales", d["cortes"], even=True)
    _set_row(ws, 33, "Adicionales / Observaciones", d["observaciones"])
    _set_note(ws, 36)

# ── DO: hojas genéricas ───────────────────────────────────────────────────────

CHECK_SEL = [
    "Perfiles de cargo levantados (descriptor por cargo)",
    "Flujo y responsables de aprobación de requerimientos",
    "Listado de reclutadores con nombre y correo",
    "Plantillas de correo al postulante (avance / rechazo / oferta)",
    "Credenciales de los portales de publicación",
]
CHECK_DES = [
    "Organigrama vigente cargado en Rex+",
    "Diccionario de competencias a evaluar",
    "Escala de evaluación y ponderaciones definidas",
    "Listado de relación evaluador – evaluado",
    "Calendario del ciclo (apertura, cierre, calibración, feedback)",
]
CHECK_CLI = [
    "Cuestionario aprobado por el cliente",
    "Universo de encuestados con correo vigente",
    "Estructura de segmentación de resultados",
    "Comunicación interna de lanzamiento",
    "Fechas de apertura y cierre confirmadas",
]

def build_do(ws, titulo, datos_cliente, nombre_seccion, campos_cfg, observaciones, checklist):
    """Construye una hoja DO con layout estándar (cliente + configuración + checklist)."""
    _col_widths(ws)
    r = 2
    _set_title(ws, r, titulo)
    r += 2
    _set_section(ws, r, "  DATOS GENERALES DEL CLIENTE")
    r += 1
    for i, (l, v) in enumerate(datos_cliente):
        _set_row(ws, r, l, v, even=(i % 2 == 1))
        r += 1
    r += 1
    _set_section(ws, r, nombre_seccion)
    r += 1
    for i, (l, v) in enumerate(campos_cfg):
        _set_row(ws, r, l, v, even=(i % 2 == 1))
        r += 1
    r += 1
    _set_section(ws, r, "  COMENTARIOS GENERALES")
    r += 1
    _set_row(ws, r, "Minuta / Observaciones", observaciones)
    ws.row_dimensions[r].height = 40
    r += 2
    _set_section(ws, r, "  ✅  CHECKLIST — INFORMACIÓN NECESARIA PARA COMENZAR")
    r += 1
    for i, item in enumerate(checklist):
        _set_check(ws, r, item, alt=(i % 2 == 1))
        r += 1
    r += 1
    _set_note(ws, r)

def generar_excel(data_rem, data_asi, data_do=None, submodulos=None,
                  incluir_rem=True, incluir_asi=True):
    wb = Workbook()
    primera = True

    def _hoja(nombre):
        nonlocal primera
        if primera:
            ws = wb.active
            ws.title = nombre
            primera = False
            return ws
        return wb.create_sheet(nombre)

    if incluir_rem:
        build_remuneraciones(_hoja("Rex - Remuneraciones"), data_rem)
    if incluir_asi:
        build_asistencia(_hoja("Rex - Asistencia"), data_asi)

    submodulos = submodulos or []
    if data_do and submodulos:
        d = data_do
        cliente = [
            ("OT (Orden de Trabajo)", d["ot"]),
            ("Empresa / Razón Social", d["empresa"]),
            ("RUT Empresa", d["rut"]),
            ("Vendedor", d["vendedor"]),
            ("Jefe de Proyecto", d["jefe_proyecto"]),
            ("Contacto Cliente", d["contacto_nombre"]),
            ("Correo de Contacto", d["correo"]),
            ("Número de Contacto", d["telefono"]),
            ("Plan Contratado", d["plan"]),
            ("Cantidad de Colaboradores", d["colaboradores"]),
            ("Submódulos DO contratados", ", ".join(submodulos)),
        ]

        if "Selección" in submodulos:
            campos = [
                ("Procesos de selección al mes (estimado)", d["sel_procesos"]),
                ("Cargos / perfiles a levantar", d["sel_cargos"]),
                ("Usuarios reclutadores", d["sel_usuarios"]),
                ("Flujo de aprobación del requerimiento", d["sel_aprobacion"]),
                ("Etapas del proceso", d["sel_etapas"]),
                ("Portales de publicación", d["sel_portales"]),
                ("Pruebas / evaluaciones aplicadas", d["sel_pruebas"]),
                ("¿Integra con onboarding / ficha de empleado?", d["sel_onboarding"]),
                ("¿Requiere firma electrónica de contratos?", d["sel_firma"]),
                ("¿Migra postulantes históricos?", d["sel_migracion"]),
            ]
            build_do(_hoja("DO - Selección"),
                     "📋  MINUTA DE IMPLEMENTACIÓN — DO SELECCIÓN",
                     cliente, "  CONFIGURACIÓN DE SELECCIÓN", campos,
                     d["sel_obs"], CHECK_SEL)

        if "Desempeño" in submodulos:
            campos = [
                ("Tipo de evaluación", d["des_tipo"]),
                ("Periodicidad del ciclo", d["des_periodicidad"]),
                ("Colaboradores a evaluar", d["des_evaluados"]),
                ("Cantidad de evaluadores", d["des_evaluadores"]),
                ("¿Evalúa competencias?", d["des_competencias"]),
                ("Cantidad de competencias", d["des_n_competencias"]),
                ("¿Evalúa objetivos / KPI?", d["des_objetivos"]),
                ("Cantidad de objetivos por colaborador", d["des_n_objetivos"]),
                ("Escala de evaluación", d["des_escala"]),
                ("¿Ponderación distinta por cargo o área?", d["des_ponderacion"]),
                ("¿Requiere calibración / comité?", d["des_calibracion"]),
                ("¿Genera plan de desarrollo / feedback?", d["des_plan"]),
                ("¿Vinculado a bono o ajuste de renta?", d["des_bono"]),
                ("Organigrama cargado en Rex+", d["des_organigrama"]),
            ]
            build_do(_hoja("DO - Desempeño"),
                     "📋  MINUTA DE IMPLEMENTACIÓN — DO DESEMPEÑO",
                     cliente, "  CONFIGURACIÓN DE DESEMPEÑO", campos,
                     d["des_obs"], CHECK_DES)

        if "Clima" in submodulos:
            campos = [
                ("Tipo de encuesta", d["cli_tipo"]),
                ("Periodicidad de aplicación", d["cli_periodicidad"]),
                ("Universo de encuestados", d["cli_universo"]),
                ("¿Encuesta anónima?", d["cli_anonima"]),
                ("Cantidad de dimensiones / secciones", d["cli_dimensiones"]),
                ("Cantidad de preguntas", d["cli_preguntas"]),
                ("Tipos de pregunta", d["cli_tipos_preg"]),
                ("Canales de difusión", d["cli_canales"]),
                ("Segmentación de resultados", d["cli_segmentacion"]),
                ("Fecha estimada de lanzamiento", d["cli_fecha"]),
                ("¿Requiere informe de resultados Rex+?", d["cli_informe"]),
                ("¿Cuestionario propio del cliente?", d["cli_propio"]),
            ]
            build_do(_hoja("DO - Clima"),
                     "📋  MINUTA DE IMPLEMENTACIÓN — DO CLIMA",
                     cliente, "  CONFIGURACIÓN DE CLIMA", campos,
                     d["cli_obs"], CHECK_CLI)

    if primera:  # ninguna hoja seleccionada
        ws = wb.active
        ws.title = "Sin datos"
        ws["C2"] = "No se seleccionó ninguna hoja para exportar."

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ── ZOHO: token ───────────────────────────────────────────────────────────────

VENDEDORES = ['Alicia Jensen', 'Camila Huber', 'Cristian Astaburuaga', 'Edgardo Verdejo',
              'Francisca Soto', 'Francisco Reig', 'Gislaine Sepulveda', 'Gonzalo Pereira',
              'Jenny Chavarro', 'Juan Carlos Rabi', 'Marcelo Baeza', 'Matías Ossandon',
              'Mauricio Bastías', 'Roberto Ramírez', 'Sebastian Ulloa', 'Tamara Castro',
              'Valentina Berrios', 'Yanin Rebolledo', 'Otro', 'Sin Definir']

portal_id = st.secrets.get("ZOHO_PORTAL_ID", "757079135")

try:
    token = get_access_token(
        st.secrets["ZOHO_REFRESH_TOKEN"],
        st.secrets["ZOHO_CLIENT_ID"],
        st.secrets["ZOHO_CLIENT_SECRET"],
    )
    ZOHO_OK = bool(token)
except Exception:
    token = None
    ZOHO_OK = False

# ── SESSION STATE ─────────────────────────────────────────────────────────────
for k, v in {"zoho_data": {}, "last_ot": "", "zoho_msg": ()}.items():
    if k not in st.session_state:
        st.session_state[k] = v

def z(key, default=""):
    return st.session_state.zoho_data.get(key, default)

def panel_ots(key_busq, key_tabla):
    """Panel ayuda memoria OTs activas — solo lectura."""
    if not ZOHO_OK:
        return
    with st.expander("📋 Ver OTs activas (Inicio sin agenda / Reunión KO / Agenda por confirmar)", expanded=False):
        with st.spinner("Cargando OTs..."):
            ots = listar_ots_activas(token, portal_id)
        if ots:
            busq = st.text_input("🔍 Filtrar", placeholder="Buscar OT, nombre o consultor...", key=key_busq)
            df_ots = pd.DataFrame(ots)
            if busq:
                mask = df_ots.apply(lambda row: row.astype(str).str.contains(busq, case=False).any(), axis=1)
                df_ots = df_ots[mask].reset_index(drop=True)
            st.dataframe(df_ots, use_container_width=True, hide_index=True,
                         column_config={"OT": st.column_config.TextColumn(width="small"),
                                        "Proyecto": st.column_config.TextColumn(width="large")},
                         key=key_tabla)
            st.caption(f"{len(df_ots)} proyectos · copia la OT y pégala en el campo de abajo")
        else:
            st.info("No hay proyectos en estos estados.")

# ── UI ────────────────────────────────────────────────────────────────────────

tab_rem, tab_asi, tab_do = st.tabs(["📊 Remuneraciones", "🕐 Asistencia", "🌱 DO"])

with tab_rem:
    st.subheader("Datos Generales del Cliente")

    panel_ots("rem_busq_ot", "rem_tabla_ots")

    c1, c2 = st.columns(2)

    ot = c1.text_input("OT (Orden de Trabajo)", placeholder="Ej: RE-2910", key="r_ot")

    if ot and ot != st.session_state.last_ot and ZOHO_OK:
        with st.spinner(f"🔍 Buscando OT {ot} en Zoho..."):
            proyecto = buscar_proyecto_por_ot(token, portal_id, ot)
        if proyecto:
            datos = extraer_datos_zoho(proyecto)
            st.session_state.zoho_data  = datos
            st.session_state.last_ot    = ot
            st.session_state["r_empresa"]       = datos.get("empresa", "")
            st.session_state["r_rut"]           = datos.get("rut", "")
            st.session_state["r_jefe_proyecto"] = datos.get("jefe_proyecto", "")
            st.session_state["r_direccion"]     = datos.get("direccion", "")
            st.session_state["r_correo"]        = datos.get("correo", "")
            st.session_state["r_telefono"]      = datos.get("telefono", "")
            st.session_state["a_empresa"]       = datos.get("empresa", "")
            st.session_state["a_rut"]           = datos.get("rut", "")
            st.session_state["a_jefe_proyecto"] = datos.get("jefe_proyecto", "")
            st.session_state["a_direccion"]     = datos.get("direccion", "")
            st.session_state["a_cont_email"]    = datos.get("correo", "")
            st.session_state["a_cont_num"]      = datos.get("telefono", "")
            st.session_state["a_cont_nombre"]   = datos.get("contacto", "")
            st.session_state["d_empresa"]       = datos.get("empresa", "")
            st.session_state["d_rut"]           = datos.get("rut", "")
            st.session_state["d_jefe_proyecto"] = datos.get("jefe_proyecto", "")
            st.session_state["d_correo"]        = datos.get("correo", "")
            st.session_state["d_telefono"]      = datos.get("telefono", "")
            st.session_state["d_cont_nombre"]   = datos.get("contacto", "")
            st.session_state["zoho_msg"]        = ("ok", f"✅ Proyecto encontrado: **{proyecto.get('name', '')}**")
            st.rerun()
        else:
            st.session_state.zoho_data  = {}
            st.session_state.last_ot    = ot
            st.session_state["zoho_msg"] = ("warn", f"⚠️ No se encontró proyecto con OT **{ot}**.")
            st.rerun()

    if st.session_state.zoho_msg:
        tipo, msg = st.session_state.zoho_msg
        if tipo == "ok":
            st.success(msg)
        else:
            st.warning(msg)

    empresa_r       = c1.text_input("Empresa / Razón Social", placeholder="Ej: Fundación Ejemplo", key="r_empresa")
    rut_r           = c1.text_input("RUT Empresa", placeholder="Ej: 65058734-0", key="r_rut")
    vendedor_z      = z("vendedor")
    v_idx           = VENDEDORES.index(vendedor_z) if vendedor_z in VENDEDORES else 0
    vendedor_r      = c1.selectbox("Vendedor", VENDEDORES, index=v_idx, key="r_vendedor")
    jefe_proyecto_r = c2.text_input("Jefe de Proyecto", placeholder="Ej: Nicolás Parra", key="r_jefe_proyecto")
    direccion_r     = c2.text_input("Dirección", placeholder="Ej: Av. Principal 123", key="r_direccion")
    correo_r        = c2.text_input("Correo de Contacto", placeholder="correo@empresa.cl", key="r_correo")
    telefono_r      = c2.text_input("Número de Contacto", placeholder="Ej: 56912345678", key="r_telefono")
    PLANES_R        = ["Express (0-100 colab)", "Base (101-200 colab)",
                       "Estandar (201-800 colab)", "Full (801-3000 colab)", "Mega Full (3001+)"]
    plan_z          = z("plan")
    plan_idx        = next((i for i, p in enumerate(PLANES_R) if plan_z.lower() in p.lower()), 0)
    plan_r          = c2.selectbox("Plan Contratado", PLANES_R, index=plan_idx, key="r_plan")
    col_r, col_rs   = st.columns(2)
    colab_z         = z("colaboradores")
    colaboradores_r = col_r.number_input("Cantidad de Colaboradores", min_value=1,
                                          value=max(1, int(colab_z) if str(colab_z).isdigit() else 1),
                                          step=1, key="r_colab")
    razones_z       = z("razones")
    razones_r       = col_rs.number_input("Cantidad de Razones Sociales", min_value=1,
                                           value=max(1, int(razones_z) if str(razones_z).isdigit() else 1),
                                           step=1, key="r_razones")

    st.divider()
    st.subheader("Configuración de Remuneraciones")
    c3, c4 = st.columns(2)
    estructura_r     = c3.selectbox("Estructura de Remuneraciones",
                                    ["Fijos y variables", "Solo fijos", "Solo variables", "Otro"], key="r_estructura")
    comision_r       = c3.selectbox("Comisión / Semana Corrida", SI_NO, key="r_comision")
    reliquidacion_r  = c3.selectbox("Reliquidación / Renta Accesoria", SI_NO, key="r_reliq")
    tres_dias_r      = c3.selectbox("3 Primeros Días (Art. 195)", ["No aplica", "si", "no"], key="r_tresdias")
    zona_extrema_r   = c4.selectbox("Zona Extrema", SI_NO2, key="r_zona")
    provision_r      = c4.selectbox("Provisión Vacaciones", SI_NO2, key="r_provision")
    centralizacion_r = c4.selectbox("Centralización Contable",
                                    ["Manager+", "Manager Time", "SAP R3", "SAP B1", "SAP RA3",
                                     "Softland", "Defontana", "Laudus", "Chipax", "Oracle",
                                     "D365", "otro", "no aplica"], key="r_central")
    if centralizacion_r == "otro":
        centralizacion_r = c4.text_input("¿Cuál sistema contable?", placeholder="Escribe el sistema...", key="r_central_otro")
    BANCOS = ["Banco BCI", "Banco BICE", "Banco Consorcio", "Banco Coopeuch",
              "Banco de Chile", "Banco del Estado de Chile", "Banco Edwards",
              "Banco Falabella", "Banco Internacional", "Banco ITAU", "Banco Ripley",
              "Banco Santander Chile", "Banco Security", "BBVA", "Citibank",
              "Corpbanca", "Global 66", "HSBC Bank Chile", "Mach", "Mercado Pago",
              "Prex Chile", "Scotiabank", "Tenpo", "Los Heroes", "Sin Banco"]
    bancos_sel      = st.multiselect("Transferencia Bancaria", BANCOS, key="r_transfer")
    transferencia_r = ", ".join(bancos_sel) if bancos_sel else ""
    usa_api_r       = c4.selectbox("¿Utiliza API?", ["no", "si"], key="r_api")

    st.divider()
    st.subheader("Comentarios Generales")
    observaciones_r = st.text_area("Minuta / Observaciones", height=80,
                                   placeholder="Ej: SB+ no tiene gratificación + Colación y movilización",
                                   key="r_obs")

with tab_asi:
    st.subheader("Datos de la Empresa")

    panel_ots("asi_busq_ot", "asi_tabla_ots")

    if st.session_state.zoho_msg:
        tipo, msg = st.session_state.zoho_msg
        if tipo == "ok":
            st.success(msg)
        else:
            st.warning(msg)

    c5, c6 = st.columns(2)
    vendedor_z      = z("vendedor")
    empresa_a       = c5.text_input("Empresa", placeholder="Ej: Municipalidad de Marchigue", key="a_empresa")
    rut_a           = c5.text_input("RUT", placeholder="Ej: 69091300-3", key="a_rut")
    jefe_proyecto_a = c5.text_input("Jefe de Proyecto", placeholder="Ej: Nicolás Parra", key="a_jefe_proyecto")
    direccion_a     = c5.text_input("Dirección", placeholder="Ej: Maria Errazuriz 1507", key="a_direccion")
    v_idx_a         = VENDEDORES.index(vendedor_z) if vendedor_z in VENDEDORES else 0
    vendedor_a      = c5.selectbox("Vendedor", VENDEDORES, index=v_idx_a, key="a_vendedor")
    EMP_VENTA       = ["REX", "Visma", "Manager", "Otro"]
    emp_venta_z     = z("empresa_venta")
    ev_idx          = EMP_VENTA.index(emp_venta_z) if emp_venta_z in EMP_VENTA else 0
    empresa_venta_a = c6.selectbox("Empresa Venta", EMP_VENTA, index=ev_idx, key="a_emp_venta")
    contacto_nombre = c6.text_input("Contacto (Nombre Completo)", placeholder="Nombre del contacto", key="a_cont_nombre")
    contacto_numero = c6.text_input("Contacto (Número)", placeholder="Ej: 56912345678", key="a_cont_num")
    contacto_email  = c6.text_input("Contacto (Email)", placeholder="correo@empresa.cl", key="a_cont_email")

    st.divider()
    st.subheader("Plan de Implementación")
    plan_a             = st.text_input("Plan Asistencia",
                                       placeholder="Ej: PLAN ASISTENCIA CON MARCAJE Reloj, APP Y/O WEB", key="a_plan")
    c7, c8             = st.columns(2)
    casino_a           = c7.text_input("Plan Casino", value="NO APLICA", key="a_casino")
    adicionales_plan_a = c8.text_input("Adicionales", value="NO APLICA", key="a_adicionales_plan")

    st.divider()
    st.subheader("Consultas Técnicas")
    c9, c10       = st.columns(2)
    sistema_a     = c9.text_input("Sistema de Asistencia Actual", placeholder="Ej: Cass, Manual", key="a_sistema")
    dispositivo_a = c9.text_input("Dispositivo de Marcaje", placeholder="Ej: APP y Reloj control", key="a_dispositivo")
    tiene_rex_a   = c9.selectbox("¿Tiene Rex+?", ["No", "Si"], key="a_tiene_rex")
    cant_rut_a    = c9.number_input("Cantidad de RUT", min_value=1, value=1, step=1, key="a_cant_rut")
    cant_emp_a    = c10.number_input("Cantidad de Empleados", min_value=1, value=1, step=1, key="a_cant_emp")
    art22_a       = c10.selectbox("Empleados Art. 22", ["No", "Si", "Parcial"], key="a_art22")
    horario_a     = c10.text_input("Tipos de Horario", placeholder="Ej: Varios, Turno fijo", key="a_horario")
    ubicaciones_a = c10.text_input("Cantidad de Ubicaciones", placeholder="Ej: 1, Varias", key="a_ubicaciones")

    st.divider()
    st.subheader("Configuración Adicional")
    c11, c12        = st.columns(2)
    concepto_a      = c11.text_input("Concepto de Asistencia (Remuneración)",
                                     placeholder="Ej: Horas atraso y extra", key="a_concepto")
    cortes_a        = c11.text_input("Cortes Mensuales", placeholder="Ej: 24c/m", key="a_cortes")
    observaciones_a = c12.text_area("Adicionales / Observaciones", height=80,
                                    placeholder="Observaciones adicionales...", key="a_obs")

with tab_do:
    st.subheader("Desarrollo Organizacional")

    panel_ots("do_busq_ot", "do_tabla_ots")

    if st.session_state.zoho_msg:
        tipo, msg = st.session_state.zoho_msg
        if tipo == "ok":
            st.success(msg)
        else:
            st.warning(msg)

    SUBMODULOS = ["Selección", "Desempeño", "Clima"]
    submodulos_do = st.multiselect(
        "Submódulos contratados",
        SUBMODULOS,
        help="Se generará una hoja en el Excel por cada submódulo seleccionado.",
        key="d_submodulos",
    )

    st.divider()
    st.subheader("Datos Generales del Cliente")
    d1, d2          = st.columns(2)
    vendedor_z      = z("vendedor")
    empresa_do      = d1.text_input("Empresa / Razón Social", placeholder="Ej: Fundación Ejemplo", key="d_empresa")
    rut_do          = d1.text_input("RUT Empresa", placeholder="Ej: 65058734-0", key="d_rut")
    v_idx_d         = VENDEDORES.index(vendedor_z) if vendedor_z in VENDEDORES else 0
    vendedor_do     = d1.selectbox("Vendedor", VENDEDORES, index=v_idx_d, key="d_vendedor")
    jefe_do         = d1.text_input("Jefe de Proyecto", placeholder="Ej: Javiera Bravo", key="d_jefe_proyecto")
    contacto_do     = d2.text_input("Contacto Cliente (Nombre Completo)", placeholder="Nombre del contacto", key="d_cont_nombre")
    correo_do       = d2.text_input("Correo de Contacto", placeholder="correo@empresa.cl", key="d_correo")
    telefono_do     = d2.text_input("Número de Contacto", placeholder="Ej: 56912345678", key="d_telefono")
    plan_do         = d2.text_input("Plan Contratado", placeholder="Ej: Plan DO Full", key="d_plan")
    colab_z_do      = z("colaboradores")
    colaboradores_do = st.number_input("Cantidad de Colaboradores", min_value=1,
                                       value=max(1, int(colab_z_do) if str(colab_z_do).isdigit() else 1),
                                       step=1, key="d_colab")

    # ── Selección ─────────────────────────────────────────────────────────────
    if "Selección" in submodulos_do:
        st.divider()
        st.subheader("🎯 Configuración de Selección")
        s1, s2 = st.columns(2)
        sel_procesos  = s1.number_input("Procesos de selección al mes (estimado)", min_value=0, value=0, step=1, key="d_sel_procesos")
        sel_usuarios  = s1.number_input("Usuarios reclutadores", min_value=1, value=1, step=1, key="d_sel_usuarios")
        sel_cargos    = s1.text_input("Cargos / perfiles a levantar", placeholder="Ej: Operario, Supervisor, Administrativo", key="d_sel_cargos")
        sel_aprobacion = s1.selectbox("Flujo de aprobación del requerimiento",
                                      ["Sin aprobación", "1 aprobador", "2 aprobadores", "3 o más", "Por definir"],
                                      key="d_sel_aprobacion")
        ETAPAS = ["Postulación", "Filtro curricular", "Pruebas psicolaborales", "Entrevista RRHH",
                  "Entrevista jefatura", "Referencias laborales", "Examen preocupacional",
                  "Oferta", "Contratación"]
        etapas_sel    = s2.multiselect("Etapas del proceso", ETAPAS, key="d_sel_etapas")
        PORTALES = ["Portal propio Rex+", "LinkedIn", "Trabajando.com", "Laborum", "Indeed",
                    "Chiletrabajos", "Redes sociales", "Bolsa interna", "No aplica"]
        portales_sel  = s2.multiselect("Portales de publicación", PORTALES, key="d_sel_portales")
        sel_pruebas   = s2.text_input("Pruebas / evaluaciones aplicadas", placeholder="Ej: Test psicolaboral externo", key="d_sel_pruebas")
        s3, s4 = st.columns(2)
        sel_onboarding = s3.selectbox("¿Integra con onboarding / ficha de empleado?", SI_NO, key="d_sel_onboarding")
        sel_firma      = s3.selectbox("¿Requiere firma electrónica de contratos?", SI_NO, key="d_sel_firma")
        sel_migracion  = s4.selectbox("¿Migra postulantes históricos?", SI_NO, key="d_sel_migracion")
        sel_obs        = st.text_area("Minuta / Observaciones — Selección", height=80,
                                      placeholder="Ej: cliente usa hoy planilla Excel para seguimiento de candidatos",
                                      key="d_sel_obs")
    else:
        sel_procesos = sel_usuarios = 0
        sel_cargos = sel_aprobacion = sel_pruebas = ""
        etapas_sel = portales_sel = []
        sel_onboarding = sel_firma = sel_migracion = sel_obs = ""

    # ── Desempeño ─────────────────────────────────────────────────────────────
    if "Desempeño" in submodulos_do:
        st.divider()
        st.subheader("📈 Configuración de Desempeño")
        TIPOS_EVAL = ["Autoevaluación", "90° (jefatura)", "180°", "270°", "360°"]
        des_tipo_sel     = st.multiselect("Tipo de evaluación", TIPOS_EVAL, key="d_des_tipo")
        e1, e2 = st.columns(2)
        des_periodicidad = e1.selectbox("Periodicidad del ciclo",
                                        ["Anual", "Semestral", "Trimestral", "Mensual", "Por proyecto", "Por definir"],
                                        key="d_des_periodicidad")
        des_evaluados    = e1.number_input("Colaboradores a evaluar", min_value=1, value=1, step=1, key="d_des_evaluados")
        des_evaluadores  = e1.number_input("Cantidad de evaluadores", min_value=1, value=1, step=1, key="d_des_evaluadores")
        des_escala       = e1.text_input("Escala de evaluación", placeholder="Ej: 1 a 5, 1 a 7, cualitativa", key="d_des_escala")
        des_competencias = e1.selectbox("¿Evalúa competencias?", SI_NO2, key="d_des_competencias")
        des_n_comp       = e2.number_input("Cantidad de competencias", min_value=0, value=0, step=1, key="d_des_n_comp")
        des_objetivos    = e2.selectbox("¿Evalúa objetivos / KPI?", SI_NO2, key="d_des_objetivos")
        des_n_obj        = e2.number_input("Objetivos por colaborador", min_value=0, value=0, step=1, key="d_des_n_obj")
        des_ponderacion  = e2.selectbox("¿Ponderación distinta por cargo o área?", SI_NO2, key="d_des_ponderacion")
        e3, e4 = st.columns(2)
        des_calibracion  = e3.selectbox("¿Requiere calibración / comité?", SI_NO, key="d_des_calibracion")
        des_plan         = e3.selectbox("¿Genera plan de desarrollo / feedback?", SI_NO, key="d_des_plan")
        des_bono         = e4.selectbox("¿Vinculado a bono o ajuste de renta?", SI_NO, key="d_des_bono")
        des_organigrama  = e4.selectbox("Organigrama cargado en Rex+", ["no", "si", "parcial"], key="d_des_organigrama")
        des_obs          = st.text_area("Minuta / Observaciones — Desempeño", height=80,
                                        placeholder="Ej: primer ciclo solo con jefaturas, 2026 se amplía a 360°",
                                        key="d_des_obs")
    else:
        des_tipo_sel = []
        des_periodicidad = des_escala = des_competencias = des_objetivos = ""
        des_ponderacion = des_calibracion = des_plan = des_bono = des_organigrama = ""
        des_evaluados = des_evaluadores = des_n_comp = des_n_obj = 0
        des_obs = ""

    # ── Clima ─────────────────────────────────────────────────────────────────
    if "Clima" in submodulos_do:
        st.divider()
        st.subheader("🌡️ Configuración de Clima")
        k1, k2 = st.columns(2)
        cli_tipo         = k1.selectbox("Tipo de encuesta",
                                        ["Clima organizacional", "Pulso", "Satisfacción", "eNPS",
                                         "Encuesta de salida", "Otro"], key="d_cli_tipo")
        cli_periodicidad = k1.selectbox("Periodicidad de aplicación",
                                        ["Anual", "Semestral", "Trimestral", "Mensual", "Única", "Por definir"],
                                        key="d_cli_periodicidad")
        cli_universo     = k1.number_input("Universo de encuestados", min_value=1, value=1, step=1, key="d_cli_universo")
        cli_anonima      = k1.selectbox("¿Encuesta anónima?", SI_NO2, key="d_cli_anonima")
        cli_dimensiones  = k1.number_input("Cantidad de dimensiones / secciones", min_value=0, value=0, step=1, key="d_cli_dimensiones")
        cli_preguntas    = k2.number_input("Cantidad de preguntas", min_value=0, value=0, step=1, key="d_cli_preguntas")
        TIPOS_PREG = ["Escala Likert", "Selección única", "Selección múltiple", "Pregunta abierta", "NPS"]
        tipos_preg_sel   = k2.multiselect("Tipos de pregunta", TIPOS_PREG, key="d_cli_tipos_preg")
        CANALES = ["Correo electrónico", "App móvil Rex+", "Portal Rex+", "WhatsApp", "Link público", "Papel"]
        canales_sel      = k2.multiselect("Canales de difusión", CANALES, key="d_cli_canales")
        SEGMENTOS = ["Área / Gerencia", "Sucursal", "Cargo", "Antigüedad", "Género", "Rango de edad", "Sin segmentar"]
        segmentos_sel    = k2.multiselect("Segmentación de resultados", SEGMENTOS, key="d_cli_segmentacion")
        k3, k4 = st.columns(2)
        cli_fecha_val    = k3.date_input("Fecha estimada de lanzamiento", value=None, format="DD/MM/YYYY", key="d_cli_fecha")
        cli_informe      = k3.selectbox("¿Requiere informe de resultados Rex+?", SI_NO, key="d_cli_informe")
        cli_propio       = k4.selectbox("¿Cuestionario propio del cliente?", SI_NO, key="d_cli_propio")
        cli_obs          = st.text_area("Minuta / Observaciones — Clima", height=80,
                                        placeholder="Ej: cliente entrega cuestionario validado por su consultora externa",
                                        key="d_cli_obs")
    else:
        cli_tipo = cli_periodicidad = cli_anonima = cli_informe = cli_propio = ""
        cli_universo = cli_dimensiones = cli_preguntas = 0
        tipos_preg_sel = canales_sel = segmentos_sel = []
        cli_fecha_val = None
        cli_obs = ""

    if not submodulos_do:
        st.info("Selecciona al menos un submódulo arriba para completar la minuta de DO.")


# ── Descarga ──────────────────────────────────────────────────────────────────
st.divider()
st.subheader("Descargar Minutas")

nombre_archivo = st.text_input(
    "Nombre del archivo (sin extensión)",
    value=f"Minuta_{empresa_r or empresa_a or empresa_do or 'cliente'}".replace(" ", "_"),
    key="nombre_archivo",
)

col_h1, col_h2 = st.columns(2)
incluir_rem = col_h1.checkbox("Incluir hoja Remuneraciones", value=True, key="inc_rem")
incluir_asi = col_h2.checkbox("Incluir hoja Asistencia", value=True, key="inc_asi")

if st.button("📥 Generar y Descargar Excel", type="primary", use_container_width=False):
    data_rem = {
        "ot": ot, "empresa": empresa_r, "rut": rut_r,
        "vendedor": vendedor_r, "jefe_proyecto": jefe_proyecto_r, "direccion": direccion_r,
        "correo": correo_r, "telefono": telefono_r,
        "plan": plan_r, "colaboradores": colaboradores_r,
        "razones_sociales": razones_r, "estructura": estructura_r,
        "comision": comision_r, "reliquidacion": reliquidacion_r,
        "tres_dias": tres_dias_r, "zona_extrema": zona_extrema_r,
        "provision": provision_r, "centralizacion": centralizacion_r,
        "transferencia": transferencia_r, "usa_api": usa_api_r, "observaciones": observaciones_r,
    }
    data_asi = {
        "empresa": empresa_a, "rut": rut_a, "jefe_proyecto": jefe_proyecto_a, "direccion": direccion_a,
        "vendedor": vendedor_a, "empresa_venta": empresa_venta_a,
        "contacto_nombre": contacto_nombre, "contacto_numero": contacto_numero,
        "contacto_email": contacto_email, "plan_asistencia": plan_a,
        "plan_casino": casino_a, "adicionales_plan": adicionales_plan_a,
        "sistema_actual": sistema_a, "dispositivo": dispositivo_a,
        "tiene_rex": tiene_rex_a, "cantidad_rut": cant_rut_a,
        "cantidad_empleados": cant_emp_a, "art22": art22_a,
        "tipos_horario": horario_a, "ubicaciones": ubicaciones_a,
        "concepto": concepto_a, "cortes": cortes_a,
        "observaciones": observaciones_a,
    }
    data_do = {
        "ot": ot, "empresa": empresa_do, "rut": rut_do, "vendedor": vendedor_do,
        "jefe_proyecto": jefe_do, "contacto_nombre": contacto_do,
        "correo": correo_do, "telefono": telefono_do, "plan": plan_do,
        "colaboradores": colaboradores_do,
        # Selección
        "sel_procesos": sel_procesos, "sel_cargos": sel_cargos,
        "sel_usuarios": sel_usuarios, "sel_aprobacion": sel_aprobacion,
        "sel_etapas": ", ".join(etapas_sel), "sel_portales": ", ".join(portales_sel),
        "sel_pruebas": sel_pruebas, "sel_onboarding": sel_onboarding,
        "sel_firma": sel_firma, "sel_migracion": sel_migracion, "sel_obs": sel_obs,
        # Desempeño
        "des_tipo": ", ".join(des_tipo_sel), "des_periodicidad": des_periodicidad,
        "des_evaluados": des_evaluados, "des_evaluadores": des_evaluadores,
        "des_competencias": des_competencias, "des_n_competencias": des_n_comp,
        "des_objetivos": des_objetivos, "des_n_objetivos": des_n_obj,
        "des_escala": des_escala, "des_ponderacion": des_ponderacion,
        "des_calibracion": des_calibracion, "des_plan": des_plan,
        "des_bono": des_bono, "des_organigrama": des_organigrama, "des_obs": des_obs,
        # Clima
        "cli_tipo": cli_tipo, "cli_periodicidad": cli_periodicidad,
        "cli_universo": cli_universo, "cli_anonima": cli_anonima,
        "cli_dimensiones": cli_dimensiones, "cli_preguntas": cli_preguntas,
        "cli_tipos_preg": ", ".join(tipos_preg_sel), "cli_canales": ", ".join(canales_sel),
        "cli_segmentacion": ", ".join(segmentos_sel),
        "cli_fecha": cli_fecha_val.strftime("%d-%m-%Y") if cli_fecha_val else "",
        "cli_informe": cli_informe, "cli_propio": cli_propio, "cli_obs": cli_obs,
    }
    excel_bytes = generar_excel(
        data_rem, data_asi,
        data_do=data_do,
        submodulos=submodulos_do,
        incluir_rem=incluir_rem,
        incluir_asi=incluir_asi,
    )
    st.download_button(
        label="⬇️ Haz clic aquí para descargar",
        data=excel_bytes,
        file_name=f"{nombre_archivo}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="descarga_excel",
    )
    st.success("✅ Minuta generada. Haz clic en el botón azul para descargar.")

if BRANDING:
    aplicar_footer()
