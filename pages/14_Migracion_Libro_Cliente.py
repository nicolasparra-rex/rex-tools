# -*- coding: utf-8 -*-
"""Rex-tools · Migración detalle desde el Libro de Remuneraciones de cualquier cliente.
Lee un libro con estructura arbitraria, autodetecta período/estructura, propone el mapeo de
conceptos contra el catálogo del cliente y genera la planilla de migración detalle, cuadrada
al peso. Empresa/mutual/contrato/caja se resuelven por RUT desde la dotación."""
import streamlit as st
import pandas as pd
import io, os, json, csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from libro_engine import (norm, _num, load_grid, detect_header_row, match_struct, classify_and_map,
                          generar_detalle, cargar_homologacion, cargar_dotacion, detectar_periodo,
                          leer_catalogo_rex, cargar_base_estandar, OUT_COLS)

try:
    from lib.branding import aplicar_branding, aplicar_footer, hero
except Exception:
    def aplicar_branding(**k): pass
    def aplicar_footer(): pass
    def hero(t, d="", i=""): st.title(t); st.caption(d)

DATA_DIR = "data"
AMBAR = "#FFF3CD"; ROJO = "#F8D7DA"; VERDE = "#D4EDDA"

st.set_page_config(page_title="Rex+ | Migración desde Libro del Cliente", page_icon="📘", layout="wide")
aplicar_branding(titulo_pagina="Migración desde Libro", badge="BETA")
hero("📘 Migración detalle desde el Libro del Cliente",
     "Sube el libro del cliente y la dotación, confirma el mapeo de conceptos y genera la migración detalle cuadrada al peso.")

# ---------- helpers de datos de referencia ----------
# Nota: la caché se invalida por la fecha de modificación del archivo (mtime). Así, si el
# archivo en data/ se agrega o se actualiza, la tabla se recarga sola (no queda un "vacío" pegado).
def _mtime(path):
    return os.path.getmtime(path) if os.path.exists(path) else 0

@st.cache_data(show_spinner=False)
def _parametros_de(path, _mtime):
    wp = load_workbook(path, data_only=True).worksheets[0]
    ph = [wp.cell(row=1, column=c).value for c in range(1, wp.max_column + 1)]
    return {wp.cell(row=r, column=1).value: {ph[c-1]: wp.cell(row=r, column=c).value
            for c in range(1, wp.max_column + 1)} for r in range(2, wp.max_row + 1) if wp.cell(row=r, column=1).value}

def cargar_parametros(file=None):
    if file is not None:
        return _parametros_de(file, getattr(file, "name", "subido"))
    p = os.path.join(DATA_DIR, "parametrosMesuales.xlsx")
    return _parametros_de(p, _mtime(p)) if os.path.exists(p) else {}

@st.cache_data(show_spinner=False)
def cargar_cot_hist():
    ch = load_workbook(os.path.join(DATA_DIR, "cot_afp_hist.xlsx"), data_only=True).worksheets[0]
    return {f"{ch.cell(row=r,column=2).value}{ch.cell(row=r,column=3).value}": (ch.cell(row=r,column=5).value or 0)
            for r in range(2, ch.max_row + 1)}

@st.cache_data(show_spinner=False)
def _homolog_de(path, _mtime):
    return cargar_homologacion(path) if os.path.exists(path) else []

def cargar_homolog_default():
    p = os.path.join(DATA_DIR, "listado_instituciones.xlsx")
    return _homolog_de(p, _mtime(p))

@st.cache_data(show_spinner=False)
def cargar_ids_estandar():
    """Lista maestra de IDs de concepto estándar de Rex (desde data/equiv_conceptos.xlsx).
    Sirve como universo base para el desplegable, además del catálogo del cliente."""
    p = os.path.join(DATA_DIR, "equiv_conceptos.xlsx")
    if not os.path.exists(p):
        return []
    try:
        x = pd.read_excel(p)
        col = "concepto_detalle" if "concepto_detalle" in x.columns else x.columns[1]
        return sorted({str(v).strip() for v in x[col].dropna() if str(v).strip()})
    except Exception:
        return []

def leer_catalogo(file):
    df = pd.read_excel(file, header=None)
    hr = 0
    for r in range(min(6, len(df))):
        vals = [norm(x) for x in df.iloc[r].values]
        if "concepto" in vals and any("nombre" in v for v in vals): hr = r; break
    hdr = [norm(x) for x in df.iloc[hr].values]
    ci = hdr.index("concepto") if "concepto" in hdr else 0
    ni = next((i for i, v in enumerate(hdr) if "nombre" in v), 2)
    names, valid = {}, {}
    for _, row in df.iloc[hr+1:].iterrows():
        cid = row[ci]; nom = row[ni]
        if pd.notna(cid) and str(cid).strip():
            valid[str(cid).strip()] = (str(nom).strip() if pd.notna(nom) else "")
            if pd.notna(nom): names[norm(nom)] = str(cid).strip()
    return names, valid

def to_csv(filas):
    """Archivo de salida en CSV UTF-8 (con BOM, separador coma) — formato de carga a Rex."""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(OUT_COLS)
    for rr in filas:
        w.writerow(rr)
    return buf.getvalue().encode("utf-8-sig")

# ---------- Barra lateral: solo lo mínimo ----------
with st.sidebar:
    st.subheader("⚙️ Datos del cliente")
    cliente = st.text_input("Nombre del cliente", value="", placeholder="ej. thoughtworks")
    st.caption("Empresa, mutual, contrato y caja se resuelven por RUT desde la **dotación**. "
               "El período se detecta del archivo.")
    with st.expander("Avanzado (valores por defecto)"):
        apv_inst = st.text_input("Institución APV / Cuenta 2", value="afp")
        caja_inst = st.text_input("Caja CCAF (si la dotación no la trae)", value="losandes")
        jornada = st.text_input("Jornada", value="C")
# fallbacks internos (se completan desde la dotación)
empresa_id, mutual_id, num_contrato = "", "", 1

_DOTACION_SQL = '''SELECT
    contr.empleado                                                          AS "Id empleado",
    contr.contrato                                                          AS "Numero de contrato",
    contr."fechaInic"                                                       AS "fechaInic",
    contr."tipoCont"                                                        AS "tipoCont",
    (SELECT emp.jubilado FROM T$empleados emp WHERE emp.empleado = contr.empleado) AS "jubilado",
    e.identificador_nacional                                                AS "Rut empresa",
    e.empresa                                                               AS "idempresa",
    e.mutual                                                                AS "Mutual",
    e."cotizacionMutu"                                                      AS "% mutual"
FROM T$empleadoscontr contr
INNER JOIN T$empresas e ON e.empresa = contr.empresa'''

# ---------- 0. Antes de empezar (orientación) ----------
st.markdown("### Antes de empezar")
st.markdown(
    "Para migrar la historia de un cliente necesitas preparar **3 archivos** y luego subirlos abajo:\n\n"
    "- 📘 **Libro de remuneraciones** del cliente — te lo entrega el cliente. **(Obligatorio)**\n"
    "- 📗 **Catálogo de conceptos** del cliente — se **exporta de su Rex** (hoja *Lista de conceptos*). **(Obligatorio)**\n"
    "- 📙 **Dotación** — la **generas tú** corriendo la consulta de abajo y exportando el resultado a Excel. **(Recomendada)**")
with st.expander("📋 Paso previo — genera la dotación con esta consulta (para el consultor)"):
    st.caption("Córrela en la base del cliente y exporta el resultado a Excel. **Ese archivo es el que subes como ③ Dotación.** "
               "Con él la app resuelve contrato / empresa / mutual por RUT.")
    st.code(_DOTACION_SQL, language="sql")
    st.caption("Es **una fila por trabajador** (no el listado de empresas). La primera columna es la llave para "
               "casar con el libro. Corre la consulta, exporta a Excel y súbelo como ③ Dotación.")
st.divider()

# ---------- 1. Archivos ----------
st.markdown("### 1 · Sube los archivos")
# Estilos SOLO de nuestro apartado (clases propias rex-upc-*, no tocan el menú lateral ni nada global).
st.markdown("""
<style>
.rex-upc{margin:-1rem -1rem .7rem -1rem;padding:.6rem 1rem;font-weight:700;color:#fff;
         border-radius:.5rem .5rem 0 0;font-size:.95rem;line-height:1.2;
         display:flex;align-items:center;min-height:2.7rem;box-sizing:border-box;}
.rex-upc .n{opacity:.85;font-weight:800;margin-right:.35rem;}
.rex-upc .rex-badge{margin-left:auto;font-size:.68rem;font-weight:700;background:rgba(255,255,255,.28);
         padding:.12rem .5rem;border-radius:1rem;white-space:nowrap;}
.rex-upc-ambar .rex-badge{background:rgba(0,0,0,.14);}
.rex-upc-azul{background:#1e5591;}
.rex-upc-ambar{background:#F5A623;color:#3d2c00;}
.rex-upc-verde{background:#7CB342;}
</style>
""", unsafe_allow_html=True)
c1, c2, c3 = st.columns(3)
with c1:
    with st.container(border=True):
        st.markdown('<div class="rex-upc rex-upc-azul"><span class="n">①</span>Libro de remuneraciones<span class="rex-badge">Obligatorio</span></div>', unsafe_allow_html=True)
        libro_file = st.file_uploader("Libro", type=["xlsx", "xls"], key="up_libro", label_visibility="collapsed")
        st.caption("Del cliente · .xlsx o .xls")
with c2:
    with st.container(border=True):
        st.markdown('<div class="rex-upc rex-upc-ambar"><span class="n">②</span>Catálogo de conceptos<span class="rex-badge">Obligatorio</span></div>', unsafe_allow_html=True)
        cat_file = st.file_uploader("Catálogo", type=["xlsx"], key="up_cat", label_visibility="collapsed")
        st.caption("Exportado del Rex del cliente")
with c3:
    with st.container(border=True):
        st.markdown('<div class="rex-upc rex-upc-verde"><span class="n">③</span>Dotación<span class="rex-badge">Recomendada</span></div>', unsafe_allow_html=True)
        dot_file = st.file_uploader("Dotación", type=["xlsx"], key="up_dot", label_visibility="collapsed")
        st.caption("RUT → contrato / empresa / mutual / caja")

_XLMIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
def _tabla_mantenible(titulo, archivo, ayuda):
    """Expander estándar: descargar la tabla actual de data/, editarla y volver a subirla."""
    with st.expander(titulo):
        st.caption(ayuda)
        p = os.path.join(DATA_DIR, archivo)
        if os.path.exists(p):
            with open(p, "rb") as f:
                st.download_button(f"⬇️ Descargar {archivo} actual", f.read(), file_name=archivo,
                                   mime=_XLMIME, key=f"dl_{archivo}")
        else:
            st.warning(f"⚠️ No existe `data/{archivo}` en el sitio.")
        return st.file_uploader(f"Reemplazar {archivo} (opcional)", type=["xlsx"], key=f"up_{archivo}")

st.divider()
st.caption("⚙️ **Tablas de referencia del sitio — revísalas antes de migrar.** En especial, confirma que los "
           "**parámetros mensuales incluyan el mes que vas a cargar**. Puedes descargarlas, editarlas y volver a subirlas.")
homolog_file = _tabla_mantenible(
    "🏛️ Homologación de instituciones (descargar / actualizar / subir)", "listado_instituciones.xlsx",
    "Traduce el texto del libro (‘MODELO’, ‘Colmena…’) al ID de Rex. Edítala y súbela para actualizarla.")

# Parámetros mensuales con estado de cobertura de meses (para que revisen si tienen el mes a migrar).
_pp = os.path.join(DATA_DIR, "parametrosMesuales.xlsx")
_ult_mes = None
if os.path.exists(_pp):
    _P = _parametros_de(_pp, _mtime(_pp))
    _con = sorted(m for m, r in _P.items()
                  if _num(r.get("topeImp_pesos_afp", 0)) > 0 and _num(r.get("sis", 0)) > 0)
    _ult_mes = _con[-1] if _con else None
with st.expander(f"📅 Parámetros mensuales — con datos hasta {_ult_mes or '—'}  (descargar / actualizar / subir)"):
    if _ult_mes:
        st.success(f"✅ Tope de salud y % SIS cargados **hasta {_ult_mes}**. "
                   f"Si el mes que vas a migrar es **posterior a {_ult_mes}**, agrégalo antes: descarga la tabla, "
                   "completa la fila del mes y vuelve a subirla.")
    else:
        st.warning("⚠️ La tabla no tiene meses con datos válidos. Actualízala antes de migrar.")
    if os.path.exists(_pp):
        with open(_pp, "rb") as f:
            st.download_button("⬇️ Descargar parametrosMesuales.xlsx actual", f.read(),
                               file_name="parametrosMesuales.xlsx", mime=_XLMIME, key="dl_param")
    param_file = st.file_uploader("Reemplazar parametrosMesuales.xlsx (opcional)", type=["xlsx"], key="up_param")

if not libro_file:
    st.info("Sube el **libro** para comenzar. El **catálogo** ayuda a proponer los IDs y la **dotación** "
            "completa empresa/mutual/contrato/caja.")
    aplicar_footer(); st.stop()

by_id, name_to_id = ({}, {})
if cat_file: by_id, name_to_id = leer_catalogo_rex(cat_file)
base_estandar = cargar_base_estandar(os.path.join(DATA_DIR, "conceptos_base.xlsx"))
if not by_id:
    st.error("🔴 Falta el **catálogo de conceptos del cliente** (② en el paso 1). Es **obligatorio**: "
             "de él salen los IDs válidos de Rex. Expórtalo de su Rex (hoja *Lista de conceptos*) y súbelo.")
    aplicar_footer(); st.stop()
saved = {}
homolog = cargar_homologacion(homolog_file) if homolog_file else cargar_homolog_default()
if not homolog:
    st.error("🔴 **No se cargó la tabla de homologación de instituciones.** Las AFP, salud, mutual y caja "
             "saldrán **sin homologar** (texto crudo del libro). Revisa que exista `data/listado_instituciones.xlsx` "
             "o súbela en el paso 1 (‘Homologación de instituciones’).")
dotacion = cargar_dotacion(dot_file) if dot_file else {}
params_all = cargar_parametros(param_file) if param_file else cargar_parametros()

df, sheet = load_grid(libro_file)
hr = detect_header_row(df)
hdr = [x if str(x) != "nan" else "" for x in df.iloc[hr].values]
struct = match_struct(hdr)
propuesta = classify_and_map(hdr, struct, catalog_names=name_to_id, saved=saved, valid_ids=set(by_id))

# ---------- 2. Período + estructura ----------
st.markdown("### 2 · Período y estructura detectados")
periodo = detectar_periodo(df, libro_file.name)
ca, cb = st.columns([1, 3])
if periodo:
    ca.success(f"📅 Período: **{periodo}**")
else:
    periodo = ca.text_input("📅 No pude detectar el período — escríbelo (AAAA-MM)", value="", placeholder="2026-06")
cb.caption(f"Hoja: **{sheet}** · Encabezado en fila **{hr+1}**")

faltan = [k for k in ["rut", "total_haberes", "total_descuentos", "liquido", "base_afp"] if k not in struct]
if faltan:
    st.error(f"❌ No detecté columnas estructurales: **{', '.join(faltan)}**. Revisa el libro antes de continuar.")
if not dotacion:
    st.warning("⚠️ No subiste la **dotación**: empresa, mutual, contrato y caja no se podrán resolver por RUT.")

# Aviso si el período no tiene parámetros mensuales cargados (tope salud / SIS).
prow = params_all.get(periodo, {}) if periodo else {}
param_ok = bool(prow) and _num(prow.get("topeImp_pesos_afp", 0)) > 0 and _num(prow.get("sis", 0)) > 0
if periodo and not param_ok:
    st.error(f"🔴 **Faltan los parámetros mensuales de {periodo}** (tope de salud y/o % de SIS vienen en 0). "
             "Si generas así, el **afecto de isapre saldrá sin tope** y la **cotización de SIS en 0**. "
             "Actualiza `parametrosMesuales.xlsx` en el paso 1 (agrega el mes) y vuelve a subirla.")

# ---------- 3. Mapeo ----------
st.markdown("### 3 · Confirma el mapeo de conceptos")
st.caption("Cada columna del libro se asocia a un concepto **del catálogo del cliente**. "
           "El **nombre y el tipo** (haber/descuento/aporte) salen del catálogo. "
           "En **ámbar** los que faltan por asociar; en **rojo** los que no existen en el catálogo (hay que corregirlos).")

valid_ids = sorted(by_id.keys())

def _info(cid):
    """Datos derivados del catálogo para un ID: nombre, legal/propio, bloque y estado."""
    if not cid:
        return {"nombre": "", "tipo": "LEGAL/PROPIO", "bloque": "", "estado": "🟠 falta asociar"}
    if cid in by_id:
        return {"nombre": by_id[cid]["nombre"], "tipo": ("LEGAL" if cid in base_estandar else "PROPIO"),
                "bloque": by_id[cid]["bloque"] or "—", "estado": "✅ en catálogo"}
    return {"nombre": "—", "tipo": "—", "bloque": "—", "estado": "❌ no está en catálogo"}

# Bloque por posición del libro (respaldo cuando el catálogo no define bloque o el concepto no está asociado).
_GRUPO_POS = {"haber": "haber", "descuento": "desc", "aporte": "aporte", "?": ""}
filas_map = []
for r in propuesta:
    inf = _info(r["id_rex"])
    # Bloque efectivo: del catálogo si lo define; si no (pendiente o Tipo Dato/Valor Guardado), el de la posición.
    if r["id_rex"] in by_id and by_id[r["id_rex"]]["bloque"]:
        inf["bloque"] = by_id[r["id_rex"]]["bloque"]
    else:
        inf["bloque"] = _GRUPO_POS.get(r["grupo"], "")
    filas_map.append({"Columna del libro": r["header"], "Concepto Rex (nombre)": inf["nombre"],
                      "Tipo": inf["tipo"], "Bloque": inf["bloque"], "ID Rex": r["id_rex"] or "",
                      "Estado": inf["estado"]})
map_df = pd.DataFrame(filas_map)

tot = len(map_df)
auto0 = int((map_df["ID Rex"].astype(str).str.len() > 0).sum())
m1, m2, m3 = st.columns(3)
m1.metric("Conceptos", tot); m2.metric("Asociados", auto0); m3.metric("Por asociar", tot - auto0)

# posición del libro por columna (para detectar conflictos y como respaldo del bloque)
pos_by_header = {norm(r["header"]): _GRUPO_POS.get(r["grupo"], "") for r in propuesta}

editor = st.data_editor(
    map_df, use_container_width=True, hide_index=True, key="mapeo", height=430,
    column_config={
        "Columna del libro": st.column_config.TextColumn(disabled=True),
        "Concepto Rex (nombre)": st.column_config.TextColumn(disabled=True,
                    help="Nombre del concepto en el catálogo del cliente."),
        "Tipo": st.column_config.TextColumn("Legal/Propio", disabled=True, width="small"),
        "Bloque": st.column_config.SelectboxColumn("Bloque", options=["haber", "desc", "aporte"], required=False,
                    width="small",
                    help="Cómo se trata en la migración. Por defecto sale del catálogo (o de la posición del "
                         "libro). Cámbialo solo si en ESTE libro corresponde otro (es un ajuste opcional)."),
        "ID Rex": st.column_config.SelectboxColumn("ID Rex (elige del catálogo)", options=valid_ids, required=False,
                    help="Elige el concepto del catálogo del cliente. Escribe para filtrar."),
        "Estado": st.column_config.TextColumn(disabled=True, width="medium"),
    })

mapping = {norm(r["Columna del libro"]): r["ID Rex"] for _, r in editor.iterrows() if str(r["ID Rex"]).strip()}
# tipo_map = el bloque que quedó en la tabla (catálogo, posición o el override del implementador)
tipo_map = {}
for _, r in editor.iterrows():
    cid = str(r["ID Rex"]).strip(); bl = str(r["Bloque"]).strip()
    if cid and bl in ("haber", "desc", "aporte"): tipo_map[cid] = bl
invalidos = sorted(v for v in set(mapping.values()) if v not in by_id)
pend_df = editor[editor["ID Rex"].astype(str).str.strip() == ""][["Columna del libro", "Bloque"]]

# Aviso GENERAL (no atado a ningún concepto puntual): dónde el catálogo y el libro difieren en el bloque.
conflictos = []
for _, r in editor.iterrows():
    cid = str(r["ID Rex"]).strip()
    if cid in by_id and by_id[cid]["bloque"]:
        pos = pos_by_header.get(norm(r["Columna del libro"]), "")
        bl = str(r["Bloque"]).strip()
        if pos in ("haber", "desc", "aporte") and by_id[cid]["bloque"] != pos and bl == by_id[cid]["bloque"]:
            conflictos.append((r["Columna del libro"], pos, by_id[cid]["bloque"]))

if len(pend_df):
    st.warning(f"🟠 Faltan **{len(pend_df)}** concepto(s) por asociar — elige su ID del catálogo arriba:")
    st.dataframe(pend_df.style.set_properties(**{"background-color": AMBAR}),
                 hide_index=True, use_container_width=True)
elif not invalidos:
    st.success("✅ Todos los conceptos están asociados a un ID del catálogo.")
if invalidos:
    st.error("❌ Estos IDs **no existen en el catálogo del cliente**: " + ", ".join(f"`{i}`" for i in invalidos)
             + ". Rex los rechazaría al importar. Corrige el mapeo (elige uno del catálogo) "
               "o agrega el concepto al catálogo del cliente y vuelve a subirlo.")
if conflictos:
    with st.expander(f"ℹ️ {len(conflictos)} concepto(s) donde el libro y el catálogo difieren en el bloque — conviene revisar"):
        st.caption("No siempre es un error (puede ser un concepto bien clasificado). Se aplica el bloque del "
                   "**catálogo**. Si esto te descuadra al generar, revisa el mapeo (quizá va a otro concepto) "
                   "o ajusta la columna **Bloque** arriba.")
        st.dataframe(pd.DataFrame(conflictos, columns=["Columna del libro", "Bloque en el libro", "Bloque en el catálogo"]),
                     hide_index=True, use_container_width=True)

# ---------- 4. Generar ----------
st.markdown("### 4 · Generar y validar")
listo = bool(periodo) and not faltan and not len(pend_df) and not invalidos
if not listo:
    faltantes = []
    if not periodo: faltantes.append("detectar el período")
    if faltan: faltantes.append("completar la estructura")
    if len(pend_df): faltantes.append(f"asociar {len(pend_df)} concepto(s)")
    if invalidos: faltantes.append(f"corregir {len(invalidos)} ID(s) que no están en el catálogo")
    st.info("Para habilitar **Generar**, falta: " + " · ".join(faltantes) + ".")

if st.button("🚀 Generar migración detalle", type="primary", disabled=not listo, use_container_width=True):
    cfg = dict(empresa_id=empresa_id, mutual_id=mutual_id, apv_inst=apv_inst, caja_inst=caja_inst,
               num_contrato=int(num_contrato), jornada=jornada, periodo=periodo)
    filas, res = generar_detalle(df, hr, struct, mapping, params_all.get(periodo, {}),
                                 cargar_cot_hist(), cfg, homolog=homolog, dotacion=dotacion, tipo_map=tipo_map)
    ok = (res["descuadre_haberes"] == 0 and res["descuadre_descuentos"] == 0 and res["descuadre_liquido"] == 0)
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Empleados en archivo", res["empleados"]); k2.metric("Omitidos", res.get("omitidos", 0))
    k3.metric("Descuadres", res["descuadre_haberes"] + res["descuadre_descuentos"] + res["descuadre_liquido"])
    k4.metric("Estado", "✅ Cuadra" if ok else "⚠️ Revisar")
    if not ok:
        st.error(f"Descuadres → haberes {res['descuadre_haberes']}, descuentos {res['descuadre_descuentos']}, "
                 f"líquido {res['descuadre_liquido']}.")
    if res.get("log_contratos"):
        lc = pd.DataFrame(res["log_contratos"]).rename(columns={"rut": "RUT", "motivo": "Motivo"})
        st.warning(f"🔴 **{len(lc)} RUT no están en la dotación — se OMITIERON del archivo.** "
                   "Revisa esta lista; si alguno debía ir, agrégalo a la dotación y vuelve a generar.")
        st.dataframe(lc.style.set_properties(**{"background-color": ROJO}), hide_index=True, use_container_width=True)
        st.download_button("⬇️ Log RUT omitidos (.csv)", lc.to_csv(index=False).encode("utf-8-sig"),
                           file_name=f"omitidos_{cliente or 'cliente'}_{periodo}.csv", mime="text/csv")
    if res.get("log_inst"):
        li = pd.DataFrame(res["log_inst"])
        sin = li[li["estado"] == "SIN HOMOLOGAR"]
        with st.expander(f"🏛️ Instituciones homologadas ({len(li)}) — "
                         + (f"⚠️ {len(sin)} sin homologar" if len(sin) else "✅ todas OK"),
                         expanded=bool(len(sin))):
            def _color(v): return f"background-color: {ROJO}" if v == "SIN HOMOLOGAR" else f"background-color: {VERDE}"
            li_show = li.rename(columns={"tipo": "Tipo", "valor_libro": "Valor en el libro",
                                         "id_rex": "ID Rex", "estado": "Estado"})
            st.dataframe(li_show.style.apply(lambda s: [_color(v) for v in s], subset=["Estado"]),
                         hide_index=True, use_container_width=True)
            if len(sin):
                st.warning("🟠 Las marcadas **SIN HOMOLOGAR** salieron con el texto crudo del libro. "
                           "Agrégalas a la tabla de homologación (paso 1) y vuelve a generar.")
    for f in res["flags"]:
        st.warning("🟠 " + f)
    st.download_button("⬇️ Descargar migración detalle (.csv)", to_csv(filas),
                       file_name=f"migracion_detalle_{cliente or 'cliente'}_{periodo}.csv",
                       mime="text/csv", type="primary")

aplicar_footer()
