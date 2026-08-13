# -*- coding: utf-8 -*-
"""
Migración de empleados Rex -> BNOVUS
====================================
Página de rex-tools que toma el "Listado de Empleados" exportado desde Rex y
genera el archivo de carga de trabajadores en el formato de la plantilla BNOVUS.

Insumos que sube el usuario:
  - Listado de Empleados de Rex (.xlsx). Header en la fila 2, datos desde la fila 3.

Parámetros ingresables (barra lateral):
  - RUT Empresa            (obligatorio; columna A de BNOVUS)
  - Alcance                (todos / solo activos)
  - Defaults de negocio    (moneda sueldo, tipo sueldo, cantidad días, modalidad,
                            gratificación) para las columnas que Rex NO trae.

Salida:
  - Archivo .xlsx con la estructura BNOVUS (hoja Sheet1 + catálogos), listo para subir.
  - Informe de cobertura: valores no reconocidos en los mapeos (para revisar).

La plantilla BNOVUS vive en  data/plantilla_bnovus.xlsx  (se puede reemplazar por
file_uploader si no está en el repo).
"""

import io
import os
import unicodedata
from datetime import datetime

import pandas as pd
import streamlit as st
import openpyxl
import base64

# ----------------------------------------------------------------------------- #
#  Config de página + branding rex-tools (igual que las demás páginas)
# ----------------------------------------------------------------------------- #
st.set_page_config(page_title="Migración BNOVUS | Rex+ Tools",
                   page_icon="👥", layout="wide")
try:
    from lib.branding import aplicar_branding, aplicar_footer, hero
    BRANDING = True
except ImportError:
    BRANDING = False

if BRANDING:
    aplicar_branding(titulo_pagina="Migración BNOVUS", badge="BETA")

# Plantilla BNOVUS (encabezado + catálogos) embebida: la página es autónoma,
# no depende de data/ ni de subir la plantilla a mano.
PLANTILLA_BNOVUS_B64 = "UEsDBBQAAAAIAK6eDF1Gx01IlQAAAM0AAAAQAAAAZG9jUHJvcHMvYXBwLnhtbE3PTQvCMAwG4L9SdreZih6kDkQ9ip68zy51hbYpbYT67+0EP255ecgboi6JIia2mEXxLuRtMzLHDUDWI/o+y8qhiqHke64x3YGMsRoPpB8eA8OibdeAhTEMOMzit7Dp1C5GZ3XPlkJ3sjpRJsPiWDQ6sScfq9wcChDneiU+ixNLOZcrBf+LU8sVU57mym/8ZAW/B7oXUEsDBBQAAAAIAK6eDF3DAWyFHgEAAGYCAAARAAAAZG9jUHJvcHMvY29yZS54bWzFksFOwzAQRH8lyj3ZOCUtsdIcoIILlZCoBPRmOdvUahxbtktavh43pCkV3Dl6ZvZ5VtqCa8qVwWejNBon0AYH2bSWcj0Pt85pCmD5FiWzsU+03twoI5nzT1ODZnzHaoQ0SaYg0bGKOQYnYKRHYlgWFafcIHPKDPiKj3i9N00PqzhggxJbZ4HEBMLyHc1OBWtmjNjsPb2AC+gEdWik/RawGsm9+ie+dyAckgcrxlTXdXE36XN+FwJvy6eXfu1ItNaxlqOfsoK6o8Z5eP75dXK/WD2EZZqQPEqyiMxWJKPJjJKb9anrVb9LYakqsRH/3DidRsltRNIVyWk2oWn+o/G5YFn482iYdctBuDuWC4G1Ch5Z84GfBfz2e+36pMovUEsDBBQAAAAIAK6eDF1YynR38gUAAI4aAAATAAAAeGwvdGhlbWUvdGhlbWUxLnhtbO1ZzYsbNxS/F/o/iLk78z3jWeIN9thO2uwmIeuk5CjPyB5lNSMzknfXhEBJTr0UCmnppdBbD6U00EBDL/1jFhLa9I+oZsYfGlvOR+OUlMYGW3r6vaef3pOepJmLl85SAk5QzjDNWpp5wdAAyiIa42zc0m4N+o2mBhiHWQwJzVBLmyGmXdr/+KOLcI8nKEVA6GdsD7a0hPPJnq6zSIghu0AnKBNtI5qnkItqPtbjHJ4KuynRLcPw9BTiTAMZTIXZgdABMQLXRyMcIW1/Yb5HxE/GWSGISH4UlX1WOhI2PjaLPzZjIcnBCSQtTfQU09MBOuMaIJBx0dDSjPKj6fsX9aUS4Vt0Jb1++ZnrzRXiY6vUy8fDpaLjuI7XXtq3KvubuJ7f83re0l4JgFEkRmpuYN1O0Om6c6wEqooK212/a5s1vGTf3sC33eJbw9srvLOB7/fDlQ8lUFV0FT7xrdCp4d0V3tvA+0a76/g1fAlKCM6ON9CG69nhYrRLyIiSK0p44Dp935rDVyhdml2Vfsa3zbUU3qV5XwDK4EKOM8BnEzSCkcCFkOBhjsEBHidi4k1gRpkQG5bRN2zxW3ydslR6BO4hKGlXoohtiAo+gEU5nvCW9qmwqkmQZ0+fnj94cv7g1/OHD88f/Dzve1PvCszGst6LH77667vPwZ+/fP/i0ddqPJPxz3/64vlvv7/MPK/R+ubx8yePn3375R8/PlLA2zkcyvABThED19ApuElTMUBFB2iYv5nGIIG4pgETgVQAezypAa/NIFHhOqjuwtu5yBQq4OXp3RrXoySfcqwAXk3SGvCQUtKhuXI4V4u+5OFMs7G683wq425CeKLqO1wLcG86EVMeq0yGCarRvEFEtOEYZYiDoo0eI6RQu4Nxza+HOMopoyMO7mDQgVjpkgEecrXSFZyKuMxUBEWoa745vA06lKjMd9FJHSmWBSQqk4jU3HgZTjlMlYxhSmTkAeSJiuTRLI9qDmdcRHqMCAW9GDGm0rmez2p0r4oMow77IZmldWTO8bEKeQAplZFdehwmMJ0oOeMskbGfsGMxRSG4QbmSBK2vkKIu4gCzreG+jRF/s2V9S2Qg9QQpWqa5akkgWl+PMzKCKJtvBLWUnuLslfl9LbO7/05mf2c5fffZvJ1j5Zpaz+HbcP/BzN2F0+wGEovlQ+L+kLj/j4l721refbpeZWhdPquXZtKtB/cRJuSIzwg6YGVuZ2J4cV8Iy0qptLwnTBJRnHdXw41zWJZBTvlnmCdHCZyIbsyyhzGbmx4zMKFM7A7aVtvl7jJND2lcSU1zcTUVCpCv5GJ3WcjFXsQrqeev7mBL82VtzGQCbmn09UlIndVJ2AoSvv16JExjVywCBYum+TIWuhQVsf4ALJ5ruE7FSMw3SFBcxKnSX0R355He5sz6sC3F8AJnZ5GukZCmW52ENA0TGKN18Y5jHQTqUFtKGn7zXcRa38wNJKvXwKlYc7YrzERw0tJG4lwoiulE2GNF3oRknLW0iM8d/U8yyyRnvAtZUsHKpmr8KeYoBwSnYq7LYSDZiptp+cb7Sy4w3j/P6etBRqMRivgWyaoq2iojyta3BBcVOhWkj5L4FAzJNL8JhaNc3ywcGGPGl96McS5N7pUX19LVfCnWHpmtligkkwTOdxQ5mVfwsrykI42jZLo+Kl3lwuG4v4td99VKa0lzywbib81i726Tl1jZalauMtcFTePlu8TbbwgStaaamq2mtm3v2OGBQOrO2+I3a2s033I3WJ+1unSuLGsb7ybo8K6Y+V1xXJ0Szqr7/5m4I4SLp8pVJiili+xyxsE0xy3tnuG2ndByw4bRdHsNx3aMRtNt242269pmzzWNbse6L5zCk9R0q7774j5DZvOXL6V84wVMujhmX4hoqtPyHKyXyuULGNOqvYCpzslgULRrAAvP3POsfmAHHa8R2O1+w+l2mo0g9DqNrhf63X43dJtB/74GTkqw07ZDx+s1G54Zhg3HMwr6zaDhO5bVdvx2s+e07899LUa++F+4t+S1/zdQSwMEFAAAAAgArp4MXeTaYLTfAAAAWQEAABMAAABkb2NQcm9wcy9jdXN0b20ueG1snZDNboMwEIRfBflOvXb4qREgATZSbz1UvSMwDVKMLbxJE1V99xq1Te+97WhGs99u+bxZpzdctI+u5rT6ihwRXUGpH4/aDP4h2GtwZruZAYPc3qid52XU0o5no1ekHCCj49mjNbG715G6/BG3aB2MrkhnVwz5l5vTTxOJZoPLVJEPmXZSppDGXIkuZsDaWBxEHsMjAG9514tGfZLI7WEeWi9YnNy7x+0buLjgf5knO+7n+9cdKQDDFcJ6gLRpWC5YxpImZyzJRJvnB9ULqaTqk1SV9M5Ql/T3yjD+PbP+AlBLAwQUAAAACACungxd0pCsajgKAADnPgAAGAAAAHhsL3dvcmtzaGVldHMvc2hlZXQxLnhtbKWb23ajuBKGX4WVB4jN2cxKZy2bnLuTzo6T9HTfKVhJ2APIG3DSM0+/hUxhGwkVMDfdtpX6rEP99UuyffLJ8r+Kd0pL43eaZMWXo/eyXP8xmRTRO01JcczWNOMtryxPScmf5m+TYp1TshJBaTKxplNvkpI4Ozo9Ea/d56cnbFMmcUbvc6PYpCnJ/17QhH1+OTKP4IWH+O29rF6YnJ6syRtd0vJpfZ/zZ5OGsopTmhUxy4ycvn45mpt/hM+mW0WIP3mO6Wex99goycuSJjQq6Uq81T+MpcuIJJQ/M6d7z++q0ST1q9UMvDD2V4W45oHTaiACU70z4f990JAm/M+/Vn9e/A86Uz1t+ltF7z+Gjl2IieMT8UIKGrLkR7wq3/k7c9KKvpJNUu69aB47rm1abtP2wD6vaD1T1rEYesSSQvxrfNZR9vHM9JypV8W90KK8iMXEGtGmKFkKbN7z8u9qKqzZkZHGmXgpJb/rNdgDWn7TjS6EXSOsGmG1EKZ77DpW4M18DQT6YdcQu92PadOPIcNyapwj9wkbFiDcGuHKCN/yXK8Pw6sZnjy7u6kZMi6/BvrtTs12nRrCm9W8WbuD9qhpD2pc0MZZx5YXjABWEtnm57SFtGdjM75JeSnnvWPTH7EoJijAbEuAZ++4dTZBD6YkCGcsEjRhtkXBF3vY6oDyTdCIKYnEG5aPDREUY7YlwzO8b0EyQSVmWyZ8PXYjRTEgDrOtDicYpQ4T5GFK+sCrLUAsUITVVoQV9C/ZoAFL0kAwYIaspvS3E98eMtEWJLvVTnYzwGttQ4H8tqSi3z91LMhoS8pop8eImkWCNLakyu/gpmibgIFEtqRE7k63XTTkr9XOX7PPIjdjgay1pKw1d/USm1gb0taW0nZk1bUhh20ph6cj3caGfLYVhXxIPWvWwG72NVIdtwcNe0eENLelMh6Mq7k25LzdzvnB9tUgIf9tqYz7g0pnAwQl2JISvHFAEIctbX2G7lUaJAjFbguFD3rUyjggGqctGnPoBrJBgmqctmp4lRyJBNU40glg6B6tQYJsHMkRvIGk5iSgOAqMHC/oxZE8YqQCHZCLI+96BpJAJ05bJ7Y5Mq0dUIojKWVseXBAKY5kKUN3t4B0QSquJJVhm7UGCEJxJXsxx/YRhOLK26WBJNCHK9nKyI23CzpxJVuZjUwctzlBS74yMKtd0IcrbafGVX8XZOJKMnFHzh+oxG2rxDaH7W8aIojEVYhkzKA9kIinODmMW2IPROIpzhHjSr8HIvEkkYzbOnigFU9xtzRy2CAWTzKV0cMGsXiSWPxx+eM1l0+SaMZagQey8aRdmD8WCbrx5CuogSTQiydtv9xxM+iDYHzJUwZetzVE0IuvOLOMSW4f1OJLJxZrIAlk4ksyGTt9oBJfspSxW2wfVOJLKhl7h+WDTPx/eVRprmibO1rVoX1Q10AbvnSAd8eOFkTit0Uycus6A43MFBqZzoKZYw892M9AJLO2SAbeFDR9BJEE7T6OLK0B9DCQDlFjLSVo+iideMZuiQPQXyCdU8zeF2LB7kLWngkdT7YfPokPts5ISU5PcvZp5CJafEbFa2MD3n7ytdfd5pMs/lbilo13rAqe1y/wt+OtcVZ9crgsc94c8zcrTx82pXGernNakJNJyXtRvTyJ6uhFHW12RIdsFb8x4zoraZ4x4zEnL+S/ZMVyBSvs0RMt4AwB3LH0Jad6xjnCmK9pksQrZtyTHkO66Eu77UO7RGivNHonRkYiRewVEntJM5oj73+NzS+JYpYRPiCy0pNuapLfQTovSh5phPFHnOhJX5E+naeEI+5pXlQ907O+9UjAG/pKjbM4p1GJzNYtQlvG2SqOCIa5w1Iop3zN4w+aGKYi/Dsi0L1wSxF+3z/cVoT/p3+4owh/QMKXm2iTF9iyLpEZDEn+hizCY68sC1m+Zjkp4w8E94SMS+RXpSVeZfIyjjYJyfXE597Eu5wdGyEpiB74owY6/YBndF3y/7TIP1GvSDcZ0q2fSLce6FvVJy3jF7KYj/Galx6WlXwpmSJ+XltmdXekBFyIOnydvXHTRDJhvujFuoi5m2v7hJnnLVvVhVlHOUMFRxNel6tv02g5mIuGlMcyY1VRCqwCzjEXrTcZNVQQVZjLnssW87TWju6qF+iR5nwjpyddI/O9JZ3RD5q9CXM1nkkkHtBCxbtBeCK160XU9esrwrnlHVjpE/IbgrjilbIweL00ljQlmR52i5UOkpUiuc9iDtWR7vrM0GVVxF/j7UyrKJir1vsXvm9OKH+gQmDO+kwSPjdoTzCH5SvF9XXL+EYq/qeTghntlhKypJOw7EWY82WK+AyrCI99VkYca6q6cU/elBTMWhcki7B6g5np3Sat9sphSY4rz89jXniQY8X8B8K8z+lHXKDeNcdMdI9TzZce9hOBXWdFGZcbUXj4nC9JslmpOL/QxQfzESsXlWpNLOaYzFlZJ7FB+d6Iu2uBEPHz6T7x6QLBhQjuirxwnZV8t7YHfmbJJuOvxcoDNGa5TantyTvvKeWeuAsE90Cr/OJ5Fr8Y87x0rP7oy17ZZ8xX3Eb7U696L9GSvm0qGdOicg8lrJ8/w06vBxAz6MOtQw8g5tRb4FNSxulBEvVAYw4+f133wWDevd2TbDlcfc+8TqgwmHHPiwpRVTrSg4YZ+D2r7LIHCLPxBc0o9++YpysvVjgP8/PvFSwjq+b4lZAXPuRExcJcvWbpEJit14h1zNNfg8G8vcaQjP7WcsDdLT0n2p4kNSDM4p8KmmckRVx9gbl6db8RbU8nehB+Nq4gEX5TsQBP77ykqzcv3XvkBWbn9XnrhvEpWpHzYk15fqumOcQMvUYYOgZm4Q9MGYYdiC/zjXIfGmKG/C3mm3vVVVuIWa+IVN2yhZjLikjVBVuImaiIVN2thZhRikhXFYm5ooj0VJGY/YlIXxWJ+ZyInKkiMRsTkYEqEnOubSZMVaHoMZP+Ltl8FUfKRMLcqYlWJhNmSWJLUEWLy3plBzAXOkQoewHm43bqT5yQV5uo/tRABVkikJDkOc3VanxEYvcPN003jOVmTblV58rL9PAJYYp9FjO2dbqEOq0iPfchEfUnb5OcfTYf/lnNB3nWljgTxOoHgbt36246624672666G66rJsCuemqO+q6u+mmu+lrd9Oyu+lXd9O8nkTbVLQtNG2aGZ5rZmR+pWn73s1c/NQs9kLTpu7nQULZRw3I1oCUbQcgZwdyNCBl2wHI3YFcDUjZdgDydiBPA1K2HYD8HcjXgJRtB6DZDjTTgJRtB6BgBwo0IGXbAcic7kjiF7ddKHXjIcvcY5k6lrLxkGXtsXQZrm48ZO3luKlLcnXjIWsvzU1dnqsbD1l7mW7qUl3duGVN9r4gUv22+5abUZwVRkJf+d9Oj3lS5tuvgojHJVuLR9V3XFhZshSevVOyonn1jE/CK2MlPNn9ZnyzNpi4/STVr7W/HK1Zzk0vLsW3VZrftp/+H1BLAwQUAAAACACungxdZI87ES4TAAAbZAAAGAAAAHhsL3dvcmtzaGVldHMvc2hlZXQyLnhtbMVdbW/jOJL+K7wsbnELdCexXu3pF0CxnW73JbZhJ1nsfjEYmUnUI0seSU53z6+/IimZsk0WFWCwNx8msuV6xHpIkQ+LRfbHH3nxe/nCWEV+btKs/HT2UlXb3y4uyviFbWh5nm9ZBnee8mJDK/hYPF+U24LRtTDapBfO5WVwsaFJdvb5o/huXnz+mO+qNMnYvCDlbrOhxa8rluY/Pp31zpovFsnzS8W/uPj8cUuf2ZJV99t5AZ8u9ijrZMOyMskzUrCnT2dR77dR3+EG4hcPCftRtq7Jn3m+WcY0ZQDrXZ6pz1Ne+LT+ljv8mOe/c5PJ+tPZ5RkvQMbIz+U2TaBIzhn5pS6rfHvDnqohSwFg6MA3NK6SVzYHk09nj3lV5RvhC3hW0Qq+eyryP1kmislSBj+G8m/FrwGr/qnEqEE5KX8ID4c9nZl8CC/FoWHk7g3h0mioe6TXV8+E6z3rnJT2dcPvtah+qM5HWrJhnv4zWVcvQCjwuWZPdJdWrS97557v9hx/f2+R//jK6voWBY3ztBT/Jz+kjQtA8a6E0jYgnM5foirPyCbJ5F/6s24xLVOnv3+cBcGpEZwjBM857/cC7zLogOHWGO5xKQZdS+HVCN4RQos2C4LfMBG4fYFyIfkUNTWiFf38sch/kEJYCNK98z3qviKg2cf8NxF/gGzqn86SjL+2y6qAuwlAVp9H0d2s9/GigqfwLy7i2uzKbuZozIYWs/FyuJjMh5O//83p9T5MNQgjHGH8bXw7v5kdGl4AH3tSHElKcB5qSCHVSxL/fpXXDB0YutLQ8cxcuqJsrqFsbAM9Z0lXxa7SMSqNPYPxYleRsQTQ8YobD/MCDLd5tmaEpoRjwVVKSV2kd6RMoNfYZVVeviMx9CDPuwTqwHM/ZOQXWcOlG354TqqcvLIieUpius4LXe3IcviGcoSBG/Q85/K9h1SQZ+fZE08JzDynMELlqzhfJ8+5jmsJEBrp4nZkklWsyHJyV9BH+t3g8dACJVkUeGuWtrDIroQ/ZAtXqiLOyTgjMS1z8seuXUHwY0bAHJoOVM8mJ8maZVVTEVB9jCTZGj4RXlXuh7Ji57rqkYXtGwrbQ2rFr2ulj/Ulfrd6MbwAPsokb7R4TeD22ncgJdUe8q97CXyUZefS7V86/cH7EKE72HMadOM0yzeP4J+O1wDlZSoNLdy2MbjUU84GqLMP0c14ejeZRiS6X8wWEeJyuHc57OYy3a62VLyiOq9D1OtoC0oogfdvTru85KHR/RB1P7p5iBbjfyNe9/de9zt7vTF73e/m9W0nr/tGr/uo11fRchbd3GGj8GDv9qCb208sXmU01vk8QH2+ZvELJVMaJ5sE+kybzzgY9M5yKpST0ej97e37f8F/2p5W4gz23L1+dgOvD6/8q4aN3qWSY5fd+HhmGSu0TaBGMLnwRRhaWLBg3EbL4f3NZDojObke346ncKllocYxNZTGFmkovZZS7XXsCqGu84xCQ6drLUE9vENsmdto6hnfkPqWyfPh1wl0jFh32HOU4043xxlMQbnsSV6TVOu4gzduYU6G3NzmuETqmYr0wAoyT2mWpKBhDnD5uDrbcoa1g9WoRjbxtsxT6Lny/6H/wKhzFXVuR+o2NElXW1aUvOq15Lk4eRyAzGsAG3041uwxTZ6hjykS0Ie0oCQqE1B0WZxQECJDWmxZRckoAS2iLeuoxj8l8ZQqT1GFq+qUPuYFTbl4W31nT0zLES6Gueb6BqZQ9ILF1q7YgnbCEjQ1ui3yR3h/G9kGWm+dx7sN7/j1zc2kiE+Z8hVTuM5tmCqlJK/0vTSuVpeNrY0j39wD4TJ0mpP6IWnyJ8CX2Aul5GgP16ON78+sEO1V6zouSL/UpjbPcRTodPISplKbLfxZs0eW8a6nhHlTQWHGBV9m+VrcSklePENX9VzQDcyERVMSg3zBZ1zQkhz/Q/GYp7wxbWi2zt8RAIO+loHhWl4WCTTncyAUphbPu4zP9cBYPgXuk12VSJa5KbRT+cQ/VUvlEw7eF1XwVN5Fwm+2FMiEsnwZT8eL6EY/vuLquzbFalap7h4uu/eteveIVi6uu5fK2la/EsgNdE3bpLpP/VP6uocL7MY/WjC9YxZpDWY2j/qIR7ii/hIt72bLpimMl1iNKmndw7V14/EuMyolXA3fZ1000gDxetC1Hh2lkB1cIat2Gu+KUj+WO7i8XdamFs9qFFfX+Tq48J2O/zmeTLAoZc8eBXNwPdwQEdNCHwRzcB085HY2CiwQ7ZgL79PWrIyLZNseoEES8ge9ExEvEceK090vJvrYlAzzrCr4MAgIsii5tiN0cLkd/et+FE3vxuRqNhp/wWS3o2S301V2C+0Y58UW+K6SVz3ZFu0t5ONQYdiIx+EE8Tmpm4CWMJPOPmVEqWmno5pec33HNf4KRtMqiXcp1Xlx5eA6eNTAkPkexsYLjgjTh5zIUFkdAQf5kzI+6pf0O/xlVaxvXyZFLXHn0TL6NiY3vI+eLP53hnXQjpLcTsdAtuIzK/KYapcArhxcLisypwVIlSGg2LjEAafwDvuXHzY8miC5rItZv9taGvEAtChZzx14fYw/JcSdjgHnA/7WbKuX5A4uyY8IHHGYcxuFOOQRhVAyaOdUzFaEyNyvBLCfMAGk+paJy/y6mP57tNNT+t7pGG8GubrL9O0QF+ZDYWjjrUvMunmDTSUZObg6vonIfDK9i9AwjKMEstMxLl2wZ2glWmpwcbwQhjZqwjeEXiQiTEBsYRcHj2Pfju8Ws/nsZmJjS8ltB5fbcT2wr6pkq38bcb19B2Z7daDlqf8Gng7Q7GThYn1+E/17Rq4n39BlaCXTHVym75l64pHsVZI9g6bSU9YlEj6R9rZW9lfFwZ3TQLjneqGjD4S7Sua7uMw/YuUpgdJoV9VxqS85uebWaGuywHTnowbqykcrtwBX+3s+NvnaHPp2cb1+29jiXOAgN7Qkef32kBJeJcAiX0GHFgmMastk/4Hs4CqT16WeLEsQfTYlX2eLaDHB3jTX6ZC4gUv9csdS6NjlH54DpeUWl+RLYSwyqHB2cZjbPKtkKoEEFDrh/vodGUnZBRKZle/IHF7xUsrZc3JDyQZqY01FzgHls6wyT5M4qSi3fkqgB3QH+hpoTxPq5hpcwn+G9qrmCi4+V9i31xjUTgESjYGkKPXCzMVF/VAgcDUw5AiWns0CNjqeoR6ha0nCJwfWyb6rZgQuPiNQrIkUlYY8M3OdEl1qF03+DW0orRyXLmx5J02qZ2hNSum7uNI/GSOhcesJwQV5M0Ryc/wtxXHeMB74bxoPlFB3caF+xEjFik2iTxlwca0tKbmT9jgnloh8d04CDSeBFxg4CTv077h0PyJrzV5Z9iyWgVev+pwDF1fykrRRg9P0JQ80pkZZObRg6vkj8lFrHjgTrVasX8Ry2SKNd2lOXvdP1bMdvqkFKoXv4gofRsu1lPf1wKnlsYPKr0dOtOnhMDXCXE63H2Hoy8ked14krzBY6LnBlb4GF+vmleR3cclf6w05aGtpw+X5rRztUcYGb5ge1Xi1u7bpkWuK6tfTIy5NsPxPNQfw8DlAyTY0o6sXkJDlapsXK/mFNuET1+9ckZYiF3MpIFDqDrAOw/7e5alc8vUvktdKZ8Wlfe1mTLOKa/PVOqHaPD/PFs+X9mQE9riDONC9TFzl/QyP3re7mjUP4QsC47wo4Gn6LNTeCUsmklQ03sMl+jMPm/OcTNFrm4IKHi6uRXfzRSGZI5lDC9IcOj/ORQFzy+8UOpsh5QxREyN41sveGHttlPT2uia9iMQcLU+WZBeZ0TOugbT84AgP5BN5SJ5BLbIPZAQfQHe/0MckTSpTX+x1TmrxlJ72pNh0TcweNptXmmonDVceroAfuF3XhuOZew88On43m4+J9y70sVagJLOHS+YcdHoJQ8yryEwQ7m/43FLrPq525Zz0tgWlddycquL5Jx2COzBOMj0lgj1cBEsf4zy1+YcrV+nfsIbR+mbOj/aC0yHBMfumws0erlmlb7xPj7nEMvuGq0rpW1TDaH2zZHQkPMelSljG3tWbEApytpxMSTS9mwwn89mZ9m3GA85te6y5KzXq4WpUSOctfTYHnL0OUlRshOOj3Jxq19SH3luizgdoVlnl4Vp0NJ7PlhM06dpTAtTDBahi65Fmx6K2pguXoFfczrZ1BoeoV3kkkpAba7aVPWuZ6GMLHi49h4vxCCgiYzKZPowXy8lsiubS+EqK+rgUVYRlu80qruhKCCA+vunI83E52l4XHFb0nCcGSDDbFhizNPXxjBTP9Qf9XniJsdHa6IMrVvq0XW0L9pqUhoUwH1eY88bW5q4lk1s2oOh6rmsrPh5cnt9E04fJHZo95yt56uPy9IAROSmuMNeufIu+VAzxfslC01uytY/wyP5J1J687eMyNopPs2MO+VQa1sc1bEnT3XqVZGWVVDvDqHzl4xp0oqx5J7zkkFr6cJi6lU2W0Xwx5hsiZtNoGWnpscSJ78cPEbmNlg+TEbbI6it56+PhYskSuiLk48pWrQiJYSo2TBiGNcwbmlgNJmm3t6wav2eak15jhClR7OOiWBIW51UjiVm22p7GK2rqbBsOGxS+yCKiHjYSLTnhPHmpTIiQejAgSv4A21TCkW/KDTmlSGlqH9fUOop2T1p+bIkgbX7ur23k4GhGco5bRs2MRHPbUc/eOc870qlyX6ly3xZJ5k695ukuq2iR0FUzqdV38ri0/trYkoarfSB5/wAtVTjspBbpS+ispnzXa2sTrBTyhCfJIQ/Vzs/9zknZvlLtvi1L5IhOc0jUx+X7PiT6JiJxzHqR9b+BrvtrPScmzX7KidLmvi0f5JQT/dTPt0aJK6yWtYy0IY8EZudk7kDJ6sCW5XHka8F4QwXtlDyuaFF5ukMmrgJcXS8EBmic5JFEHOONHARmkR2YRPYpB0pMB7bMjiMOpOpZ0fUm0cqeAFfFUvaQiJu/1XPzfsfApKZPPa/zMXoD7CiBwJKRwZ53RR6zkgdALL1sgEtp1csuBSgZClS53V9PAw7YtX/t+LhRYNkSiSifoDm0JESYtijtQ6atSWgBrpYPk9A6M26JR+uXRkfskRU1qojFsz92yStNxURa7Ed4UounskCHR0LIHzVnc2jrxn3L4mlQn23SC9CmbxH1ugpB1vcDXOEfru93rhActPtafw006NBvKBUfWFS8jqFdWiUb2hKuWq66pIfcCyTS0q+dWfurskaCzsI+aJ0kYhH2h6zxiIX8atV8p2UMl+XR0/aQHcPAgpwvgidxdwnTBHWGiNtDXztc2UsuXmEubIwgB7julskMkgzoch4SrZAdWlDERIfpTwnimedlVTCyy0i5fxAvM8nrDaZ8cwHdb5kud2WruzMkiASdhX3Q70S0JW9EEU3FNS+avu1ZtmQq6y60W1JI/vO0d547BINOtFtyThTt25yv42kZx2cUc2HYhWwc5/+B7M6Tl/CyC9mhLXNlT/Yj6MGnJE5EFrSO8xCfzVy17TtQb4H7z1Mfdp4zhc2BhuG5g3HfbV9szlnL6H5vmJZ7fCo1O4JANlwOLVCjg71z9damuojvyFpsod2CSKMZi8VRcOjywSjEFze+3UdTMhqT0WS2JLfRzfhqAV+4gYPlFoRqtSPEZ2dHFGuZxedQNbMNoTxJL4uTrYFax6geQsuKxMN4OhlFZL4YLyejMd8hvJiMl8MZ8cMedhxeqFYqQnz+dMTFNtFPm0LLkTDKGm1krpmJ9uKDnKkMLg2J16FaYQi7nQpTewet86fePctZLi1z1D9z8kzY+SSXUE0kwm4nudTemfc8WnBmbXvUvzbOkX/4Ps/5vWi9lmPXQjUbCDtu9NzBxDmjG+0ydoiL//va1Ha4nln8h8fJM6cedcj+DnFl3z75w3xkaIjL8mhvbPMWxzk8MJT8/W+9gfsBsHkcSYzDzTkJL/QX3R8as87lTth0t4EGNn8n4hi0iF/4SQIAw49Y0AcwwtDKcL8Dw93OVpEEGQ+kCHEJPlTWNo4tQIccD+X5E935Xb6F376V30EHfjtuEc12G8OpgKEl12Ynk02QrGALghCOMilYysVSMMkFy3eakY1MExZLoDtxFpPca9ukEW+ScpO3RI2ezIGNzL5aWuhLXdlzT/uVK+TeELk30t87LIGqtX4PKYH53hC5N9LfOyxBE2B3dTOUoyOu65LWGSKerqTme0Pk3kh/T5b0onVYOTS8Z3EwPd+lvsvqcu2/JfXR+MFv/MAdfty5Mvj88eUXiGJoir+3rxsTKEAK/QU/Ev/TmcxZFocQ/hc/9HydlNuU/vp01u2sQv7k4wfwqLp6QLPXHkHXbsfXIvfbyDxw1LyZCLp5E7vuEXzpWz1CLo/KDR5/2SO8Ay9EtiPPdETwzRtNdPi+c8wSHxai67nNgw7pTdrnecfP4/kKuchgsT3SnO4i2nS7HfN/D+MWBpQkK0nKnuCFuDyHllbIN1hcV/lWXMH7Lf+Vh+bTC2g2VvBP0EE85XnVfFD/zsZuS3KRTFj7sc0LqM6kEuXY/3sgn/8PUEsDBBQAAAAIAK6eDF2i5Is3NAIAALcGAAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDMueG1spZXfj9owDMf/lSpP20sLhQI6tZUYt2mTtgkduttzaF2ILo27JMDd/fVzUlrQfsAkHoDYcfz92KRuekD9bLYANnippTIZ21rb3EWRKbZQcxNiA4p2KtQ1t2TqTWQaDbz0h2oZxYPBJKq5UCxPvW+p8xR3VgoFSx2YXV1z/foBJB4yNmSd40FsttY5ojxt+AZWYB+bpSYr6rOUogZlBKpAQ5Wx+fBuPnPxPuBJwMGcrYM3xHpVcAmUNRmwk/3dscuj19W7Rnx2R76UGRswp68geF01UniiwGLzFSq7AEmn5jELeGHFHpYUlrE1Wou126daLLfkqjS+gfJkIIFiibj5I7hNckpqfrZVxawv2kGdr7v6PvnuUzfX3MAC5Q9R2i2hUj0lVHwn7ZlzGI6T0TBO+r0HPHyGY7vjMHFyBUrjv4NDeyqOw2k8SSbuWLEzBN2lY0EtVPvLX45/WNSe92z33PI81XgItI9yMvG4y9ILU5sLFzF3VyBjI9Iht1Dumqyspm1BGW3+kXpaYrAQeyHTyJKW80cFfUijF4qPQoPwRNx2yUP8Lu4kh0nPEHeOvzOsUFrQ+I6/v0AwuolgdJlgwQ114TLA+CaA8WWAe7FHXYirEMlNEMlliCexu6Y/uUl/cuUeQMP11RZMb0KYXrkIqPb0KICycAFhdhPC7L8R/v1YRmezwI3zb1xvhDKBpNFHQzakMnUr79c0Yf2KNNsR2VlberOAdhY9IRWi7Qw3dfp3Vf4LUEsDBBQAAAAIAK6eDF1icBIfZQMAALQNAAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDQueG1spVdtk9o4DP4rnnSm345AILBbXmaATq+d2b1jlrb3WQQRPHXs1DZl2V9f2YEsaXPcBT4sayvyIz2yLMujvdLfzBbRsudMSDMOttbm78LQJFvMwLRUjpK+bJTOwNJUp6HJNcLaL8pEGLXb/TADLoPJyMsWejJSOyu4xIVmZpdloA8zFGo/DjrBSfDE0611gnAyyiHFJdov+ULTLCxR1jxDabiSTONmHEw772adtlvgNb5y3JuzMXtRKlsmIJBg43bwOv/LOS+OUkd4pdQ3t+TTehy0A+eARHZY5oJ7l5hV+QNu7BwFrZpGAYPE8h+4ILVxsFLWqsx9JzIWLIk2Wr2g9J6hQNIll/PflAuQI2iPFn8vaPWCkrVz6nx84vfBh5/CuQKDcyX+4Wu7JVeJzxo3sBP2TNhp9eJuJ4rLb09q/xGP8Y5asTOXKGH8L9sXq3p3AVuhsR+OIUh2hhw/QQYs47L4D8/HXTtfPGgNon7cdzb/EyQ6gkSeduGIJ/keLExGWu2Z9trO36jdcqAFUEmCtixxSlOXT+OgSxok5tLl3NJq+swJ1E6eMKWtQDMKLZlysjChPzJR2olq7BTRPmPwatiZ68Sl/egk+Hf7b99Eve5Q0l6wz6Ahh4QkUTys+uThZo3gBJtq2CUgSdQdDOECye5NJLuNSE6lVRtIgY5GHcVmYEIZ9gCpurSDvZvI9ZqRs5BAVkusIdDBuG077eSfKFGDYHPQjvKnVZEjbkFniC9eZw5Zri7EIb4pDnEj9+fq+45nK1UXiGZIj5QqQlCtNOzABLj0AW15Amy+5QLlpbTu38S438jPryByOr7c1HJugPWIVqtc0VUDF8kNbiI3aFZJHvgKtYW10mUqzlBL0GvF/mYfeZpyaeqIN7Ljj/NTUa4uHum7m7jfNeP+SFck1nFrhEPF2CXtgS1AcwmJ+rX8VQje30TwvhnB2THif5wGdWQbYTIatu+Hu9WvgauQdN3J9Sxdq3fRpSWX7D1uuOT6khN1jcT/caKKcm2bUEW59h6uolx74VVRrr0uqijXluAqyrW1ropybdUoUMKzHjRDnfpG3bBE7eSx5SylZ88R38O+qhdvmUfQrl4yQW0/PTBaRE8XVv2YXhd+5Jpl/zw4zbb0rELtZpQoG6XsaeKMlA+1yU9QSwMEFAAAAAgArp4MXWNN1jRRAgAA3AYAABgAAAB4bC93b3Jrc2hlZXRzL3NoZWV0NS54bWyllUuP2jAQgP+KlVN7IRAI0BVEolSIlfpASx9nk0zAXceT2qZZ+PUdOxCQisKBQxLPwzPfTJzJpEL9anYAlr0VUplpsLO2fApDk+6g4KaDJSiy5KgLbknU29CUGnjmNxUyjLrdYVhwoYJk4nUrnUxwb6VQsNLM7IuC68NHkFhNg15wVryI7c46RZhMSr6FNdgf5UqTFDZRMlGAMgIV05BPg1nvaTZ2/t7hp4DKXK3ZEbFYp1wCRY27wUX+6tjlSevq3SC+ui3P2TToBi6/AnZYl1J4Imax/Ay5nYOkXbMoYDy14i+syG0abNBaLJydarHckirXeATlyUAC+RJx+Z9zHeQU1DXiz6mqoCnaQV2vz/UtfPepmxtuYI7yl8jsjlCpngxyvpf2StnrDOJ+L4ob2wtWSzi1O+rELl2K0vg7q+pdfTIMog/D8Yj2bcDYxakX6d5QBefYASuEqp/87fT2wjqYB/3ELU8mGiumvZfLGXU7DqYO1IBQ21Pn5DtB+cmD1EK5Y7O2msyCgtrkuyjRUCVsjspqbnESWsrpjGFKF+VqEkY3EtatuyrlQuDy9uIGJDorboM8K2qnUCJrI+g/RNBvJ5ixleRHZAvxu41h8BDDoJ1hiQo11wJNC0H8EEHcTnDpAeu1MAwfYhjeYUDNvm00ZwTCQfEWjtFDHKP7HEt6Hy35xw/lH985kcWGPs536z3IDNmBXU7H+1tI4dWMcDP/C9dboQyTNB9pEneoWl2D+DWNYb9yE8nP0bO0o98PaCfRF5Mj2rPgplHzQ0v+AVBLAwQUAAAACACungxdcf/qQCoCAACeBgAAGAAAAHhsL3dvcmtzaGVldHMvc2hlZXQ2LnhtbKWVbW/aMBCA/4rlH0AgEGBViMTaoU0aEyrr9tkkF2LVsTPbGW1//c7OS5GG0g/5APhen7vDucQXpZ9NAWDJSymk2dDC2uouCExaQMnMRFUg0ZIrXTKLoj4HptLAMh9UiiCcTpdBybikSex1B53EqraCSzhoYuqyZPr1Mwh12dAZ7RSP/FxYpwiSuGJnOIJ9qg4apaDPkvESpOFKEg35hm5nd9u18/cOvzhczNWZvClVHlMmALNGU/ou/3C1i1br+j0p9exCvmUbOqWOL4G8HivBfUXEquo75PYeBEZtQ0pYavlfOKDbhp6Utap0duzFMouqXKs3kL4yEIC+WHH1n3OTpE3qBvGn7Yr2Tbuirs9dfzs/fZzmiRm4V+I3z2yBpWI/GeSsFvZKOZssovksjHrbo7p8hXbc4SRyuFQJ47/JpYmao2ERflquVxh3AmN37SzS2mAHXW5KSi6bX/bS/ntBk8wX+sAsS2KtLkR7L8cMpxNXTJOoLwTHnjonPwnkowequXTX5mg1mjkmtcleSciYIccaRKbiwCLQWYIUPwjqaeENWjO3qz7e8Q46i/oqwk5xu4oDmCH2fBR7Psx+UILpAfhiFHwxDP9S66HGo1HsaJj9tBsgL0eRlx+Qf+4H0KtR6NUY9HoUev3RvMkeDNlKC5qrm1cuuHrK3dbeM33m0hCBGw536QSb0w3cn3GR+pPbKX4TdlKBLxDQTsLLnytlO8Htk/6VlPwDUEsDBBQAAAAIAK6eDF0aAGU8MwMAAPMNAAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDcueG1srVdtc6IwEP4rGX6ACIovN+qM2nbuZu7Fqdfe5wiLZAoJDVHb/vrbBEGuZdByfqglm+yzz7O7hGRyEPIpiwAUeUlink2tSKn0i21nfgQJzToiBY4zoZAJVTiUWztLJdDAOCWx7Xa7AzuhjFuzibGt5GwidipmHFaSZLskofJ1AbE4TC3HKgz3bBspbbBnk5RuYQ3qIV1JHNklSsAS4BkTnEgIp9bc+bJwjINZ8cjgkFWeyZsQydqnMSCs17VO45+afHy0asEbIZ60y7dganUtTYADeV2nMTOUiBLpdwjVEmL0mrsWob5ie1jhsqm1EUqJRM+jGEUVmkIp3oAbZhADrkXK6YfFOcgRVGfi+STLKoVrXtXnQuKdqQBmdEMzWIr4DwtUhGxRUgAh3cWqYnQ6fa/nuF45dy8OX+GYcrfj6XC+iDPzSw65Vw8n+u54MBqi3wYydXdMh7/LUESBbZGE8fw/fTlWsILijjtDd+ANdPBaR/fo6BrNOQuj8IYqOptIcSDSrNZk3W7nBFQqwJL5epHJIhLHFWhmXPfcWkmcZgiqZiZnmAKyolsxsRXG0xO2j38Ypwzm5sGcUREpT3ZF/Cm0Duh4JQO3MNQzuA11O+zfBTeuizOu843gggAnyx1wRckjw25rENFrL6LXzGQZwfMO6iRc5Ehu9TshBWd+UxH67fn3z9AwCVxJltA6ERd5k/uH3w3kvfbkvebwN5CKjKnaDjrjWjSfbiKIyYLyxgIM2msYNBN5xE14/7F9cxFnfH/JAOmff4eH7dkPmxmsQe4ZfqPqyJ9x1cLzFxc3VA4h8xmVrEnGqL2M0af2k3kkpKztqjM4FU2GUMCadqVxez3jT+lZqtod6gxIRQxL8FyjTxsNavTXtq0c7XshlVRI4mN5oLY+nwBCPjQQskmR8x+KnIuJYIUkbJELuTVpbmoZx6359l/C6l+U3lVQ+ldB8a6CMrgKyvAqKKOroIyvgeJ2r4JSd+K8HMWuHGETkFtzzs/wLd7x44m1tJJ3x/7T8vwq9IPKLeMZifHWgPeTDhZM5lHNM15OzJM+pJvbRTGK8FYGUo+w9UMhVDHQQcp73uwvUEsDBBQAAAAIAK6eDF1E1UZbQwIAAD4GAAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDgueG1srVVtb9owEP4rlidV2xcSAoGuJUisU7VJ6xSVbvtskgt4dXKZbZqWX7+zE1KmTawa+wDc+z332D5mDep7swGw7LFUlUn4xtr6IghMtoFSmAHWUJGnQF0KS6peB6bWIHKfVKogCsNJUApZ8fnM21I9n+HWKllBqpnZlqXQT+9AYZPwId8bbuV6Y50hmM9qsYYl2C91qkkL+iq5LKEyEiumoUj4YnixiF28D/gqoTEHMtshlstMKKCqccif9c8Ou+qsbt4V4r1L+ZgnPOSufwXsaVkr6RExi/UnKOwVKMpaRJyJzMoHSCks4Su0Fkvnp1mssGQqNO6g8shAAcUS4vq34LZIV9QR8aObivdDO1CH8n6+a88+sbkSBq5QfZO53RBUmieHQmyVPTAOB+N4NIzi3neLzQfo6I4GnsEMlfHfrGmzRuQYR28n51PKW4Gx1x0X2dbQBPvanJWyan/FY3d6QVvMA30vrJjPNDZM+yjXcxxSz65Oj4NYz1yMJ4L0KbGecFm5W7O0mtySatr5nayR3WmxEt9FjpqlGh6kuxJCzQJLzV1YkNGHmvado7ZzFA6ivnXL4cFMz1g8griHFO0Nf4a0cOeIR7qPTuo+Ot49bV8EkcFeZ2jlTrw5AmV8EpTxy6FUyP6OJj4JTfySY2Fna3vJJiE7exWNRyTFTDhxeInmCLLJPyL7pcj0fxQ5P6lIcPAK3Va9EXotK8MUbSDadQOCqNtUL9Oi85J7835T7bUNLXjQTqPrWCDaveLee/+XMf8JUEsDBBQAAAAIAK6eDF3cf/WmOQIAAKoGAAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDkueG1spVVtb5swEP4rlqV9mgSEhCStACnrVG3SOqFG3T47cASrBjPbCW1//c4modlL00hECvGd7+55nnN8xJ1Uj7oCMOSpFo1OaGVMe+37Oq+gZtqTLTS4U0pVM4Om2vq6VcAKl1QLPwyCuV8z3tA0dr5MpbHcGcEbyBTRu7pm6vkTCNkldEKPjnu+rYx1+Gncsi2swTy0mULLH6oUvIZGc9kQBWVCV5Pr1dLGu4AfHDp9siYvUtbrnAnAqlFAX+3vlrs4eK3ejZSPNuVrkdCAWvwGyPO6FdwxIka236A0NyAwaxVSwnLD95BhWEI30hhZ233UYphBV6nkCzSOGQjAWGTc/hPcFzkUtY34dVBFB9GW1On6qO/WdR+7uWEabqT4yQtTIVXUU0DJdsKcOCfeLJpOwmjYu5fdFzi0O/QiC5dLod2TdH1WuByy8p1GzsdqlNS86X/Z0+G8/D7dUfvMDEtjJTuiXJRFmQXeUGdAxj7nNsZJR3uBfU4ob+z/ZG0UbnOsadI7WTDBC1aQDDvGCokyyJqJXRH7BrFtlJ/jFzEH4LAHDgPvVUHfNEfqbyqOQDQwCnvH8g1GGWipz4BPR4FPz4M/3J5Bno1CnvWOK4dsr+8+DbxgEfzxmcT+/j/I0Sjk6LzmxQfykZwVPh8FP78IHh/vnfxiFI3Feyd/AYPlKAbLCxrxJgP/5P7bCX7H1JY3mgicdjhXPZSnely3xqHqVkizn4pHq8KXCShr4VUopTRHw06a4fWU/gZQSwMEFAAAAAgArp4MXcIVUBnPBAAAay0AAA0AAAB4bC9zdHlsZXMueG1s5Vr/b6M2FP9XEJX2w7QWCISGNcndLVKlSbupuusPJ21TRYJJPBnIjNMl99fPNoRA45fQlNycjaoC/Px57/O+GJuYYc42BH1eIMSMdULSfGQuGFv+aFn5bIGSML/JlijlkjijScj4LZ1b+ZKiMMoFKCFWz7Z9Kwlxao6H6Sq5T1huzLJVykamWzUZxennaGQ6vmcahbpJFqGRad/99glFf9impezdb/Z+Mr43rn64urJvbH48GXei4ffrvabv/lpl7O66OL17J1vfPxmAEb9pJIqsJLE2/Lh7LxBW6dh4GGfpzj+nZxYtXGeYIOM5JCPzA8UhEaA4TDDZFI090ZB/LW4c6apVADuGT0KCpxQrFUxLvLieZSSjBk4jtEY8AoO6geBE/afqhF2uM67c996sqnuib85NU4EcfCUqwWlGO8h4ZcHp0sLRUHZl6CxK94uCCcUjkz+iGBZD/Nq+8YIgGHi3nn3r9Xv+PplVcZ3jdE7QMXbyJJ4jmJDGc1I0jIfLkDFE03t+IzGycU9klNePmyXnOqfhxun1zdaAPCM4EibnE+kznU9H5r08bFl+01JQpdUvxlxNY2VLnrhD04xGiFYuBea2aTwkKGYcTvF8Ic4sWwobGWNZwi8iHM6zNJT+bhF1pCGnKB5JFOFVYu6VXMlNdC1ttETIvpJOSwDvueW9RbCFnPfU/YuuWnl4kO//0L8jBi7YQz1GlmYD3TprfTeUlQ9VWx4Q8SOII9nfBxwbwfuI81V4uyd9ecEnkBki5LMw8iVuTIzruLZSt8U6PZWXYsnOp56ytVBT3FgvQJ57CipoCbLqxAs3ah549gEXbFCvscTPGftpxSOXynvxGoMeKIrxWt6vYzXrmvbbs2ofdK/dq4U8ANQ7phEul2TzgeB5miD5Ftbe4ngYbnHGM6IMz8QajZepafxNw+UjWrNyNWitY9h3rck5UFV1x26RUfyVmxP8xIg3X823Tap7OkRTa3Jgqrtj9/ZUO96lEfY9nfMfnDmcvc7D2T8zY6djxi3m0HOl/1Q+7r/E58AKpnM+M36L6PGHoqPdkK2x884esHr9lyHbD+KrOJ8/yYfG7KG0B5dAUr/sH6PZB2jefrNZ+uiiUVOKLUaNPgyhGH6bIdOGIfTirWUMe5fFUJ8s23rGsP7ioivFYuNa61JscNQ0ji1e+D0dVo8Qub7O5HwdyF3IygtaNPQ7+c0W0u6frl3zH0su4scHq9wzqO2BNHZAqlZDfNUwMn8VXyqRXYCM6QoThtMqQy8BH7nTIUX5jnMd4qogDxkVJRv+iXa/C9VB/ebWB+fOwilBTfKcWoTicEXYYyUcmbvrj3IzqVf1ehARL3vtrn8Ru0mOX32jwW2V206T8pbOp7UvGWx7t7/1UlJ85aCWQJhCppYIGWQHYgBhChRk57/kzwD0p5BB3AZKyQDEDEBMgVJJJvIPsqPGBPxQexoEruv7UEQnEyWDCRQ33xf/am0QN4GA7AhLr4s1nG24Qg7XAZTTQxUCeQpXIuQpHGshUcdNIIJAnW3IjkBAWYBqR9hX2xE1pca4rsgqxA0awbAkCCCJqEV1jfo+EB1f/KnzA40S1w0CtUTI1AxcF5KI0QhLIAaCAyRx5eRpvZiPrO08Ze0+rx7/A1BLAwQUAAAACACungxdQknwHtAAAACmAgAACwAAAF9yZWxzLy5yZWxztZLLbgIxDEV/ZZQ9mEfFomJYsWGHED/gJp6HZhJHjhHD35OyKYPaUiF16eTm6OTK6wP1qC2H1LQxFYPvQypNoxrfAZJtyGOacqSQbyoWj5pHqSGi7bAmWMxmK5B7htms75nF8RLpL0SuqtbSlu3JU9BvwA8JUxxRatLSDD2cWboP5m6aoabYudLIzs0NvOjy8+/Ak6JDRbAsNImSX4u2lL50HNt9Pk63xEho8bLQ83JoUAqO3O9KGOPIaPmPRvaUlP2Tim6ZkdLbpxKM1nJzBVBLAwQUAAAACACungxdQbNoEQUCAABJBwAADwAAAHhsL3dvcmtib29rLnhtbLWVUW/aMBSF/0oW9XULpEABEaQJxlppW6OC2kdk4htyVceObAfa/vrdmGV1NSnaS57sc53c++lEOV6clX4+KPUcvJRCmiQsrK3mUWSyAkpmvqgKJJ3kSpfMktTHyFQaGDcFgC1FFA8Gk6hkKMPlou2V6mi5aDaPCGfzXm9kcEKDBxRoX5PQ7QWEQYkSS3wDnoSDMDCFOt8qjW9KWia2mVZCJOHwcvAI2mL2T3nb8OzYwbjKyxNKrs5JGE8nU+r42urPw5jk2akn5LagR2aD0d/aLeCxsNRjfNM8aNnhgVlUSTgZkMxRG+smOU6WWTwBDb2o2qoNCgt6zSx816quUB4bHHIj8uxw1rXrxfe5/h/nVZ5jBmuV1SVIe7Feg2gApSmwMmEgWQlJ6BCdMbTe8YtJlqg8y/Uc6UDfccfXH8saTKaxyqjuAcUdQHG/QN9oMFcrPKHwgK47gK77BXqAI5XAeDSjDppRvzQ7rNSKfjzNrPKIxh1E436JfpI5nG1rENwnmnQQTfol2jQvpezo49x04Nz0/8l2mh2+blIPaNoBNO0fKKV0VFsmau4xzTqYZi4n23DkkCN99l/U76P6M4KugX0qmEQh2J7D/t4lTPMTCdXcDu9Bs/QS6NPV6moYLyKv3/KDoln0dpbqoFlcdE7i2ZDSIa+FWFHtXv5QjLex3l5ty99QSwMEFAAAAAgArp4MXXsM+vndAAAAggYAABoAAAB4bC9fcmVscy93b3JrYm9vay54bWwucmVsc8XVzQ6CMAwH8FchewCLgPgR4OTFq/EFFiwfEdiy1qhvL8ED1njwYnZa2mX//k5ddsROc2sGalpLwb3vBspVw2x3AFQ22GtaGIvDeFMZ12seS1eD1eVF1whRGKbg3jNUkb1nBqeHxV8STVW1Je5Nee1x4C/BcDPuQg0iq+CkXY2cK7h3c5tgOpaLMVkFh3Ou3OG8VOAbFAlQ5B8UC1DsH5QIUOIftBKglX9QKkCpf9BagNb+QRsB2vgHbQVo+0cQ8aNDmjWvWq7B8I/zeXyL8/ipfDU/dvG0jEF8OcUTUEsDBBQAAAAIAK6eDF2fMBxmRQEAAIwIAAATAAAAW0NvbnRlbnRfVHlwZXNdLnhtbM2W3U7CQBCFX6XpLaGLqPgT4Ea9VS58gXU7pRu6P9kZEN7eaQskGmwgNXFvumln5nxnezZpp+87D5hsTWVxlpZE/lEIVCUYiZnzYLlSuGAk8W1YCi/VSi5BjEejiVDOElgaUq2RzqfPUMh1RcnLlh+jdnaWBqgwTZ7axpo1S6X3lVaSuC42Nv9BGe4JGU82PVhqjwNuSMVJQl35HbCfe9tACDqHZCEDvUrDXWJbCaRdBZh1S5zw6IpCK8idWhseydAHkDmWAGSqrBUddJOJ3zC016ve/EamC8idi+A8cmIBLscdIqmnh56FIJDu3uKRyNK99wd12jnkl7LVGsmZ3vhW5kw4Z/vpwqo5DCiapX/A3w/YUf9CH+NIfFxH4uMmEh+3kfiYROLjLhIf95H4ePhHHx/Orf76+1ivmZHaHvii+QmZfwFQSwECFAMUAAAACACungxdRsdNSJUAAADNAAAAEAAAAAAAAAAAAAAAgAEAAAAAZG9jUHJvcHMvYXBwLnhtbFBLAQIUAxQAAAAIAK6eDF3DAWyFHgEAAGYCAAARAAAAAAAAAAAAAACAAcMAAABkb2NQcm9wcy9jb3JlLnhtbFBLAQIUAxQAAAAIAK6eDF1YynR38gUAAI4aAAATAAAAAAAAAAAAAACAARACAAB4bC90aGVtZS90aGVtZTEueG1sUEsBAhQDFAAAAAgArp4MXeTaYLTfAAAAWQEAABMAAAAAAAAAAAAAAIABMwgAAGRvY1Byb3BzL2N1c3RvbS54bWxQSwECFAMUAAAACACungxd0pCsajgKAADnPgAAGAAAAAAAAAAAAAAAgIFDCQAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1sUEsBAhQDFAAAAAgArp4MXWSPOxEuEwAAG2QAABgAAAAAAAAAAAAAAICBsRMAAHhsL3dvcmtzaGVldHMvc2hlZXQyLnhtbFBLAQIUAxQAAAAIAK6eDF2i5Is3NAIAALcGAAAYAAAAAAAAAAAAAACAgRUnAAB4bC93b3Jrc2hlZXRzL3NoZWV0My54bWxQSwECFAMUAAAACACungxdYnASH2UDAAC0DQAAGAAAAAAAAAAAAAAAgIF/KQAAeGwvd29ya3NoZWV0cy9zaGVldDQueG1sUEsBAhQDFAAAAAgArp4MXWNN1jRRAgAA3AYAABgAAAAAAAAAAAAAAICBGi0AAHhsL3dvcmtzaGVldHMvc2hlZXQ1LnhtbFBLAQIUAxQAAAAIAK6eDF1x/+pAKgIAAJ4GAAAYAAAAAAAAAAAAAACAgaEvAAB4bC93b3Jrc2hlZXRzL3NoZWV0Ni54bWxQSwECFAMUAAAACACungxdGgBlPDMDAADzDQAAGAAAAAAAAAAAAAAAgIEBMgAAeGwvd29ya3NoZWV0cy9zaGVldDcueG1sUEsBAhQDFAAAAAgArp4MXUTVRltDAgAAPgYAABgAAAAAAAAAAAAAAICBajUAAHhsL3dvcmtzaGVldHMvc2hlZXQ4LnhtbFBLAQIUAxQAAAAIAK6eDF3cf/WmOQIAAKoGAAAYAAAAAAAAAAAAAACAgeM3AAB4bC93b3Jrc2hlZXRzL3NoZWV0OS54bWxQSwECFAMUAAAACACungxdwhVQGc8EAABrLQAADQAAAAAAAAAAAAAAgAFSOgAAeGwvc3R5bGVzLnhtbFBLAQIUAxQAAAAIAK6eDF1CSfAe0AAAAKYCAAALAAAAAAAAAAAAAACAAUw/AABfcmVscy8ucmVsc1BLAQIUAxQAAAAIAK6eDF1Bs2gRBQIAAEkHAAAPAAAAAAAAAAAAAACAAUVAAAB4bC93b3JrYm9vay54bWxQSwECFAMUAAAACACungxdewz6+d0AAACCBgAAGgAAAAAAAAAAAAAAgAF3QgAAeGwvX3JlbHMvd29ya2Jvb2sueG1sLnJlbHNQSwECFAMUAAAACACungxdnzAcZkUBAACMCAAAEwAAAAAAAAAAAAAAgAGMQwAAW0NvbnRlbnRfVHlwZXNdLnhtbFBLBQYAAAAAEgASAK8EAAACRQAAAAA="

# ----------------------------------------------------------------------------- #
#  Utilidades de normalización
# ----------------------------------------------------------------------------- #
def _norm(s):
    """minúsculas, sin acentos, sin espacios extremos -> para comparar llaves."""
    if s is None:
        return ""
    s = str(s).strip()
    s = "".join(c for c in unicodedata.normalize("NFD", s)
                if unicodedata.category(c) != "Mn")
    return s.lower()


def _clean(s):
    if s is None:
        return ""
    return str(s).strip()


def _to_date(v):
    """Devuelve datetime o None."""
    if v is None or v == "" or (isinstance(v, str) and v.strip() == ""):
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, (int, float)):
        # serial excel improbable aquí; ignorar
        return None
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(str(v).strip()[:10], fmt)
        except ValueError:
            continue
    return None


# ----------------------------------------------------------------------------- #
#  Catálogos de mapeo  (Rex -> BNOVUS)
#  Las llaves se comparan normalizadas (_norm), así toleran mayúsculas/acentos.
# ----------------------------------------------------------------------------- #
SEXO_MAP = {"m": "MASCULINO", "f": "FEMENINO"}

# Formato según archivo BNOVUS aceptado: sin "(a)", en título.
ESTCIVIL_MAP = {
    "s": "Soltero",
    "c": "Casado",
    "d": "Divorciado",
    "v": "Viudo",
    "u": "Conviviente Civil",   # 'U' = unión/conviviente civil (confirmar con negocio)
}

REGION_MAP = {
    _norm("Antofagasta"): "Región de Antofagasta",
    _norm("Metropolitana de Santiago"): "Región Metropolitana",
    _norm("Libertador General Bernardo O'Higgins"):
        "Región del Libertador General Bernardo O Higgins",
    _norm("Valparaíso"): "Región de Valparaiso",
    _norm("Coquimbo"): "Región de Coquimbo",
    _norm("Tarapacá"): "Región de Tarapacá",
    _norm("Arica y Parinacota"): "Región de Arica y Parinacota",
    _norm("Biobío"): "Región del Bío-Bío",
    _norm("Maule"): "Región del Maule",
    _norm("Ñuble"): "Región de Ñuble",
    _norm("De los Lagos"): "Región de los Lagos",
    _norm("Los Lagos"): "Región de los Lagos",
    _norm("Atacama"): "Región de Atacama",
    _norm("La Araucanía"): "Región del Araucanía",
    _norm("Araucanía"): "Región del Araucanía",
    _norm("Los Ríos"): "Región de los Ríos",
    _norm("Magallanes"): "Región de Magallanes y la Antartica Chilena",
    _norm("Aysén"): "Región de Aysén del General Carlos Ibáñez del Campo",
}

FORMAPAGO_MAP = {
    "actacorr": "Abono en CuentaCte",
    "actavis": "Abono en Cuenta Vista",
    "actaaho": "Abono en CuentaAhorro",
    "actarut": "Cuenta RUT",
    "efectivo": "Efectivo",
    "cheque": "Cheque",
    "sindefinir": "",
}

# Formato según archivo BNOVUS aceptado: en MAYÚSCULAS.
TIPOCONTRATO_MAP = {
    "i": "INDEFINIDO",
    "f": "A PLAZO FIJO",
    "o": "POR OBRA O FAENA",
    "h": "HONORARIOS",
}

# Jubilado? -> Prevision Tipo Trabajador (catálogo TipoTrabAFP)
PREVTIPO_MAP = {
    _norm("Activo (No Pensionado)"): "Activo",
    _norm("Pensionado y no cotiza"): "Pensionado (no cotiza)",
    _norm("Pensionado y cotiza"): "Pensionado (cotiza)",
    _norm("Activo > 60 ó 65 años"): "Activo > 60 ó 65 años",
}

ESTADO_MAP = {"a": "V", "p": "D"}   # Estado contrato Rex -> Estado Empleado BNOVUS

# AFP: se pasan a MAYÚSCULAS; casos con nombre distinto van acá
AFP_MAP = {
    "afp": "",   # genérico sin definir
}

ISAPRE_MAP = {
    "fonasa": "FONASA",
    "nuevamasvida": "NUEVA MASVIDA",
    "banmedica": "BANMEDICA",
    "colmena": "COLMENA",
    "consalud": "CONSALUD",
    "cruzblanca": "CRUZ BLANCA",
    "esencial": "ESENCIAL",
    "vidatres": "VIDA TRES",
    "fundacion": "FUNDACION",
}

# Banco (código Rex -> nombre banco BNOVUS). Best-effort: ajustar a catálogo BNOVUS.
BANCO_MAP = {
    "estado": "BANCO ESTADO",
    "falabella": "BANCO FALABELLA",
    "chile": "BANCO DE CHILE",
    "santander": "BANCO SANTANDER",
    "bci": "CREDITO E INVERSIONES",
    "scotia": "SCOTIABANK",
    "itau": "ITAU",
    "mercadopago": "MERCADO PAGO",
    "ripley": "BANCO RIPLEY",
    "losandes": "CAJA LOS ANDES",
    "security": "BANCO SECURITY",
    "tenpo": "TENPO",
    "copeuch": "COOPEUCH",
    "bice": "BANCO BICE",
    "edwards": "BANCO EDWARDS",
    "bbva": "BBVA",
    "corpbanca": "CORPBANCA",
}

# País (Rex) -> Nacionalidad (título), según archivo BNOVUS aceptado.
NACIONALIDAD_MAP = {
    "chile": "Chilena", "chilena": "Chilena",
    "peru": "Peruana", "peruana": "Peruana",
    "bolivia": "Boliviana", "boliviana": "Boliviana",
    "venezuela": "Venezolana", "venezolana": "Venezolana",
    "colombia": "Colombiana", "colombiana": "Colombiana",
    "argentina": "Argentina",
    "ecuador": "Ecuatoriana", "ecuatoriana": "Ecuatoriana",
    "haiti": "Haitiana", "haitiana": "Haitiana",
    "brasil": "Brasileña",
}

# Fecha término placeholder para contratos indefinidos (archivo aceptado usa 31-12-2030).
FEC_TERMINO_INDEFINIDO = datetime(2030, 12, 31)

# Índice de columnas BNOVUS por encabezado exacto de la fila 1 de la plantilla
BNOVUS_HEADERS = [
    "Rut Empresa", "Codigo Interno Trabajador", "Rut Trabajador", "Nombre Trabajador",
    "Apellido Paterno Trabajador", "Apellido Materno Trabajador", "fecha nac",
    "Genero Trabajador", "Nacionalidad Trabajador", "Estado Civil Trabajador",
    "Email Personal Trabajador", "Rut Jefe Directo Trabajador", "Sindicato Trabajador",
    "Area nivel 1", "Area nivel 2", "Area nivel 3", "Area nivel 4", "Sucursal Trabajador",
    "Cargo Trabajador", "Email Corporativo Trabajador", "Direccion Particular Trabajador",
    "Direccion Nro. Casa Trabajador", "Direccion Nro. Depto. Trabajador", "Comuna Trabajador",
    "Region Trabajador", "Tipo Contrato", "Fecha Ingreso Trabajador", "Fecha Firma Contrato",
    "Modalidad Contrato", "Sueldo base Contrato", "Centro de Costo Trabajador",
    "Codigo Centro Costo", "Fecha Inicio Contrato", "Fecha Termino Contrato",
    "Fecha Devengacion Vacaciones", "Tipo Sueldo Contrato", "Moneda Contrato",
    "Horas por Semana Contrato", "Cantidad Dias Contrato", "Tipo Gratificacion",
    "Estado Empleado", "Valor Gratificacion", "Monto Movilizacion", "Monto Colacion",
    "Monto Anticipo", "Tipo Forma de Pago", "Banco Trabajador",
    "Numero Cta. Corriente Trabajador", "Prevision Trabajador", "Prevision Tipo Trabajador",
    "Institucion de Salud", "Modalidad de Pactado", "Cotizacion en Pesos de Pactado",
    "Cotizacion en UF de Pactado", "Habilitar Cotizacion Voluntaria",
    "Moneda Cotizacion Voluntaria", "Monto Cotizacion Voluntaria",
    "Rebaja Trib Art42 Cotizacion Voluntaria", "Instit Admin Cotizacion Voluntaria",
    "Habilitar Seguro Cesantia", "Fecha Ingreso Seguro Cesantia",
    "Fecha Termino Seguro Cesantia", "Fecha Ultima Cotizacion Seguro Cesantia",
    "Afp Seguro Cesantia", "Tipo Seguro de Vida", "Aseguradora Seguro de Vida",
    "Poliza Seguro de Vida", "Beneficiarios Seguro de Vida", "Oficinadireccion laboral",
    "Oficina laboral", "Oficinapiso laboral", "Oficinaanexo laboral", "Oficinacomuna laboral",
    "Username Trabajador", "Areacodigo Trabajador", "Codigocargo Trabajador",
    "Numero Contrato", "Codigo JornadaEspecial", "Jornada Especial", "Rol", "Grupo",
    "Lista1", "Lista2", "Lista3", "Lista4", "Lista5", "Lista6", "Lista7", "Lista8",
    "Lista9", "Lista10", "TextoAdic1", "TextoAdic2", "FechaAdicional1", "FechaAdicional2",
    "Grado Educacional", "Carrera1", "Institucion Educacion Superior 1",
    "Ultimo cargo trabajado", "Ultima Empresa",
]
COL = {h: i for i, h in enumerate(BNOVUS_HEADERS)}   # nombre -> índice 0-based


# ----------------------------------------------------------------------------- #
#  Split de nombre "APELLIDO_P APELLIDO_M NOMBRES..."
# ----------------------------------------------------------------------------- #
def split_nombre(nombre):
    toks = _clean(nombre).split()
    if not toks:
        return "", "", ""
    if len(toks) == 1:
        return "", toks[0], ""           # solo un token -> paterno
    if len(toks) == 2:
        return "", toks[0], toks[1]      # paterno, materno (sin nombres)
    paterno, materno = toks[0], toks[1]
    nombres = " ".join(toks[2:])
    return nombres, paterno, materno


# ----------------------------------------------------------------------------- #
#  Transformación principal
# ----------------------------------------------------------------------------- #
def transformar(df, rut_empresa, incluir_todos, defaults, avisos):
    """
    df       : DataFrame del Listado de Empleados de Rex (columnas = header fila 2)
    devuelve : (lista_de_filas, dict_no_mapeados)
    """
    no_map = {"region": set(), "afp": set(), "isapre": set(), "banco": set(),
              "estado_civil": set(), "forma_pago": set(), "tipo_contrato": set(),
              "prev_tipo": set()}

    def g(row, *names):
        """primer valor no nulo de las columnas 'names' (tolerante a duplicados .1)."""
        for n in names:
            if n in row.index and pd.notna(row[n]):
                return row[n]
        return None

    filas = []
    for _, row in df.iterrows():
        rut = _clean(g(row, "Rut"))
        if not rut:
            continue

        estado_src = _norm(g(row, "Estado"))
        estado_bn = ESTADO_MAP.get(estado_src, "V")
        if not incluir_todos and estado_bn == "D":
            continue

        nombres, paterno, materno = split_nombre(g(row, "Nombre"))

        # --- códigos con catálogo ---
        sexo = SEXO_MAP.get(_norm(g(row, "Sexo")), _clean(g(row, "Sexo")).upper())

        ec_src = _norm(g(row, "Estado civil"))
        ec = ESTCIVIL_MAP.get(ec_src)
        if ec is None and ec_src:
            no_map["estado_civil"].add(_clean(g(row, "Estado civil")))
            ec = ""

        reg_src = _norm(g(row, "Región"))
        reg = REGION_MAP.get(reg_src)
        if reg is None and reg_src:
            no_map["region"].add(_clean(g(row, "Región")))
            reg = _clean(g(row, "Región"))

        fp_src = _norm(g(row, "Forma Pago"))
        fp = FORMAPAGO_MAP.get(fp_src)
        if fp is None and fp_src:
            no_map["forma_pago"].add(_clean(g(row, "Forma Pago")))
            fp = ""

        afp_src = _norm(g(row, "AFP"))
        afp = AFP_MAP.get(afp_src, _clean(g(row, "AFP")).upper())

        isa_src = _norm(g(row, "Isapre"))
        isa = ISAPRE_MAP.get(isa_src)
        if isa is None and isa_src:
            no_map["isapre"].add(_clean(g(row, "Isapre")))
            isa = _clean(g(row, "Isapre")).upper()

        banco_src = _norm(g(row, "Banco"))
        banco = BANCO_MAP.get(banco_src)
        if banco is None and banco_src:
            no_map["banco"].add(_clean(g(row, "Banco")))
            banco = _clean(g(row, "Banco")).upper()

        tc_src = _norm(g(row, "Tipo contr."))
        tc = TIPOCONTRATO_MAP.get(tc_src)
        if tc is None and tc_src:
            no_map["tipo_contrato"].add(_clean(g(row, "Tipo contr.")))
            tc = ""

        prev_src = _norm(g(row, "Jubilado?"))
        prev_tipo = PREVTIPO_MAP.get(prev_src, "Activo")

        nacion = NACIONALIDAD_MAP.get(_norm(g(row, "País")),
                                      _clean(g(row, "País")).title() or "Chilena")

        # --- salud: modalidad de pactado ---
        moneda_isa = _norm(g(row, "Moneda Isapre"))
        cot_uf = g(row, "Cotización UF")
        cot_pesos = g(row, "Cotización $")
        if moneda_isa in ("u.f.", "uf"):
            modalidad_pactado = "UF"
            salud_uf = cot_uf if (cot_uf not in (None, 0)) else ""
            salud_pesos = ""
        else:   # 7% (Fonasa u opción legal)
            modalidad_pactado = 0.07           # opción "7%" del catálogo TipoPactoSalud
            salud_uf = ""
            salud_pesos = ""

        # --- fechas ---
        fec_ini_contr = _to_date(g(row, "Fecha Inicio contrato"))
        fec_term_contr = _to_date(g(row, "Fecha término contrato"))
        # placeholder de "indefinido" (año >= 2999 en Rex) -> ignorar
        if fec_term_contr is not None and fec_term_contr.year >= 2999:
            fec_term_contr = None
        if tc == "INDEFINIDO":
            # archivo BNOVUS aceptado usa 31-12-2030 como término de indefinidos
            fec_term_bn = FEC_TERMINO_INDEFINIDO
        else:
            fec_term_bn = fec_term_contr
        fec_venc_vac = _to_date(g(row, "Fecha inicio vacaciones")) or fec_ini_contr
        fec_cesantia = _to_date(g(row, "Fecha inc. Seguro Cesa.")) or fec_ini_contr

        # --- seguro cesantía ---
        afecto_ces = g(row, "Afecto Seguro Cesantéa", "Afecto Seguro Cesantia")
        habilita_ces = "S" if bool(afecto_ces) else "N"

        # --- código interno: la columna "Código Interno" de Rex trae basura
        #     (valores como 'SI', 'CONTRATO 1', códigos de centro de costo),
        #     y el archivo BNOVUS aceptado la deja vacía -> siempre en blanco.
        cod_int = ""

        # --- sindicato ---
        sind_src = _norm(g(row, "Sindicato"))
        if sind_src in ("", "no tiene", "sindivpten", "no sindicalizado",
                        "no sindicalizados"):
            sindicato = "No Sindicalizados"
        else:
            sindicato = _clean(g(row, "Sindicato"))

        # ------------------------------------------------------------------ #
        #  Armado de la fila BNOVUS
        # ------------------------------------------------------------------ #
        fila = [None] * len(BNOVUS_HEADERS)

        def s(header, value):
            fila[COL[header]] = value

        s("Rut Empresa", rut_empresa)
        s("Codigo Interno Trabajador", cod_int)
        s("Rut Trabajador", rut)
        s("Nombre Trabajador", nombres)
        s("Apellido Paterno Trabajador", paterno)
        s("Apellido Materno Trabajador", materno)
        s("fecha nac", _to_date(g(row, "Fecha Nacimiento")))
        s("Genero Trabajador", sexo)
        s("Nacionalidad Trabajador", nacion)
        s("Estado Civil Trabajador", ec)
        s("Email Personal Trabajador",
          _clean(g(row, "Email Personal", "Correo electrónico")))
        s("Sindicato Trabajador", sindicato)
        # organigrama no viene en Rex -> GENERAL en nivel 1 (regla plantilla)
        s("Area nivel 1", defaults["area_nivel1"])
        area_src = _clean(g(row, "Área"))
        s("Area nivel 2", area_src)
        s("Sucursal Trabajador", _clean(g(row, "Sede")))
        s("Cargo Trabajador", _clean(g(row, "Cargo")))
        s("Email Corporativo Trabajador", "")
        s("Direccion Particular Trabajador", _clean(g(row, "Dirección")))
        s("Comuna Trabajador", _clean(g(row, "Comuna")))
        s("Region Trabajador", reg)
        s("Tipo Contrato", tc)
        s("Fecha Ingreso Trabajador", fec_ini_contr)
        s("Fecha Firma Contrato", fec_ini_contr)
        s("Modalidad Contrato", defaults["modalidad_contrato"])
        s("Sueldo base Contrato", g(row, "Sueldo Base"))
        s("Centro de Costo Trabajador", _clean(g(row, "Centro Costo")))
        s("Codigo Centro Costo", _clean(g(row, "Id Centro de Costo")))
        s("Fecha Inicio Contrato", fec_ini_contr)
        s("Fecha Termino Contrato", fec_term_bn)
        s("Fecha Devengacion Vacaciones", fec_venc_vac)
        s("Tipo Sueldo Contrato", defaults["tipo_sueldo"])
        s("Moneda Contrato", defaults["moneda_sueldo"])
        s("Horas por Semana Contrato", g(row, "Horas Semanales"))
        s("Cantidad Dias Contrato", defaults["cantidad_dias"])
        s("Tipo Gratificacion", defaults["tipo_gratificacion"])
        s("Estado Empleado", estado_bn)
        s("Valor Gratificacion", defaults["valor_gratificacion"])
        s("Monto Movilizacion", g(row, "Movilización"))
        s("Monto Colacion", g(row, "Colación"))
        s("Monto Anticipo", "SIN ANTICIPO")
        s("Tipo Forma de Pago", fp)
        s("Banco Trabajador", banco)
        s("Numero Cta. Corriente Trabajador", _clean(g(row, "Cuenta Banco")))
        s("Prevision Trabajador", afp)
        s("Prevision Tipo Trabajador", prev_tipo)
        s("Institucion de Salud", isa)
        s("Modalidad de Pactado", modalidad_pactado)
        s("Cotizacion en Pesos de Pactado", salud_pesos)
        s("Cotizacion en UF de Pactado", salud_uf)
        s("Habilitar Seguro Cesantia", habilita_ces)
        s("Fecha Ingreso Seguro Cesantia", fec_cesantia)
        s("Afp Seguro Cesantia", afp)
        s("Numero Contrato", 1)
        s("Rol", "empleados")
        s("Grupo", "Todos los empleados")

        filas.append(fila)

    return filas, no_map


DATE_HEADERS = {
    "fecha nac", "Fecha Ingreso Trabajador", "Fecha Firma Contrato",
    "Fecha Inicio Contrato", "Fecha Termino Contrato", "Fecha Devengacion Vacaciones",
    "Fecha Ingreso Seguro Cesantia", "Fecha Termino Seguro Cesantia",
    "Fecha Ultima Cotizacion Seguro Cesantia", "FechaAdicional1", "FechaAdicional2",
}


def construir_workbook(filas):
    """Escribe las filas en la plantilla BNOVUS embebida y devuelve bytes del .xlsx."""
    wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(PLANTILLA_BNOVUS_B64)))
    ws = wb["Sheet1"]
    # limpiar datos previos (por si la plantilla trae ejemplo)
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None

    date_cols = {COL[h] for h in DATE_HEADERS if h in COL}
    for i, fila in enumerate(filas, start=2):
        for j, val in enumerate(fila):
            cell = ws.cell(i, j + 1)
            cell.value = val
            if j in date_cols and isinstance(val, datetime):
                cell.number_format = "DD-MM-YYYY"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# ----------------------------------------------------------------------------- #
#  UI
# ----------------------------------------------------------------------------- #
def main():
    if BRANDING:
        hero("👥 Migración de empleados · Rex → BNOVUS",
             "Convierte el Listado de Empleados de Rex al archivo de carga de "
             "trabajadores de BNOVUS.")
    else:
        st.title("Migración de empleados · Rex → BNOVUS")
        st.caption("Convierte el Listado de Empleados de Rex al archivo de carga de "
                   "trabajadores de BNOVUS.")

    # --- Parámetros en el cuerpo principal (siempre visibles) ---
    c1, c2 = st.columns([2, 1])
    with c1:
        rut_empresa = st.text_input(
            "RUT Empresa *",
            help="Sin puntos, con guión y dígito verificador. Ej: 76361420-4",
            placeholder="76361420-4",
        ).strip()
    with c2:
        alcance = st.radio(
            "Alcance", ["Todos", "Solo activos"], horizontal=True,
            help="Solo activos excluye a los trabajadores con estado 'P'.",
        )

    with st.expander("Parámetros avanzados (valores por defecto)"):
        d1, d2, d3 = st.columns(3)
        with d1:
            moneda_sueldo = st.selectbox("Moneda Contrato",
                                         ["Peso", "UF", "Dolar", "Euro", "UTM"], index=0)
            tipo_sueldo = st.selectbox("Tipo Sueldo Contrato",
                                       ["Sueldo Privado", "Sueldo Público"], index=0)
        with d2:
            modalidad_contrato = st.selectbox(
                "Modalidad Contrato",
                ["Con Horario", "Sin Horario", "Honorarios"], index=0)
            cantidad_dias = st.number_input("Cantidad Dias Contrato", 1, 7, 5)
        with d3:
            tipo_gratificacion = st.selectbox("Tipo Gratificación",
                                              ["", "Calculada", "Fija"], index=1)
            valor_gratificacion = st.text_input("Valor Gratificación", value="TOPE 4,75")
        area_nivel1 = st.text_input("Area nivel 1 (organigrama)", value="GENERAL")

    archivo = st.file_uploader("Listado de Empleados de Rex (.xlsx)", type=["xlsx"])

    if st.button("Generar archivo BNOVUS", type="primary", disabled=archivo is None):
        if not rut_empresa:
            st.error("Ingresa el RUT Empresa antes de generar.")
            st.stop()
        # header en la fila 2 (skiprows=1). Título en la fila 1.
        df = pd.read_excel(archivo, sheet_name=0, skiprows=1)
        df = df[df["Rut"].notna()]

        defaults = dict(
            moneda_sueldo=moneda_sueldo, tipo_sueldo=tipo_sueldo,
            modalidad_contrato=modalidad_contrato, cantidad_dias=cantidad_dias,
            tipo_gratificacion=tipo_gratificacion,
            valor_gratificacion=valor_gratificacion, area_nivel1=area_nivel1,
        )
        avisos = []
        filas, no_map = transformar(df, rut_empresa,
                                    alcance == "Todos", defaults, avisos)

        data = construir_workbook(filas)

        st.success(f"Archivo generado: {len(filas)} trabajadores.")
        rut_slug = rut_empresa.replace("-", "").replace(".", "")
        st.download_button(
            "⬇️ Descargar archivo BNOVUS",
            data=data,
            file_name=f"bnovus_{rut_slug}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # informe de cobertura
        pend = {k: v for k, v in no_map.items() if v}
        if pend:
            st.warning("Valores no reconocidos en los catálogos (se copiaron tal cual, "
                       "revisar):")
            for k, v in pend.items():
                st.write(f"**{k}**: {', '.join(sorted(map(str, v)))}")
        else:
            st.info("Todos los valores codificados se mapearon correctamente.")

    if BRANDING:
        aplicar_footer()


if __name__ == "__main__" or True:
    main()
